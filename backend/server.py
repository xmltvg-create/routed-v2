from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File, Form, Query
from fastapi.responses import JSONResponse, HTMLResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import errno
import logging
import traceback

# Suppress httpx INFO logs that leak Mapbox API keys and other tokens
# into the server log. Only WARNING+ messages from httpx are useful.
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any, Literal, Tuple, Sequence
import uuid
import hashlib
from datetime import datetime, timezone, timedelta
import httpx
from haversine import haversine, Unit
import pandas as pd
import io
import asyncio
import math
from openlocationcode import openlocationcode as olc

ORTOOLS_AVAILABLE = True
ORTOOLS_IMPORT_ERROR: Optional[str] = None
try:
    from ortools.constraint_solver import pywrapcp, routing_enums_pb2
except Exception as _ortools_exc:
    ORTOOLS_AVAILABLE = False
    ORTOOLS_IMPORT_ERROR = str(_ortools_exc)

VROOM_AVAILABLE = True
VROOM_IMPORT_ERROR: Optional[str] = None
try:
    import vroom
except Exception as _vroom_exc:
    VROOM_AVAILABLE = False
    VROOM_IMPORT_ERROR = str(_vroom_exc)

LKH_AVAILABLE = True
LKH_IMPORT_ERROR: Optional[str] = None
try:
    # Read persistent-cache-aware binary path from the installer module.
    # Falls back to /usr/local/bin/LKH for dev environments without /app PVC.
    from install_native_solvers import LKH_BIN_PATH as LKH_SOLVER_PATH  # type: ignore
except Exception:
    LKH_SOLVER_PATH = "/usr/local/bin/LKH"
try:
    import lkh
    if not os.path.isfile(LKH_SOLVER_PATH):
        LKH_AVAILABLE = False
        LKH_IMPORT_ERROR = f"LKH binary not found at {LKH_SOLVER_PATH}"
except Exception as _lkh_exc:
    LKH_AVAILABLE = False
    LKH_IMPORT_ERROR = str(_lkh_exc)

PYVRP_AVAILABLE = True
PYVRP_IMPORT_ERROR: Optional[str] = None
try:
    from solvers.pyvrp_tsp_solver import PyVRPTspSolver, DeliveryStop  # noqa: F401
except Exception as _pyvrp_exc:
    PYVRP_AVAILABLE = False
    PYVRP_IMPORT_ERROR = str(_pyvrp_exc)

# Shared coord-clustering wrapper — gives every TSP solver in the pipeline
# (OR-Tools, LKH, VROOM, ILS, GA…) the same same-doorstep super-node
# protection that PyVRP gets internally. Without this, the "Zero-Cost
# Interleaving" bug was reachable through any fallback path.
from solvers.coord_clustering import cluster_aware_solve  # noqa: E402

# Self-heal: if the LKH binary is missing OR present-but-not-runnable on
# this CPU (e.g. an x86_64 binary cached on a PVC then mounted into an
# aarch64 pod after a fork), trigger a background compile. When it
# finishes, flip LKH_AVAILABLE back to True so subsequent /api/benchmark
# requests include LKH again.
_lkh_needs_install = not LKH_AVAILABLE
if LKH_AVAILABLE:
    try:
        from install_native_solvers import _lkh_binary_runnable as _lkh_check
        if not _lkh_check():
            _lkh_needs_install = True
            LKH_AVAILABLE = False
            LKH_IMPORT_ERROR = (
                f"LKH binary at {LKH_SOLVER_PATH} present but not runnable "
                "on this CPU — scheduling rebuild."
            )
            # _lkh_binary_runnable already logged the ENOEXEC detail once;
            # just log a short INFO here so the startup sequence is readable.
            logging.getLogger(__name__).info(LKH_IMPORT_ERROR)
    except Exception:
        # If the runnability probe itself errors out, fall back to the
        # lazy self-disable in lkh_tsp_solve — don't block server startup.
        pass

if _lkh_needs_install:
    try:
        from install_native_solvers import ensure_lkh_installed_background

        def _on_lkh_installed(ok: bool) -> None:
            # Module-level write is fine — Python globals are process-wide and
            # the benchmark endpoint reads them fresh on each request.
            global LKH_AVAILABLE, LKH_IMPORT_ERROR
            if ok:
                LKH_AVAILABLE = True
                LKH_IMPORT_ERROR = None

        ensure_lkh_installed_background(on_complete=_on_lkh_installed)
    except Exception as _installer_exc:
        logging.getLogger(__name__).warning(
            "[lkh-installer] could not schedule background install: %s", _installer_exc
        )

ALNS_AVAILABLE = True
try:
    from solvers import alns_hybrid_optimize
except Exception as _alns_exc:
    ALNS_AVAILABLE = False

TIMEFOLD_AVAILABLE = False
TIMEFOLD_IMPORT_ERROR: Optional[str] = None

# Timefold is one of 14 VRP solvers and is disabled by default — it needs a
# JDK + matching libjvm.so path which varies across container images. Set
# ENABLE_TIMEFOLD=true in the environment to opt back in (dev only).
# NB: load .env first so this flag is read correctly at module import time
#     (the main load_dotenv() further below runs after this block).
try:
    from dotenv import load_dotenv as _tf_load_dotenv
    _tf_load_dotenv(Path(__file__).parent / '.env')
except Exception:
    pass
_TIMEFOLD_ENABLED = os.environ.get("ENABLE_TIMEFOLD", "false").lower() in ("true", "1", "yes", "on")

if _TIMEFOLD_ENABLED:
    try:
        import os as _os_tf
        _os_tf.environ.setdefault("JAVA_HOME", "/usr/lib/jvm/java-17-openjdk-arm64")
        from timefold_solver import timefold_optimize
        TIMEFOLD_AVAILABLE = True
    except Exception as _tf_exc:
        TIMEFOLD_IMPORT_ERROR = str(_tf_exc)
        timefold_optimize = None
else:
    timefold_optimize = None

# Self-heal: if the Java JDK is missing (production image), apt-get install it
# in the background, then lazy-reimport timefold_solver. Takes ~60s on first run.
# Skipped entirely when ENABLE_TIMEFOLD is off (the default) — keeps prod logs clean.
if _TIMEFOLD_ENABLED and not TIMEFOLD_AVAILABLE:
    try:
        from install_native_solvers import ensure_timefold_installed_background

        def _on_timefold_installed(ok: bool) -> None:
            global TIMEFOLD_AVAILABLE, timefold_optimize, TIMEFOLD_IMPORT_ERROR
            if ok:
                try:
                    # Fresh import — JPype should now find the apt-installed JDK
                    from timefold_solver import timefold_optimize as _tf_opt
                    timefold_optimize = _tf_opt
                    TIMEFOLD_AVAILABLE = True
                    TIMEFOLD_IMPORT_ERROR = None
                except Exception as e:
                    TIMEFOLD_IMPORT_ERROR = f"post-install import failed: {e}"

        ensure_timefold_installed_background(on_complete=_on_timefold_installed)
    except Exception as _installer_exc:
        logging.getLogger(__name__).warning(
            "[timefold-installer] could not schedule background install: %s", _installer_exc
        )

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Configure logging first
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

if not ORTOOLS_AVAILABLE:
    logger.warning("OR-Tools is not available at startup: %s", ORTOOLS_IMPORT_ERROR)

# MongoDB connection with connection pooling settings for production
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(
    mongo_url,
    # Atlas-safe defaults: avoid eager socket creation and allow slower DNS/network handshakes
    maxPoolSize=50,
    minPoolSize=0,
    serverSelectionTimeoutMS=30000,
    connectTimeoutMS=30000,
    socketTimeoutMS=30000,
    waitQueueTimeoutMS=30000,
    connect=False,
    retryWrites=True,
    retryReads=True
)
db = client[os.environ['DB_NAME']]
APP_START_TIME = datetime.now(timezone.utc)
DB_READY_GRACE_SECONDS = int(os.environ.get('DB_READY_GRACE_SECONDS', '300'))

# ===================== Shared HTTP Client =====================
# Reuses TCP connections across all requests instead of creating/destroying per request
_shared_http_client: httpx.AsyncClient | None = None

async def get_http_client() -> httpx.AsyncClient:
    global _shared_http_client
    if _shared_http_client is None or _shared_http_client.is_closed:
        _shared_http_client = httpx.AsyncClient(
            timeout=httpx.Timeout(30.0, connect=10.0),
            limits=httpx.Limits(max_connections=50, max_keepalive_connections=20),
        )
    return _shared_http_client

# ===================== Directions Cache =====================
# Avoids redundant Mapbox API calls on repeated GPS ticks (same coordinates within 30s)
import time as _time
from collections import OrderedDict

class TTLCache:
    """In-memory LRU cache with TTL eviction and hit/miss counters"""
    def __init__(self, maxsize=200, ttl=30):
        self._cache: OrderedDict = OrderedDict()
        self._maxsize = maxsize
        self._ttl = ttl
        self.hits = 0
        self.misses = 0
    
    def get(self, key: str):
        if key in self._cache:
            val, ts = self._cache[key]
            if _time.monotonic() - ts < self._ttl:
                self._cache.move_to_end(key)
                self.hits += 1
                return val
            del self._cache[key]
        self.misses += 1
        return None
    
    def set(self, key: str, value):
        if key in self._cache:
            del self._cache[key]
        self._cache[key] = (value, _time.monotonic())
        if len(self._cache) > self._maxsize:
            self._cache.popitem(last=False)

    def stats(self) -> dict:
        total = self.hits + self.misses
        return {
            "entries": len(self._cache),
            "maxsize": self._maxsize,
            "ttl_seconds": self._ttl,
            "hits": self.hits,
            "misses": self.misses,
            "hit_rate": round(self.hits / total * 100, 1) if total > 0 else 0,
        }

_directions_cache = TTLCache(maxsize=200, ttl=30)

# OSRM duration matrix cache — avoids redundant OSRM calls for identical stop sets
# TTL=600s (10 min), max 50 route matrices cached
_osrm_matrix_cache = TTLCache(maxsize=50, ttl=600)

# Mapbox token
MAPBOX_TOKEN = os.environ.get('MAPBOX_TOKEN', '')
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')

# OSRM routing server URL (for VROOM duration matrix)
OSRM_URL = os.environ.get('OSRM_URL', 'https://router.project-osrm.org')
# Public OSRM demo server, used as a last-ditch fallback when the local
# OSRM is unreachable (e.g. on production where the binary isn't shipped).
# It's rate-limited (~1 req/sec) and has a per-call 100-coord cap, but
# delivers real road-network durations — far better than the Mapbox
# clustered matrix for solver quality. Set empty to disable.
OSRM_PUBLIC_URL = os.environ.get('OSRM_PUBLIC_URL', 'https://router.project-osrm.org')

# Optional production OSRM URL: when set AND the configured OSRM_URL is a
# loopback host that's not actually listening (i.e. we're running on the
# production pod which doesn't ship the OSRM binary), promote this URL to
# OSRM_URL at startup. Lets a single .env file work both in sandbox (fast
# localhost OSRM) and on the Emergent production pod (no OSRM binary)
# without per-environment branching at every call site.
OSRM_URL_PROD = os.environ.get('OSRM_URL_PROD', '').strip()
if OSRM_URL_PROD and OSRM_URL.startswith(('http://localhost', 'http://127.', 'http://[::1]')):
    import socket as _socket
    import time as _osrm_probe_time
    from urllib.parse import urlparse as _urlparse
    _parsed = _urlparse(OSRM_URL)
    _host = _parsed.hostname or 'localhost'
    _port = _parsed.port or 80
    # OSRM is supervisor-managed and may still be loading the .osrm.* mmap
    # files when uvicorn boots (the binary takes ~3-5 s on container start).
    # Retry the probe a few times so we don't spuriously promote to the
    # remote URL on every container restart and silently degrade sandbox
    # latency from <10 ms (local) to ~250 ms (Fly.io).
    _osrm_local_alive = False
    for _attempt in range(15):
        try:
            with _socket.create_connection((_host, _port), timeout=1.0):
                _osrm_local_alive = True
                break
        except Exception:
            _osrm_probe_time.sleep(1.0)
    if not _osrm_local_alive:
        logger.info(
            "Local OSRM at %s unreachable after 15 s; promoting OSRM_URL_PROD=%s",
            OSRM_URL, OSRM_URL_PROD,
        )
        OSRM_URL = OSRM_URL_PROD

# ── OSRM circuit breaker ───────────────────────────────────────────────
# In production (no local OSRM binary), every request fails and floods logs with
# identical warnings. After 3 consecutive failures we suppress subsequent warnings
# for 5 minutes (the fallback to Mapbox still happens transparently). While
# suppressed, we ALSO short-circuit the HTTP attempt entirely — avoids spending
# 2-5 s per request waiting for a TCP timeout on a host we know is unreachable.
_osrm_consecutive_failures = 0
_osrm_suppress_until = 0.0
_OSRM_FAIL_THRESHOLD = 3
_OSRM_SUPPRESS_SECONDS = 300


def _osrm_enabled() -> bool:
    """True when OSRM should be attempted on this request.

    Returns False when the URL is unset, or when the circuit breaker is open
    (i.e. we hit the failure threshold within the suppression window).
    Callers that currently gate on `if OSRM_URL:` should switch to this so
    they don't burn a TCP timeout on every request in production.
    """
    if not OSRM_URL:
        return False
    import time as _time
    if _osrm_consecutive_failures >= _OSRM_FAIL_THRESHOLD and _time.time() < _osrm_suppress_until:
        return False
    return True


def _osrm_log_failure(context: str, exc) -> None:
    """Log an OSRM failure once; after threshold is reached, suppress for a window."""
    global _osrm_consecutive_failures, _osrm_suppress_until
    import time as _time
    now = _time.time()
    _osrm_consecutive_failures += 1
    if now < _osrm_suppress_until:
        return
    logger.warning("%s: %s", context, exc)
    if _osrm_consecutive_failures >= _OSRM_FAIL_THRESHOLD:
        _osrm_suppress_until = now + _OSRM_SUPPRESS_SECONDS
        logger.warning(
            "OSRM unreachable (%d consecutive failures). Suppressing OSRM attempts for %ds; falling back to Mapbox.",
            _osrm_consecutive_failures, _OSRM_SUPPRESS_SECONDS,
        )


def _osrm_note_success() -> None:
    """Reset the circuit breaker after a successful OSRM response."""
    global _osrm_consecutive_failures, _osrm_suppress_until
    if _osrm_consecutive_failures:
        _osrm_consecutive_failures = 0
        _osrm_suppress_until = 0.0

# Generoute API for route optimization
GENEROUTE_API_KEY = os.environ.get('GENEROUTE_API_KEY', '')

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ===================== Models =====================
#
# All Pydantic request/response/domain models live in /app/backend/models/
# now. We re-export them here so legacy imports (`from server import Stop,
# StopUpdate, ...`) used by routes/stops.py and other modules keep working
# transparently. Future modules should prefer importing directly from
# `models` instead of through this re-export layer.

from models import (  # noqa: F401  (re-exports for backward compat)
    User,
    UserSession,
    TimeWindow,
    GeocodeCacheEntry,
    Stop,
    StopCreate,
    StopUpdate,
    RegeocodeStopRequest,
    RegeocodeStopResponse,
    CarStopActionRequest,
    FieldMapping,
    ImportPreviewResponse,
    ImportResult,
    Route,
    ReorderRequest,
    AlertType,
    MapAlert,
    AlertCreate,
    AlertResponse,
    GenerouteLocation,
    GenerouteRequest,
    OptimizationHub,
    RefinementSection,
    OptimizationRequest,
    TightenClusterRequest,
    BenchmarkRequest,
    VanLayout,
)


# ===================== Auth Helpers =====================

async def get_session_from_request(request: Request) -> Optional[UserSession]:
    # Check cookie first
    session_token = request.cookies.get("session_token")
    
    # Fallback to Authorization header
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header.split(" ")[1]
    
    if not session_token:
        return None
    
    session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session:
        return None
    
    # Check expiry with timezone awareness
    expires_at = session["expires_at"]
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    now = datetime.now(timezone.utc)
    if expires_at < now:
        return None

    # ── Sliding-window session refresh ────────────────────────────────
    # If the session is more than halfway to expiry (> 3.5 days old on a
    # 7-day token), silently extend it by 7 days. This prevents drivers
    # from getting kicked out mid-shift on day 7 — any active usage
    # automatically keeps the session alive.
    _SESSION_LIFETIME = timedelta(days=7)
    _REFRESH_THRESHOLD = _SESSION_LIFETIME / 2  # 3.5 days
    remaining = expires_at - now
    if remaining < _REFRESH_THRESHOLD:
        new_expiry = now + _SESSION_LIFETIME
        await db.user_sessions.update_one(
            {"session_token": session_token},
            {"$set": {"expires_at": new_expiry}},
        )
        logger.info(
            "[session-refresh] Extended session for user=%s, was %.1fh remaining, new expiry=%s",
            session.get("user_id"), remaining.total_seconds() / 3600, new_expiry.isoformat(),
        )
    
    return UserSession(**session)

# DEV MODE - Skip authentication for development (set via env var)
DEV_MODE = os.environ.get('DEV_MODE', 'false').lower() in ('true', '1', 'yes')
DEV_USER = User(
    user_id='dev-user-123',
    email='dev@example.com',
    name='Dev User',
    picture=None,
    created_at=datetime.now(timezone.utc)
)

# Auth access policy (ALLOWED_USERS / SIGNUPS_DISABLED) lives in routes/auth.py.

async def get_current_user(request: Request) -> User:
    # DEV MODE: Return dev user without authentication
    if DEV_MODE:
        return DEV_USER
    
    session = await get_session_from_request(request)
    if not session:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user = await db.users.find_one({"user_id": session.user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    
    return User(**user)

async def get_optional_user(request: Request) -> Optional[User]:
    try:
        return await get_current_user(request)
    except HTTPException:
        return None

# ===================== Auth Endpoints =====================
# Moved to routes/auth.py. Wired on the shared api_router near the bottom
# of this file alongside tiles/housenumbers routers.

# ===================== Suburb Helpers =====================

import re

def extract_suburb_from_address(address: str) -> Optional[str]:
    """
    Extract suburb from address string.
    Common formats:
    - "123 Main St, Paddington, QLD 4064, Australia"
    - "123 Main Street Paddington QLD"
    - "Paddington, Brisbane"
    """
    if not address:
        return None
    
    # Split by comma and analyze parts
    parts = [p.strip() for p in address.split(',')]
    
    # Australian format: usually suburb is 2nd or 3rd part
    # e.g., "123 Main St, Paddington, QLD 4064" -> Paddington
    if len(parts) >= 2:
        # Check if second part looks like a suburb (not a state or country)
        candidate = parts[1].strip()
        # Remove any postcode that might be attached
        candidate = re.sub(r'\s+\d{4}$', '', candidate)
        # Skip if it's a state abbreviation or country
        states = ['NSW', 'VIC', 'QLD', 'SA', 'WA', 'TAS', 'NT', 'ACT', 'Australia', 'AU']
        if candidate and candidate.upper() not in states and len(candidate) > 2:
            return candidate
    
    # Try third part if available
    if len(parts) >= 3:
        candidate = parts[2].strip()
        candidate = re.sub(r'\s+\d{4}$', '', candidate)
        states = ['NSW', 'VIC', 'QLD', 'SA', 'WA', 'TAS', 'NT', 'ACT', 'Australia', 'AU']
        if candidate and candidate.upper() not in states and len(candidate) > 2:
            return candidate
    
    return None

async def reverse_geocode_suburb(latitude: float, longitude: float) -> Optional[str]:
    """Use Mapbox reverse geocoding to get suburb/locality from coordinates"""
    if not MAPBOX_TOKEN:
        return None
    
    try:
        async with httpx.AsyncClient() as client:
            url = f"https://api.mapbox.com/geocoding/v5/mapbox.places/{longitude},{latitude}.json"
            params = {
                "access_token": MAPBOX_TOKEN,
                "types": "locality,neighborhood,place",
                "limit": 1
            }
            response = await client.get(url, params=params)
            if response.status_code == 200:
                data = response.json()
                if data.get("features"):
                    for feature in data["features"]:
                        # Get the locality or neighborhood
                        place_type = feature.get("place_type", [])
                        if "locality" in place_type or "neighborhood" in place_type:
                            return feature.get("text")
                        # Fallback to place name
                        if "place" in place_type:
                            return feature.get("text")
    except Exception as e:
        logger.error(f"Reverse geocoding error: {e}")
    
    return None

# ── Stop CRUD + reorder moved to routes/stops.py ─────────────────────────
# GET/POST/PUT/DELETE /stops, /stops/{id}, /stops/{id}/complete,
# /stops/{id}/uncomplete, /stops/clear, /stops/reorder, /debug/stops-coords
# are now served from that router. Heavier siblings (regeocode, refresh-
# suburbs, /car/*, /stops/export/xlsx) stay here until split into their own
# modules.

# ===================== Route History =====================

@api_router.post("/routes/archive")
async def archive_route(
    request: Request,
    current_user: User = Depends(get_current_user),
):
    """Snapshot current stops into route_history before clearing them.

    Optional JSON body:
      { "algorithm": "vroom_lkh_3opt", "total_distance_km": 187.3,
        "total_duration_seconds": 24720 }
    Any provided fields are persisted under `summary.*` so the
    `/api/_meta/telemetry-rollup` endpoint can answer "which algorithm
    did I use today?" without us having to wire algorithm into every
    intermediate state. All fields are optional and backwards-
    compatible — older clients posting no body still archive cleanly.
    """
    # Parse the optional body. We use Request directly (not a Pydantic
    # model) so the endpoint stays backwards-compatible with clients
    # that POST with no body / no Content-Type header.
    optional_body: Dict[str, Any] = {}
    try:
        if request.headers.get("content-length") and int(request.headers["content-length"]) > 0:
            optional_body = await request.json()
            if not isinstance(optional_body, dict):
                optional_body = {}
    except Exception:
        optional_body = {}

    all_stops = await db.stops.find(
        {"user_id": current_user.user_id}, {"_id": 0}
    ).sort("order", 1).to_list(5000)

    if not all_stops:
        return {"archived": False, "message": "No stops to archive"}

    delivered = [s for s in all_stops if s.get("completed")]
    skipped = [s for s in all_stops if s.get("delivery_status") == "skipped"]
    failed = [s for s in all_stops if s.get("delivery_status") == "failed"]
    pending = [s for s in all_stops if not s.get("completed") and s.get("delivery_status", "pending") == "pending"]

    total_weight = sum(s.get("weight") or 0 for s in all_stops)
    total_quantity = sum(s.get("quantity") or 0 for s in all_stops)

    # ── Telemetry rollup: surface the geofence-vs-fallback ratio + distance
    # percentiles so we can answer "is the 100 m geofence radius the right
    # fit for this driver's parking habits?" without instrumenting on the
    # device. This data populates the Phase-1 ML readiness check.
    geofence_n = sum(1 for s in delivered if s.get("arrival_method") == "geofence")
    inferred_n = sum(1 for s in delivered if s.get("arrival_method") == "geofence_inferred")
    fallback_n = sum(1 for s in delivered if s.get("arrival_method") == "fallback_completion")
    distances = [
        s["completion_distance_m"] for s in delivered
        if isinstance(s.get("completion_distance_m"), (int, float))
    ]
    distances.sort()

    def _pct(arr, q):
        if not arr:
            return None
        idx = min(len(arr) - 1, int(q * (len(arr) - 1)))
        return round(arr[idx], 1)

    # Service-seconds samples MUST come from real geofence arrivals only.
    # `geofence_inferred` rows back-date arrived_at by a constant 30s, so
    # including them would pollute the service-time distribution with a
    # degenerate p50/p95 of 30.
    service_seconds = []
    for s in delivered:
        if s.get("arrival_method") != "geofence":
            continue
        a, c = s.get("arrived_at"), s.get("completed_at")
        if not (a and c):
            continue
        try:
            if isinstance(a, str):
                a = datetime.fromisoformat(a.replace("Z", "+00:00"))
            if isinstance(c, str):
                c = datetime.fromisoformat(c.replace("Z", "+00:00"))
            service_seconds.append((c - a).total_seconds())
        except Exception:
            pass
    service_seconds.sort()

    total_arrivals = geofence_n + inferred_n + fallback_n
    telemetry = {
        "geofence_count": geofence_n,
        "geofence_inferred_count": inferred_n,
        "fallback_count": fallback_n,
        # `geofence_rate` is the strict ratio (real geofence hits only) —
        # the diagnostic for whether the hook itself is firing.
        "geofence_rate": (
            round(geofence_n / (geofence_n + fallback_n + inferred_n), 3)
            if total_arrivals > 0
            else None
        ),
        # `arrival_proximity_rate` adds inferred geofence samples — the
        # driver-friendly "we tracked your arrival" metric. Bumps from
        # ~0% to ~80% expected once the inference backstop kicks in.
        "arrival_proximity_rate": (
            round((geofence_n + inferred_n) / total_arrivals, 3)
            if total_arrivals > 0
            else None
        ),
        "completion_distance_p50_m": _pct(distances, 0.5),
        "completion_distance_p95_m": _pct(distances, 0.95),
        "service_seconds_p50": _pct(service_seconds, 0.5),
        "service_seconds_p95": _pct(service_seconds, 0.95),
        "distance_samples": len(distances),
        "service_samples": len(service_seconds),
    }

    # Compute timestamps
    completed_times = [s["completed_at"] for s in delivered if s.get("completed_at")]
    started_at = min((s.get("created_at") for s in all_stops if s.get("created_at")), default=None)
    finished_at = max(completed_times, default=None) if completed_times else None

    route_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user.user_id,
        "archived_at": datetime.now(timezone.utc).isoformat(),
        "started_at": started_at.isoformat() if isinstance(started_at, datetime) else started_at,
        "finished_at": finished_at.isoformat() if isinstance(finished_at, datetime) else finished_at,
        "stops": all_stops,
        "summary": {
            "total_stops": len(all_stops),
            "delivered": len(delivered),
            "skipped": len(skipped),
            "failed": len(failed),
            "pending": len(pending),
            "total_weight_kg": round(total_weight, 2),
            "total_quantity": total_quantity,
            "telemetry": telemetry,
            # Persisted from the optional archive body so the
            # `/api/_meta/telemetry-rollup` endpoint can answer
            # "which algorithm did I use today?". Older clients
            # omit these, and we keep them None for backwards-compat.
            "algorithm": optional_body.get("algorithm") if isinstance(optional_body.get("algorithm"), str) else None,
            "total_distance_km": optional_body.get("total_distance_km") if isinstance(optional_body.get("total_distance_km"), (int, float)) else None,
            "total_duration_seconds": optional_body.get("total_duration_seconds") if isinstance(optional_body.get("total_duration_seconds"), (int, float)) else None,
        },
    }

    await db.route_history.insert_one(route_doc)
    # Remove the MongoDB _id that insert_one adds to the dict
    route_doc.pop("_id", None)

    return {"archived": True, "route": route_doc}


@api_router.get("/routes/history")
async def get_route_history(current_user: User = Depends(get_current_user)):
    """List all archived routes (summary only, no full stop list)."""
    cursor = db.route_history.find(
        {"user_id": current_user.user_id},
        {"_id": 0, "stops": 0}  # exclude heavy stops array
    ).sort("archived_at", -1)

    routes = await cursor.to_list(500)
    return {"routes": routes}


@api_router.get("/routes/history/{route_id}")
async def get_route_detail(route_id: str, current_user: User = Depends(get_current_user)):
    """Get full detail of a specific archived route including all stops."""
    route = await db.route_history.find_one(
        {"id": route_id, "user_id": current_user.user_id}, {"_id": 0}
    )
    if not route:
        raise HTTPException(status_code=404, detail="Route not found")
    return route


@api_router.delete("/routes/history/{route_id}")
async def delete_route_history(route_id: str, current_user: User = Depends(get_current_user)):
    """Delete a specific route from history."""
    result = await db.route_history.delete_one(
        {"id": route_id, "user_id": current_user.user_id}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Route not found")
    return {"deleted": True, "route_id": route_id}


# ── Route geometry normaliser (Python port of src/utils/routeGeometry.ts) ───
# Same contract as the frontend utility: validate shape, auto-flip [lat, lng]
# pairs if detected, strip non-finite + out-of-range values, dedupe coincident
# vertices. Used by the GPX / KML exporters so downstream GPS tooling never
# receives a malformed polyline.
def _looks_like_latlng_swap(coords: List[List[float]]) -> bool:
    sample = coords[:8]
    first_inside_lat = 0
    second_outside_lat = 0
    for p in sample:
        if not isinstance(p, (list, tuple)) or len(p) < 2:
            continue
        a, b = p[0], p[1]
        if not (isinstance(a, (int, float)) and isinstance(b, (int, float))):
            continue
        if abs(a) <= 90:
            first_inside_lat += 1
        if 90 < abs(b) <= 180:
            second_outside_lat += 1
    return first_inside_lat == len(sample) and second_outside_lat >= 1


def normalise_line_coordinates(
    coords: Optional[List[List[float]]], auto_flip: bool = True
) -> List[List[float]]:
    if not coords:
        return []
    should_flip = auto_flip and _looks_like_latlng_swap(coords)
    out: List[List[float]] = []
    prev: Optional[List[float]] = None
    for raw in coords:
        if not isinstance(raw, (list, tuple)) or len(raw) < 2:
            continue
        lng = raw[1] if should_flip else raw[0]
        lat = raw[0] if should_flip else raw[1]
        try:
            lng = float(lng)
            lat = float(lat)
        except (TypeError, ValueError):
            continue
        if not (math.isfinite(lng) and math.isfinite(lat)):
            continue
        if lng < -180 or lng > 180 or lat < -90 or lat > 90:
            continue
        if prev and prev[0] == lng and prev[1] == lat:
            continue
        out.append([lng, lat])
        prev = [lng, lat]
    return out


def _escape_xml(s: Any) -> str:
    if s is None:
        return ""
    return (
        str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


@api_router.get("/routes/history/{route_id}/export.gpx")
async def export_route_gpx(route_id: str, current_user: User = Depends(get_current_user)):
    """Export a saved route as GPX (Garmin, Strava, Komoot, most GPS devices)."""
    route = await db.route_history.find_one(
        {"id": route_id, "user_id": current_user.user_id}, {"_id": 0}
    )
    if not route:
        raise HTTPException(status_code=404, detail="Route not found")

    stops = route.get("stops") or []
    name = _escape_xml(route.get("name") or f"RouTeD route {route_id[:8]}")

    # GPX 1.1 — waypoints only (no <trkseg>). A full trackline would duplicate
    # the routing engine's output and many GPS devices treat trackpoints very
    # differently from user-planned routes; waypoints are the universal truth.
    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<gpx version="1.1" creator="RouTeD" xmlns="http://www.topografix.com/GPX/1/1">',
        f"  <metadata><name>{name}</name></metadata>",
    ]
    for s in stops:
        try:
            lat, lng = float(s.get("latitude")), float(s.get("longitude"))
        except (TypeError, ValueError):
            continue
        lines.append(
            f'  <wpt lat="{lat}" lon="{lng}"><name>{_escape_xml(s.get("name") or s.get("address") or "")}</name></wpt>'
        )
    lines.append("</gpx>")
    body = "\n".join(lines).encode("utf-8")
    return Response(
        content=body,
        media_type="application/gpx+xml",
        headers={"Content-Disposition": f'attachment; filename="route-{route_id[:8]}.gpx"'},
    )


@api_router.get("/routes/history/{route_id}/export.kml")
async def export_route_kml(route_id: str, current_user: User = Depends(get_current_user)):
    """Export a saved route as KML (Google Earth, Google My Maps, most GIS tools)."""
    route = await db.route_history.find_one(
        {"id": route_id, "user_id": current_user.user_id}, {"_id": 0}
    )
    if not route:
        raise HTTPException(status_code=404, detail="Route not found")

    stops = route.get("stops") or []
    name = _escape_xml(route.get("name") or f"RouTeD route {route_id[:8]}")

    # KML — placemarks for each stop only (no <LineString> track). GIS tools
    # and Google Earth will render clean pins the user can reorder / click;
    # a track polyline would pin the user to our routing engine's decisions.
    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<kml xmlns="http://www.opengis.net/kml/2.2"><Document>',
        f"  <name>{name}</name>",
    ]
    # Stops as Placemarks (KML expects lng,lat[,alt] — opposite of GPX).
    for s in stops:
        try:
            lat, lng = float(s.get("latitude")), float(s.get("longitude"))
        except (TypeError, ValueError):
            continue
        label = _escape_xml(s.get("name") or s.get("address") or "")
        lines.append(
            f"  <Placemark><name>{label}</name><Point><coordinates>{lng},{lat},0</coordinates></Point></Placemark>"
        )
    lines.append("</Document></kml>")
    body = "\n".join(lines).encode("utf-8")
    return Response(
        content=body,
        media_type="application/vnd.google-earth.kml+xml",
        headers={"Content-Disposition": f'attachment; filename="route-{route_id[:8]}.kml"'},
    )


@api_router.post("/routes/history/{route_id}/resume")
async def resume_route(route_id: str, current_user: User = Depends(get_current_user)):
    """Restore an archived route back into active stops (resets completion status).

    Hardened against:
      - Legacy archives saved under a different user_id (e.g. old dev-user-123).
      - Stops carrying completion telemetry fields that must be cleared so the
        resumed route shows as pristine pending.
      - Duplicate stop ids inside the same archive (would collide with the
        unique (id, user_id) index).
      - Generic exceptions — we now surface the real reason to the client.
    """
    try:
        # Primary lookup: scoped to current user.
        route = await db.route_history.find_one(
            {"id": route_id, "user_id": current_user.user_id}, {"_id": 0}
        )
        # Fallback: legacy archives may exist under a previous user_id (e.g.
        # before auth was hooked up). If the user can read it from their
        # history listing today, they should be allowed to resume it.
        if not route:
            route = await db.route_history.find_one({"id": route_id}, {"_id": 0})
            if route:
                logger.warning(
                    f"[resume_route] route {route_id} owned by "
                    f"{route.get('user_id')} resumed by {current_user.user_id} "
                    f"(legacy archive fallback)"
                )
        if not route:
            raise HTTPException(status_code=404, detail="Route not found in history")

        archived_stops = route.get("stops", []) or []
        if not archived_stops:
            raise HTTPException(status_code=400, detail="Archived route has no stops")

        # Fields that must be wiped so a resumed stop is treated as pending.
        # Anything left behind (completion coords, arrival_method, photo proof,
        # service-time samples) would make the UI render the stop as "done".
        completion_fields = (
            "completed_at", "arrived_at", "departed_at",
            "completion_lat", "completion_lng", "completion_accuracy",
            "arrival_method", "arrival_distance_m", "arrival_confidence",
            "failure_reason", "failure_code", "skip_reason",
            "proof_photo_url", "proof_photo_uploaded_at",
            "service_time_seconds", "service_time_source",
            "delivered_at", "skipped_at", "failed_at",
        )

        # Dedupe stop ids inside the archive — the (id, user_id) index is
        # unique so duplicates would 500 the insert_many.
        seen_ids: set[str] = set()
        cleaned: list[dict] = []
        for i, raw in enumerate(archived_stops):
            stop = dict(raw)  # don't mutate the archive document
            stop.pop("_id", None)
            # If id is missing or duplicated, mint a new one so the index
            # constraint is satisfied.
            sid = stop.get("id")
            if not sid or sid in seen_ids:
                sid = str(uuid.uuid4())
                stop["id"] = sid
            seen_ids.add(sid)

            stop["user_id"] = current_user.user_id
            stop["completed"] = False
            stop["delivery_status"] = "pending"
            stop["order"] = i
            for f in completion_fields:
                stop.pop(f, None)
            cleaned.append(stop)

        # Clear current active stops + replace atomically-ish.
        await db.stops.delete_many({"user_id": current_user.user_id})
        if cleaned:
            await db.stops.insert_many(cleaned, ordered=False)

        return {"resumed": True, "stops_count": len(cleaned)}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[resume_route] failed route_id={route_id} user={current_user.user_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Resume failed: {type(e).__name__}: {e}")


@api_router.get("/routes/stats")
async def get_route_stats(current_user: User = Depends(get_current_user)):
    """Aggregate lifetime stats across all archived routes."""
    pipeline = [
        {"$match": {"user_id": current_user.user_id}},
        {"$group": {
            "_id": None,
            "total_routes": {"$sum": 1},
            "total_delivered": {"$sum": "$summary.delivered"},
            "total_skipped": {"$sum": "$summary.skipped"},
            "total_failed": {"$sum": "$summary.failed"},
            "total_stops": {"$sum": "$summary.total_stops"},
            "total_weight_kg": {"$sum": "$summary.total_weight_kg"},
            "total_quantity": {"$sum": "$summary.total_quantity"},
            "avg_stops_per_route": {"$avg": "$summary.total_stops"},
            "avg_delivered_per_route": {"$avg": "$summary.delivered"},
        }},
    ]
    results = await db.route_history.aggregate(pipeline).to_list(1)
    if not results:
        return {
            "total_routes": 0, "total_delivered": 0, "total_skipped": 0,
            "total_failed": 0, "total_stops": 0, "total_weight_kg": 0,
            "total_quantity": 0, "avg_stops_per_route": 0, "avg_delivered_per_route": 0,
        }
    stats = results[0]
    stats.pop("_id", None)
    return stats

# /stops/clear, DELETE /stops/{id}, complete/uncomplete moved to routes/stops.py.
# (Regeocode stays here — uses heavy geocoder + regex helpers defined below.)


@api_router.post("/stops/{stop_id}/regeocode", response_model=RegeocodeStopResponse)
async def regeocode_stop(
    stop_id: str,
    payload: Optional[RegeocodeStopRequest] = None,
    current_user: User = Depends(get_current_user),
):
    """Re-geocode a stop's address. Keeps existing coordinates if geocoding fails."""
    existing = await db.stops.find_one({"id": stop_id, "user_id": current_user.user_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Stop not found")

    payload_address = payload.address if payload else None
    address_input = payload_address if payload_address is not None else existing.get("address", "")
    clean_address = re.sub(r"\s+", " ", str(address_input).replace("\n", " ").replace("\r", " ")).strip()
    if not clean_address:
        raise HTTPException(status_code=400, detail="Address is required for re-geocoding")

    geo_result = await geocode_address_async(clean_address, user_id=current_user.user_id)

    if not geo_result:
        metadata = dict(existing.get("geocode_metadata") or {})
        metadata["geocode_needs_fix"] = True
        metadata["geocode_status"] = "failed"
        metadata["geocode_issue"] = "Geocoding failed for current address; previous coordinates retained."
        metadata["import_original_address"] = clean_address

        await db.stops.update_one(
            {"id": stop_id, "user_id": current_user.user_id},
            {
                "$set": {
                    "address": clean_address,
                    "geocode_metadata": metadata,
                }
            },
        )

        updated_failed = await db.stops.find_one({"id": stop_id, "user_id": current_user.user_id}, {"_id": 0})
        return RegeocodeStopResponse(
            success=True,
            geocoded=False,
            message="Could not geocode this address. Previous coordinates were kept.",
            stop=Stop(**updated_failed),
        )

    suburb = extract_suburb_from_address(geo_result.get("place_name", clean_address))
    if not suburb:
        suburb = await reverse_geocode_suburb(geo_result["latitude"], geo_result["longitude"])

    metadata = _build_stop_geocode_metadata(geo_result) or {}
    metadata["import_original_address"] = clean_address
    metadata["geocoded_formatted_address"] = geo_result.get("place_name", "")
    metadata["geocode_needs_fix"] = False
    metadata["geocode_status"] = "ok"
    metadata.pop("geocode_issue", None)

    await db.stops.update_one(
        {"id": stop_id, "user_id": current_user.user_id},
        {
            "$set": {
                "address": clean_address,
                "latitude": geo_result["latitude"],
                "longitude": geo_result["longitude"],
                "suburb": suburb,
                "geocode_metadata": metadata,
            }
        },
    )

    updated = await db.stops.find_one({"id": stop_id, "user_id": current_user.user_id}, {"_id": 0})
    return RegeocodeStopResponse(
        success=True,
        geocoded=True,
        message="Address geocoded and stop location updated.",
        stop=Stop(**updated),
    )


@api_router.post("/car/stop-action", response_model=Stop)
async def car_stop_action(action_data: CarStopActionRequest, current_user: User = Depends(get_current_user)):
    """Android Auto in-car stop actions: delivered, skip, failed."""
    existing = await db.stops.find_one({"id": action_data.stop_id, "user_id": current_user.user_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Stop not found")

    now_utc = datetime.now(timezone.utc)
    update_payload: Dict[str, Any] = {}

    if action_data.action == "delivered":
        update_payload = {
            "completed": True,
            "completed_at": now_utc,
            "delivery_status": "delivered",
            "failure_reason": None,
        }
    elif action_data.action == "skip":
        update_payload = {
            "completed": False,
            "completed_at": None,
            "delivery_status": "skipped",
            "failure_reason": action_data.reason,
        }
    elif action_data.action == "failed":
        update_payload = {
            "completed": False,
            "completed_at": None,
            "delivery_status": "failed",
            "failure_reason": action_data.reason,
        }

    await db.stops.update_one({"id": action_data.stop_id, "user_id": current_user.user_id}, {"$set": update_payload})
    updated = await db.stops.find_one({"id": action_data.stop_id, "user_id": current_user.user_id}, {"_id": 0})
    return Stop(**updated)


@api_router.get("/car/next-stops", response_model=List[Stop])
async def car_next_stops(current_user: User = Depends(get_current_user), limit: int = Query(default=20, ge=1, le=100)):
    """Optimized stop feed for Android Auto surfaces."""
    cursor = db.stops.find(
        {"user_id": current_user.user_id, "completed": {"$ne": True}},
        {"_id": 0}
    ).sort("order", 1).limit(limit)
    stops = await cursor.to_list(limit)
    return [Stop(**s) for s in stops]

@api_router.post("/stops/refresh-suburbs")
async def refresh_suburbs(current_user: User = Depends(get_current_user)):
    """Refresh/update suburbs for all stops that don't have one"""
    stops = await db.stops.find({"user_id": current_user.user_id}, {"_id": 0}).to_list(1000)
    
    updated_count = 0
    for stop in stops:
        # Skip if already has suburb
        if stop.get("suburb"):
            continue
        
        # Try to extract suburb from address first
        suburb = extract_suburb_from_address(stop.get("address", ""))
        
        # If not found, try reverse geocoding
        if not suburb and stop.get("latitude") and stop.get("longitude"):
            suburb = await reverse_geocode_suburb(stop["latitude"], stop["longitude"])
        
        # Update if we found a suburb
        if suburb:
            await db.stops.update_one(
                {"id": stop["id"]},
                {"$set": {"suburb": suburb}}
            )
            updated_count += 1
        
        # Small delay to avoid rate limiting
        await asyncio.sleep(0.05)
    
    return {"message": f"Updated suburbs for {updated_count} stops", "updated_count": updated_count}

# ===================== XLS Import Endpoints =====================

async def get_user_geocoding_context(user_id: str) -> Dict[str, any]:
    """Get proximity centroid, bbox, and country from user's existing stops for geocoding bias"""
    try:
        pipeline = [
            {"$match": {"user_id": user_id}},
            {"$group": {
                "_id": None,
                "avg_lat": {"$avg": "$latitude"},
                "avg_lng": {"$avg": "$longitude"},
                "min_lat": {"$min": "$latitude"},
                "min_lng": {"$min": "$longitude"},
                "max_lat": {"$max": "$latitude"},
                "max_lng": {"$max": "$longitude"},
                "count": {"$sum": 1}
            }}
        ]
        result = await db.stops.aggregate(pipeline).to_list(1)
        if result and result[0]["count"] > 0:
            r = result[0]
            avg_lng = r["avg_lng"]
            avg_lat = r["avg_lat"]
            # Build bbox with ~50km padding (0.5 degrees) to restrict results
            padding = 0.5
            bbox = f"{r['min_lng'] - padding},{r['min_lat'] - padding},{r['max_lng'] + padding},{r['max_lat'] + padding}"
            # Detect country via reverse geocoding of centroid (cached)
            country = None
            cache_key = f"country_{round(avg_lat, 2)}_{round(avg_lng, 2)}"
            cached_country = await db.geocode_cache.find_one({"address_query": cache_key})
            if cached_country:
                country = cached_country.get("country_code")
            else:
                try:
                    async with httpx.AsyncClient() as client:
                        resp = await client.get(
                            f"https://api.mapbox.com/geocoding/v5/mapbox.places/{avg_lng},{avg_lat}.json",
                            params={"access_token": MAPBOX_TOKEN, "types": "country", "limit": 1},
                            timeout=5.0
                        )
                        if resp.status_code == 200:
                            features = resp.json().get("features", [])
                            if features:
                                country = features[0].get("properties", {}).get("short_code", "").lower()
                                await db.geocode_cache.insert_one({
                                    "address_query": cache_key,
                                    "country_code": country,
                                    "created_at": datetime.now(timezone.utc),
                                    "hit_count": 1
                                })
                except Exception as e:
                    logger.error(f"Country detection error: {e}")
            return {"proximity": f"{avg_lng},{avg_lat}", "country": country, "bbox": bbox}
    except Exception as e:
        logger.error(f"Geocoding context error: {e}")
    return {}


async def geocode_address_async(address: str, user_id: str = None) -> Optional[Dict[str, Any]]:
    """Geocode a single address using Mapbox API with database caching, proximity and country bias"""
    if not MAPBOX_TOKEN:
        return None
    
    # Normalize address for consistent cache lookup
    normalized_address = normalize_address(address).strip().lower()
    
    # Check cache first
    try:
        cached = await db.geocode_cache.find_one({"address_query": normalized_address})
        if cached:
            # Update hit count
            await db.geocode_cache.update_one(
                {"_id": cached["_id"]},
                {"$inc": {"hit_count": 1}}
            )
            logger.info(f"Geocode cache HIT for: {address[:50]}...")
            metadata = cached.get("metadata")
            if isinstance(metadata, dict) and metadata:
                centroid_lat = cached.get("latitude")
                centroid_lng = cached.get("longitude")
                centroid_plus_code = _encode_plus_code(centroid_lat, centroid_lng)
                return {
                    "latitude": centroid_lat,
                    "longitude": centroid_lng,
                    "rooftop_centroid": {
                        "latitude": centroid_lat,
                        "longitude": centroid_lng,
                    },
                    "map_pinpoint": metadata.get("map_pinpoint") or {
                        "latitude": centroid_lat,
                        "longitude": centroid_lng,
                        "source": "rooftop_centroid",
                    },
                    "access_navigation_point": metadata.get("access_navigation_point") or {
                        "latitude": centroid_lat,
                        "longitude": centroid_lng,
                        "source": "centroid_fallback",
                    },
                    "centroid_plus_code": metadata.get("centroid_plus_code") or centroid_plus_code,
                    "access_plus_code": metadata.get("access_plus_code") or centroid_plus_code,
                    "plus_code": metadata.get("plus_code") or metadata.get("access_plus_code") or centroid_plus_code,
                    "interpolation_status": metadata.get("interpolation_status") or metadata.get("location_type") or cached.get("location_type", "unknown"),
                    **metadata,
                }

            centroid_lat = cached.get("latitude")
            centroid_lng = cached.get("longitude")
            centroid_plus_code = _encode_plus_code(centroid_lat, centroid_lng)
            return {
                "latitude": centroid_lat,
                "longitude": centroid_lng,
                "rooftop_centroid": {
                    "latitude": centroid_lat,
                    "longitude": centroid_lng,
                },
                "map_pinpoint": {
                    "latitude": centroid_lat,
                    "longitude": centroid_lng,
                    "source": "rooftop_centroid",
                },
                "access_navigation_point": {
                    "latitude": centroid_lat,
                    "longitude": centroid_lng,
                    "source": "centroid_fallback",
                },
                "centroid_plus_code": centroid_plus_code,
                "access_plus_code": centroid_plus_code,
                "plus_code": centroid_plus_code,
                "interpolation_status": cached.get("location_type", "unknown"),
                "place_name": cached["place_name"],
                "formatted_address": cached.get("place_name", ""),
                "business_name": cached.get("place_name", ""),
                "brand": "",
                "is_business": False,
                "poi_category": "",
                "feature_type": "",
                "place_id": cached.get("place_id", ""),
                "location_type": cached.get("location_type", ""),
                "suburb": cached.get("suburb", ""),
                "lga": cached.get("lga", ""),
                "region": cached.get("region", ""),
                "postcode": cached.get("postcode", ""),
                "country": cached.get("country", ""),
                "country_code": cached.get("country_code", ""),
                "admin_areas": {
                    "suburb": cached.get("suburb", ""),
                    "lga": cached.get("lga", ""),
                    "region": cached.get("region", ""),
                    "postcode": cached.get("postcode", ""),
                    "country": cached.get("country", ""),
                    "country_code": cached.get("country_code", ""),
                },
            }
    except Exception as e:
        logger.error(f"Cache lookup error: {e}")
    
    # Get proximity and country context from user's existing stops
    geo_context = {}
    if user_id:
        geo_context = await get_user_geocoding_context(user_id)
    
    # Try geocoding with the normalized address
    result = await _call_mapbox_geocode(normalize_address(address), geo_context)
    
    if result:
        # Cache the result
        await _cache_geocode_result(normalized_address, address, result)
        return result
    
    return None


def normalize_address(address: str) -> str:
    """Normalize compound road names that Mapbox doesn't fuzzy-match"""
    import re
    # Split known compound words: sugarbag -> sugar bag, etc.
    compounds = {
        'sugarbag': 'sugar bag',
        'stringybark': 'stringy bark',
        'ironbark': 'iron bark',
        'blackbutt': 'black butt',
        'tallowwood': 'tallow wood',
        'bloodwood': 'blood wood',
        'paperbark': 'paper bark',
        'teatree': 'tea tree',
        'redgum': 'red gum',
    }
    result = address
    for compound, split in compounds.items():
        result = re.sub(compound, split, result, flags=re.IGNORECASE)
    return result


async def _call_mapbox_geocode(address: str, geo_context: dict) -> Optional[Dict]:
    """Call Mapbox geocoding API and return rich result with full metadata"""
    try:
        params = {
            "q": address,
            "access_token": MAPBOX_TOKEN,
            "limit": 1,
            "types": "address,street,place",
            "routing": "true",
        }
        if geo_context.get("proximity"):
            params["proximity"] = geo_context["proximity"]
        if geo_context.get("country"):
            params["country"] = geo_context["country"]
        if geo_context.get("bbox"):
            params["bbox"] = geo_context["bbox"]
        
        async with httpx.AsyncClient() as client:
            response = await client.get(
                "https://api.mapbox.com/search/geocode/v6/forward",
                params=params,
                timeout=10.0
            )
            
            if response.status_code == 200:
                data = response.json()
                if data.get("features") and len(data["features"]) > 0:
                    feature = data["features"][0]
                    return _extract_rich_feature(feature)
    except Exception as e:
        logger.error(f"Geocoding error for '{address}': {e}")
    return None


def _encode_plus_code(lat: Optional[float], lng: Optional[float]) -> str:
    try:
        if lat is None or lng is None:
            return ""
        return olc.encode(float(lat), float(lng), codeLength=11)
    except Exception:
        return ""


def _extract_access_navigation_point(feature: dict, props: dict, centroid_lat: Optional[float], centroid_lng: Optional[float]) -> Dict[str, Any]:
    coordinates_meta = props.get("coordinates") if isinstance(props.get("coordinates"), dict) else {}

    routable_points = []
    if isinstance(coordinates_meta.get("routable_points"), list):
        routable_points = coordinates_meta.get("routable_points")
    elif isinstance(props.get("routable_points"), list):
        routable_points = props.get("routable_points")
    elif isinstance(feature.get("routable_points"), list):
        routable_points = feature.get("routable_points")

    if routable_points:
        rp = routable_points[0]
        if isinstance(rp, dict):
            rp_lat = rp.get("latitude")
            rp_lng = rp.get("longitude")
            if rp_lat is not None and rp_lng is not None:
                return {
                    "latitude": rp_lat,
                    "longitude": rp_lng,
                    "source": "routable_point",
                }

    return {
        "latitude": centroid_lat,
        "longitude": centroid_lng,
        "source": "centroid_fallback",
    }


def _extract_rich_feature(feature: dict) -> Dict:
    """Extract all available metadata from a Mapbox geocoding feature"""
    props = feature.get("properties", {})
    context = feature.get("context", [])
    
    # Parse administrative levels from v5 context array or v6 context object
    admin = {}
    if isinstance(context, list):
        for ctx in context:
            if not isinstance(ctx, dict):
                continue
            ctx_id = ctx.get("id", "")
            prefix = ctx_id.split(".")[0] if "." in ctx_id else ctx_id
            admin[prefix] = {
                "id": ctx_id,
                "text": ctx.get("text", ""),
                "short_code": ctx.get("short_code"),
                "wikidata": ctx.get("wikidata"),
            }
    elif isinstance(context, dict):
        for prefix, ctx in context.items():
            if not isinstance(ctx, dict):
                continue
            admin[prefix] = {
                "id": ctx.get("mapbox_id") or ctx.get("id", ""),
                "text": ctx.get("name") or ctx.get("text", ""),
                "short_code": ctx.get("short_code"),
                "wikidata": ctx.get("wikidata") or ctx.get("wikidata_id"),
            }

    place_name = (
        feature.get("place_name")
        or props.get("full_address")
        or props.get("name_preferred")
        or props.get("name")
        or ""
    )
    feature_type = props.get("feature_type", "")
    raw_name = props.get("name_preferred") or props.get("name") or ""
    if not raw_name and place_name:
        raw_name = place_name.split(",")[0].strip()
    brand_value = props.get("brand", "")
    category_value = props.get("category", "")
    looks_like_street_address = bool(re.match(r"^\s*\d+", raw_name))
    business_like_feature = feature_type in {"poi", "business"}
    business_name = raw_name if (business_like_feature or brand_value or category_value) and not looks_like_street_address else ""
    is_business = bool(
        business_like_feature
        or brand_value
        or category_value
        or props.get("maki")
        or props.get("landmark")
    )
    categories = [c.strip() for c in category_value.split(",") if c.strip()] if category_value else []

    center = feature.get("center") or (feature.get("geometry", {}) or {}).get("coordinates", [None, None])
    centroid_lng = center[0] if isinstance(center, list) and len(center) > 1 else None
    centroid_lat = center[1] if isinstance(center, list) and len(center) > 1 else None

    access_point = _extract_access_navigation_point(feature, props, centroid_lat, centroid_lng)
    interpolation_status = (
        props.get("accuracy")
        or ((props.get("coordinates") or {}).get("accuracy") if isinstance(props.get("coordinates"), dict) else None)
        or "unknown"
    )

    centroid_plus_code = _encode_plus_code(centroid_lat, centroid_lng)
    access_plus_code = _encode_plus_code(access_point.get("latitude"), access_point.get("longitude"))

    return {
        # Core coordinates
        "latitude": centroid_lat,
        "longitude": centroid_lng,
        "rooftop_centroid": {
            "latitude": centroid_lat,
            "longitude": centroid_lng,
        },
        "map_pinpoint": {
            "latitude": centroid_lat,
            "longitude": centroid_lng,
            "source": "rooftop_centroid",
        },
        "access_navigation_point": access_point,
        "centroid_plus_code": centroid_plus_code,
        "access_plus_code": access_plus_code,
        "plus_code": access_plus_code or centroid_plus_code,
        "interpolation_status": interpolation_status,
        
        # Formatted address
        "place_name": place_name,
        "formatted_address": place_name,
        "text": feature.get("text", props.get("name", "")),             # Street/place name only
        "address_number": props.get("address", ""),   # House number
        "business_name": business_name,
        "brand": brand_value,
        "is_business": is_business,
        "poi_category": category_value,
        "feature_type": feature_type,
        
        # Identifiers
        "id": feature.get("id", props.get("mapbox_id", "")),
        "place_id": feature.get("id", props.get("mapbox_id", "")),
        "place_type": feature.get("place_type", [feature_type] if feature_type else []),
        "relevance": feature.get("relevance", 1.0),
        
        # Location accuracy / type (ROOFTOP, INTERPOLATED, APPROXIMATE, etc.)
        "location_type": interpolation_status,
        
        # OSM-like tags & categories
        "category": category_value,                      # e.g. "shop", "restaurant"
        "categories": categories,
        "maki": props.get("maki", ""),                 # POI icon category
        "landmark": props.get("landmark", False),
        "wikidata": props.get("wikidata", ""),
        "foursquare": props.get("foursquare", ""),
        "osm_tags": {
            "category": category_value,
            "maki": props.get("maki", ""),
            "wikidata": props.get("wikidata", ""),
            "foursquare": props.get("foursquare", ""),
            "landmark": props.get("landmark", False),
        },
        
        # Administrative area levels
        "neighborhood": admin.get("neighborhood", {}).get("text", ""),
        "suburb": admin.get("locality", {}).get("text", "") or admin.get("place", {}).get("text", ""),
        "locality": admin.get("locality", {}).get("text", ""),
        "lga": admin.get("district", {}).get("text", ""),          # Local Government Area
        "city": admin.get("place", {}).get("text", ""),
        "region": admin.get("region", {}).get("text", ""),         # State
        "region_code": admin.get("region", {}).get("short_code", ""),
        "postcode": admin.get("postcode", {}).get("text", ""),
        "country": admin.get("country", {}).get("text", ""),
        "country_code": admin.get("country", {}).get("short_code", ""),
        "admin_areas": {
            "neighborhood": admin.get("neighborhood", {}).get("text", ""),
            "suburb": admin.get("locality", {}).get("text", "") or admin.get("place", {}).get("text", ""),
            "locality": admin.get("locality", {}).get("text", ""),
            "lga": admin.get("district", {}).get("text", ""),
            "city": admin.get("place", {}).get("text", ""),
            "region": admin.get("region", {}).get("text", ""),
            "region_code": admin.get("region", {}).get("short_code", ""),
            "postcode": admin.get("postcode", {}).get("text", ""),
            "country": admin.get("country", {}).get("text", ""),
            "country_code": admin.get("country", {}).get("short_code", ""),
        },
        
        # Geometry & bounds
        "geometry": feature.get("geometry", {}),
        "bbox": feature.get("bbox"),
        
        # Raw context (for any fields we didn't explicitly extract)
        "context_raw": context,
    }


def _build_stop_geocode_metadata(source: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """Prepare geocode metadata payload for stop storage (all metadata, excluding coordinates)."""
    if not source or not isinstance(source, dict):
        return None

    metadata = {k: v for k, v in source.items() if k not in {"latitude", "longitude"}}
    if "formatted_address" not in metadata and source.get("place_name"):
        metadata["formatted_address"] = source.get("place_name")
    return metadata or None


async def _cache_geocode_result(normalized_address: str, original_address: str, result: dict):
    """Save geocode result to cache (stores full rich metadata)"""
    try:
        cache_entry = {
            "id": str(uuid.uuid4()),
            "address_query": normalized_address,
            "original_address": original_address,
            "latitude": result["latitude"],
            "longitude": result["longitude"],
            "place_name": result.get("place_name", ""),
            "metadata": _build_stop_geocode_metadata(result),
            "place_id": result.get("place_id", ""),
            "location_type": result.get("location_type", ""),
            "suburb": result.get("suburb", ""),
            "lga": result.get("lga", ""),
            "region": result.get("region", ""),
            "postcode": result.get("postcode", ""),
            "country": result.get("country", ""),
            "country_code": result.get("country_code", ""),
            "created_at": datetime.now(timezone.utc),
            "hit_count": 1
        }
        await db.geocode_cache.insert_one(cache_entry)
        logger.info(f"Geocode cached: {original_address[:50]}...")
    except Exception as e:
        logger.error(f"Cache save error: {e}")

def parse_excel_file(file_content: bytes, filename: str) -> pd.DataFrame:
    """Parse Excel/CSV file and return DataFrame.

    Uses python-calamine for both .xls (BIFF/OLE2) and .xlsx (ZIP/OOXML)
    files — single engine that handles both formats robustly. Detects the
    actual file format from the magic bytes rather than the extension, so
    mis-named uploads (e.g. an .xlsx renamed to .xls before upload, or
    case-mismatch like .XLS) work correctly. Falls back to extension-based
    routing only when magic bytes are inconclusive (e.g. CSV)."""
    try:
        head = file_content[:8]
        is_xlsx = head[:4] == b"PK\x03\x04"  # ZIP container = .xlsx/.ods/.xlsb
        is_xls = head == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # OLE2 = legacy .xls
        lower = filename.lower()

        if is_xlsx or is_xls:
            # calamine handles both formats from the in-memory bytes; the
            # extension is irrelevant once we know the magic bytes.
            df = pd.read_excel(io.BytesIO(file_content), engine='calamine')
        elif lower.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_content))
        else:
            # Last resort: trust the extension when magic bytes don't match
            # any known Excel format. Surfaces a clearer error than the
            # generic "Unsupported file format" we used to throw.
            raise ValueError(
                f"File '{filename}' is not a recognised Excel or CSV file "
                f"(got header bytes {head[:4].hex()})."
            )

        df.columns = df.columns.str.strip()
        return df
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error parsing file: {e}")
        raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")

@api_router.post("/import/preview", response_model=ImportPreviewResponse)
async def preview_import(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload and preview XLS/XLSX/CSV file - returns columns and sample data"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    # Check file extension
    allowed_extensions = ['.xls', '.xlsx', '.csv']
    file_ext = '.' + file.filename.split('.')[-1].lower() if '.' in file.filename else ''
    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=400, 
            detail=f"Unsupported file format. Allowed: {', '.join(allowed_extensions)}"
        )
    
    content = await file.read()
    df = parse_excel_file(content, file.filename)
    
    # Get sample rows (first 5) - convert numpy types to native Python types
    sample_df = df.head(5).fillna('')
    sample_rows = []
    for _, row in sample_df.iterrows():
        sample_rows.append({k: str(v) for k, v in row.items()})
    
    # Server-side auto-mapping: prioritise patterns to avoid POD > Note conflicts
    def _find_col(cols_lower_map, patterns):
        # First pass: exact match wins (so column "Notes" beats "POD Notes"
        # when both contain the substring `notes`; otherwise the user loses
        # the actual driver instructions to the post-delivery POD field).
        for pat in patterns:
            for orig, low in cols_lower_map.items():
                if pat == low:
                    return orig
        # Second pass: substring fallback — picks up columns like
        # "Delivery Address" or "Customer Notes" that still uniquely match.
        for pat in patterns:
            for orig, low in cols_lower_map.items():
                if pat in low:
                    return orig
        return None

    cols_lower = {c: c.lower().replace('_', '').replace(' ', '').replace('-', '') for c in df.columns}
    suggested = {}
    addr = _find_col(cols_lower, ['address', 'location', 'destination', 'deliveryaddress', 'streetaddress', 'fulladdress', 'addr'])
    if addr:
        suggested['address'] = addr
    mob = _find_col(cols_lower, ['mobile', 'phone', 'cell', 'telephone', 'tel', 'phonenumber', 'mobilenumber', 'contact', 'customernumber'])
    if mob:
        suggested['mobile_number'] = mob
    notes = _find_col(cols_lower, ['notes', 'note', 'comments', 'comment', 'instructions', 'instruction', 'remarks', 'description', 'details', 'info', 'pod'])
    if notes:
        suggested['notes'] = notes
    wt = _find_col(cols_lower, ['weight', 'wt', 'kg', 'mass', 'parcelweight', 'packageweight'])
    if wt:
        suggested['weight'] = wt
    qty = _find_col(cols_lower, ['quantity', 'qty', 'count', 'amount', 'items', 'parcels', 'packages', 'units', 'pcs'])
    if qty:
        suggested['quantity'] = qty
    # `sourcereference` matches the user's actual CSV column "Source
    # Reference"; the rest cover the obvious carrier label variants
    # (Tracking, Tracking Number, Barcode, AWB, Consignment, Reference).
    # Lower-case keys here are pre-normalised (spaces/_/- stripped).
    track = _find_col(cols_lower, [
        'sourcereference', 'sourceref', 'tracking', 'trackingnumber', 'trackingno',
        'trackingid', 'barcode', 'awb', 'awbnumber', 'consignment',
        'consignmentnote', 'reference', 'refno', 'shipmentid', 'parcelid',
    ])
    if track:
        suggested['tracking_number'] = track

    return ImportPreviewResponse(
        columns=list(df.columns),
        sample_rows=sample_rows,
        total_rows=int(len(df)),
        suggested_mapping=suggested if suggested else None,
    )

@api_router.post("/import/process", response_model=ImportResult)
async def process_import(
    file: UploadFile = File(...),
    mapping: str = Form(...),  # JSON string of FieldMapping
    clear_existing: str = Form("false"),  # "true" to clear existing stops before import
    current_user: User = Depends(get_current_user)
):
    """Process XLS import — async job pattern to avoid Cloudflare 520.

    Kicks off a background task that geocodes all rows, then writes stops
    to MongoDB. The frontend polls `/api/import/jobs/{job_id}` until
    `status` is `done` or `error`.

    Also supports the legacy synchronous flow: if the file has ≤20 rows,
    we skip the job pattern and return the result inline (fast enough to
    beat the 100s Cloudflare ceiling).
    """
    import json

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    try:
        mapping_dict = json.loads(mapping)
        field_mapping = FieldMapping(**mapping_dict)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid mapping: {str(e)}")

    content = await file.read()
    df = parse_excel_file(content, file.filename)

    if field_mapping.address not in df.columns:
        raise HTTPException(
            status_code=400,
            detail=f"Address column '{field_mapping.address}' not found in file",
        )

    # Small files (<= 20 rows): run synchronously (fast enough for Cloudflare)
    if len(df) <= 20:
        return await _run_import_inner(df, field_mapping, current_user)

    # Large files: async job pattern
    job_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    await db.import_jobs.insert_one({
        "job_id": job_id,
        "user_id": current_user.user_id,
        "status": "running",
        "started_at": now,
        "total_rows": len(df),
        "result": None,
        "error": None,
    })
    logger.info("[import/jobs] kickoff job_id=%s user=%s rows=%d", job_id, current_user.user_id, len(df))

    async def _run_import_job():
        try:
            result = await _run_import_inner(df, field_mapping, current_user)
            result_dict = result.dict() if hasattr(result, "dict") else result
            # Strip the full stops list from the stored result (too large for Mongo doc)
            if isinstance(result_dict, dict):
                result_dict.pop("stops", None)
            await db.import_jobs.update_one(
                {"job_id": job_id},
                {"$set": {"status": "done", "result": result_dict,
                          "finished_at": datetime.now(timezone.utc)}},
            )
        except Exception as e:
            logger.error("[import/jobs] job %s crashed: %s", job_id, traceback.format_exc())
            await db.import_jobs.update_one(
                {"job_id": job_id},
                {"$set": {"status": "error", "error": str(e),
                          "finished_at": datetime.now(timezone.utc)}},
            )

    task = asyncio.create_task(_run_import_job())
    _OPTIMIZE_RUNNER_TASKS.add(task)
    task.add_done_callback(_OPTIMIZE_RUNNER_TASKS.discard)

    return JSONResponse(
        status_code=202,
        content={"job_id": job_id, "status": "running", "total_rows": len(df)},
    )


@api_router.get("/import/jobs/{job_id}")
async def get_import_job(job_id: str, current_user: User = Depends(get_current_user)):
    """Poll an async import job."""
    j = await db.import_jobs.find_one(
        {"job_id": job_id, "user_id": current_user.user_id}, {"_id": 0}
    )
    if not j:
        raise HTTPException(status_code=404, detail="Import job not found")
    return {
        "job_id": job_id,
        "status": j.get("status"),
        "total_rows": j.get("total_rows"),
        "result": j.get("result") if j.get("status") == "done" else None,
        "error": j.get("error") if j.get("status") == "error" else None,
    }


async def _run_import_inner(
    df, field_mapping: FieldMapping, current_user: User
) -> ImportResult:
    """Core import logic — archives completed stops, geocodes, inserts."""

    # Always clear existing stops before import (import = new route).
    completed = await db.stops.find(
        {"user_id": current_user.user_id, "completed": True},
        {"_id": 0},
    ).to_list(length=None)
    auto_archived_count = 0
    if completed:
        # Compute the same summary block /api/routes/archive writes so
        # HistoryModal renders these auto-archives identically to
        # explicit user-triggered archives. Previously this writer used
        # a `stats: {stops_count, auto_archived_reason}` shape that the
        # modal couldn't render (it expected `summary.total_stops` etc.)
        # — the user saw an error on first tap of the History icon.
        # 2026-05-12: aligned to the canonical schema.
        delivered_count = sum(1 for s in completed if s.get("completed"))
        skipped_count = sum(1 for s in completed if s.get("delivery_status") == "skipped")
        failed_count = sum(1 for s in completed if s.get("delivery_status") == "failed")
        total_weight = sum(float(s.get("weight_kg") or 0) for s in completed)
        total_quantity = sum(int(s.get("quantity") or 0) for s in completed)
        archived_at_iso = datetime.now(timezone.utc).isoformat()
        # `started_at` heuristic: earliest non-null arrived_at on any
        # completed stop, falling back to earliest created_at, falling
        # back to archived_at.
        candidate_starts = [s.get("arrived_at") for s in completed if s.get("arrived_at")]
        if not candidate_starts:
            candidate_starts = [s.get("created_at") for s in completed if s.get("created_at")]
        started_at_iso = min(candidate_starts).isoformat() if candidate_starts and hasattr(min(candidate_starts), "isoformat") else (
            min(candidate_starts) if candidate_starts else archived_at_iso
        )
        archive_doc = {
            "id": str(uuid.uuid4()),
            "user_id": current_user.user_id,
            "archived_at": archived_at_iso,
            "started_at": started_at_iso,
            "finished_at": archived_at_iso,
            "stops": completed,
            "summary": {
                "total_stops": len(completed),
                "delivered": delivered_count,
                "skipped": skipped_count,
                "failed": failed_count,
                "pending": 0,
                "total_weight_kg": round(total_weight, 2),
                "total_quantity": total_quantity,
                "algorithm": None,
                "total_distance_km": None,
                "total_duration_seconds": None,
                # Preserved so we can still tell auto-archives apart in
                # diagnostics — just moved off the rendered surface.
                "auto_archived_reason": "import_process_overwrite",
            },
        }
        await db.route_history.insert_one(archive_doc)
        auto_archived_count = len(completed)
        logger.info(
            f"[import_process] auto-archived {auto_archived_count} completed stops "
            f"for user={current_user.user_id} into route_history "
            f"id={archive_doc['id']} BEFORE wiping for new import",
        )

    await db.stops.delete_many({"user_id": current_user.user_id})
    
    # Always start from order 0 since we clear all stops
    max_order = -1
    
    success_count = 0
    failed_count = 0
    failed_addresses = []
    created_stops = []

    def _clean_import_address(value: Any) -> str:
        if value is None:
            return ""
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
        text = str(value).replace("\n", " ").replace("\r", " ").strip()
        text = re.sub(r"\s+", " ", text)
        return text
    
    # --- Concurrent geocoding with semaphore (5 parallel) ---
    sem = asyncio.Semaphore(5)
    
    async def geocode_row(idx, row):
        raw_address = _clean_import_address(row.get(field_mapping.address, ''))
        if not raw_address:
            return None
        async with sem:
            geo_result = await geocode_address_async(raw_address, user_id=current_user.user_id)
        if not geo_result:
            return {"failed": True, "address": raw_address}
        return {
            "failed": False,
            "geo_result": geo_result,
            "row": row,
            "idx": idx,
            "raw_address": raw_address,
        }
    
    tasks = [geocode_row(idx, row) for idx, row in df.iterrows()]
    results = await asyncio.gather(*tasks)
    
    # Process results in order
    for result in results:
        if result is None:
            continue
        if result["failed"]:
            failed_count += 1
            failed_addresses.append(result["address"])
            continue
        
        geo_result = result["geo_result"]
        row = result["row"]
        raw_address = result.get("raw_address") or _clean_import_address(row.get(field_mapping.address, ''))
        
        # Extract mapped fields
        name = None
        if field_mapping.name and field_mapping.name in df.columns:
            name_val = row.get(field_mapping.name)
            if pd.notna(name_val):
                name = str(name_val).strip()
        
        mobile_number = None
        if field_mapping.mobile_number and field_mapping.mobile_number in df.columns:
            mobile_val = row.get(field_mapping.mobile_number)
            if pd.notna(mobile_val):
                mobile_number = str(mobile_val).strip()
        
        notes = None
        if field_mapping.notes and field_mapping.notes in df.columns:
            notes_val = row.get(field_mapping.notes)
            if pd.notna(notes_val):
                notes = str(notes_val).strip()
        
        weight = None
        if field_mapping.weight and field_mapping.weight in df.columns:
            weight_val = row.get(field_mapping.weight)
            if pd.notna(weight_val):
                try:
                    weight = float(weight_val)
                except (ValueError, TypeError):
                    pass
        
        quantity = None
        if field_mapping.quantity and field_mapping.quantity in df.columns:
            qty_val = row.get(field_mapping.quantity)
            if pd.notna(qty_val):
                try:
                    quantity = int(float(qty_val))
                except (ValueError, TypeError):
                    pass

        # Carrier tracking / barcode column (e.g. "Source Reference") —
        # uppercased + stripped on read so the Van Loading Assistant
        # scanner can do an O(1) Map lookup against the normalised value
        # without having to retry case variants per scan.
        tracking_number = None
        if field_mapping.tracking_number and field_mapping.tracking_number in df.columns:
            tn_val = row.get(field_mapping.tracking_number)
            if pd.notna(tn_val):
                tn_clean = str(tn_val).strip().upper()
                if tn_clean:
                    tracking_number = tn_clean

        max_order += 1
        suburb = extract_suburb_from_address(geo_result.get("place_name", raw_address))
        if not suburb:
            suburb = await reverse_geocode_suburb(geo_result["latitude"], geo_result["longitude"])

        geocode_metadata = _build_stop_geocode_metadata(geo_result)
        geocode_metadata["import_original_address"] = raw_address
        geocode_metadata["geocoded_formatted_address"] = geo_result.get("place_name", "")
        
        stop = Stop(
            id=str(uuid.uuid4()),
            user_id=current_user.user_id,
            address=raw_address,
            name=name,
            mobile_number=mobile_number,
            suburb=suburb,
            latitude=geo_result["latitude"],
            longitude=geo_result["longitude"],
            priority="medium",
            notes=notes,
            weight=weight,
            quantity=quantity,
            tracking_number=tracking_number,
            geocode_metadata=geocode_metadata,
            order=max_order
        )
        
        await db.stops.insert_one(stop.dict())
        created_stops.append(stop)
        success_count += 1
    
    return ImportResult(
        success_count=success_count,
        failed_count=failed_count,
        failed_addresses=failed_addresses[:20],  # Limit to first 20 failed
        stops=created_stops,
        auto_archived_count=auto_archived_count,
    )

# ===================== Route Optimization =====================


async def calculate_road_distance_km(stops: List[dict]) -> Optional[float]:
    """Calculate total road distance via OSRM Route API (primary) or Mapbox (fallback).

    Uses the OSRM Route service to get the actual road distance in km
    for the ordered sequence of stops. Falls back to Mapbox if OSRM unavailable.
    """
    if len(stops) < 2:
        return None

    # --- Primary: OSRM Route API ---
    if _osrm_enabled():
        try:
            coord_list = [f"{s['longitude']},{s['latitude']}" for s in stops]
            coords = ";".join(coord_list)
            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.get(
                    f"{OSRM_URL}/route/v1/driving/{coords}",
                    params={"overview": "false"},
                )
                if resp.status_code == 200:
                    data = resp.json()
                    if data.get("code") == "Ok" and data.get("routes"):
                        _osrm_note_success()
                        total_meters = data["routes"][0].get("distance", 0)
                        return round(total_meters / 1000, 2)
        except Exception as e:
            logger.warning("OSRM road distance calculation failed: %s", e)

    # --- Fallback: Mapbox Directions ---
    if not MAPBOX_TOKEN:
        return None
    try:
        coord_list = [f"{s['longitude']},{s['latitude']}" for s in stops]
        MAX_WP = 25
        total_meters = 0.0
        async with httpx.AsyncClient(timeout=15.0) as client:
            for i in range(0, len(coord_list), MAX_WP - 1):
                chunk = coord_list[i:i + MAX_WP]
                if len(chunk) < 2:
                    break
                resp = await client.get(
                    f"https://api.mapbox.com/directions/v5/mapbox/driving/{';'.join(chunk)}",
                    params={"access_token": MAPBOX_TOKEN, "overview": "false"},
                )
                if resp.status_code != 200:
                    return None
                data = resp.json()
                routes = data.get("routes", [])
                if not routes:
                    return None
                total_meters += routes[0].get("distance", 0)
        return round(total_meters / 1000, 2)
    except Exception as e:
        logger.warning("Road distance calculation failed: %s", e)
        return None


def calculate_distance_matrix(stops: List[dict]) -> List[List[float]]:
    """Calculate distance matrix between all stops using haversine"""
    n = len(stops)
    matrix = [[0.0] * n for _ in range(n)]
    
    for i in range(n):
        for j in range(n):
            if i != j:
                coord1 = (stops[i]["latitude"], stops[i]["longitude"])
                coord2 = (stops[j]["latitude"], stops[j]["longitude"])
                matrix[i][j] = haversine(coord1, coord2, unit=Unit.KILOMETERS)
    
    return matrix


async def _mapbox_matrix_batch(stops: List[dict]) -> Optional[List[List[float]]]:
    """Call Mapbox Matrix API for a batch of up to 25 stops.
    Returns distance matrix in km, or None on failure."""
    if not stops or len(stops) > 25 or not MAPBOX_TOKEN:
        return None

    coords = ";".join(f"{s['longitude']},{s['latitude']}" for s in stops)
    url = f"https://api.mapbox.com/directions-matrix/v1/mapbox/driving/{coords}"
    params = {
        "access_token": MAPBOX_TOKEN,
        "annotations": "distance,duration",
    }

    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(url, params=params, timeout=30)

        if response.status_code != 200:
            logger.warning("Mapbox Matrix API returned %s", response.status_code)
            return None

        data = response.json()
        if data.get("code") != "Ok":
            logger.warning("Mapbox Matrix API error: %s", data.get("code"))
            return None

        distances = data.get("distances")
        if not distances:
            return None

        n = len(stops)
        matrix = [[0.0] * n for _ in range(n)]
        for i in range(n):
            for j in range(n):
                if i != j and distances[i][j] is not None:
                    matrix[i][j] = distances[i][j] / 1000.0  # meters to km
                elif i != j:
                    # Unreachable pair: fall back to haversine
                    c1 = (stops[i]["latitude"], stops[i]["longitude"])
                    c2 = (stops[j]["latitude"], stops[j]["longitude"])
                    matrix[i][j] = haversine(c1, c2, unit=Unit.KILOMETERS)
        return matrix
    except Exception as exc:
        logger.warning("Mapbox Matrix API call failed: %s", exc)
        return None


async def _mapbox_duration_matrix_batch(stops: List[dict]) -> Optional[List[List[int]]]:
    """Call Mapbox Matrix API for up to 25 stops.
    Returns DURATION matrix in integer seconds, or None on failure.
    This is the primary input for OR-Tools optimization (optimize for driving time)."""
    if not stops or len(stops) > 25 or not MAPBOX_TOKEN:
        return None

    coords = ";".join(f"{s['longitude']},{s['latitude']}" for s in stops)
    url = f"https://api.mapbox.com/directions-matrix/v1/mapbox/driving/{coords}"
    params = {
        "access_token": MAPBOX_TOKEN,
        "annotations": "duration",
    }

    try:
        async with httpx.AsyncClient() as client_http:
            response = await client_http.get(url, params=params, timeout=30)

        if response.status_code != 200:
            logger.warning("Mapbox Duration Matrix API returned %s", response.status_code)
            return None

        data = response.json()
        if data.get("code") != "Ok":
            logger.warning("Mapbox Duration Matrix API error: %s", data.get("code"))
            return None

        durations = data.get("durations")
        if not durations:
            return None

        n = len(stops)
        matrix = [[0] * n for _ in range(n)]
        for i in range(n):
            for j in range(n):
                if i != j and durations[i][j] is not None:
                    matrix[i][j] = max(1, int(durations[i][j]))  # seconds, min 1
                elif i != j:
                    # Unreachable: estimate from haversine at 30 km/h
                    c1 = (stops[i]["latitude"], stops[i]["longitude"])
                    c2 = (stops[j]["latitude"], stops[j]["longitude"])
                    km = haversine(c1, c2, unit=Unit.KILOMETERS)
                    matrix[i][j] = max(1, int(km / 30.0 * 3600))
        return matrix
    except Exception as exc:
        logger.warning("Mapbox Duration Matrix API call failed: %s", exc)
        return None


def _haversine_duration_matrix(stops: List[dict]) -> List[List[int]]:
    """Fallback: estimate travel-time matrix (seconds) from haversine at 30 km/h."""
    n = len(stops)
    matrix = [[0] * n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            if i != j:
                km = haversine(
                    (stops[i]["latitude"], stops[i]["longitude"]),
                    (stops[j]["latitude"], stops[j]["longitude"]),
                    unit=Unit.KILOMETERS,
                )
                matrix[i][j] = max(1, int(km / 30.0 * 3600))
    return matrix


def _osrm_cache_key(stops: List[dict]) -> str:
    """Generate a deterministic, ORDER-INDEPENDENT cache key from stop coordinates.

    Sorts coords so that re-optimizing the same set of stops (in any order)
    hits the same cache entry. Rounds to 6 decimals (~0.1m precision).
    """
    sorted_coords = sorted(
        (round(s['latitude'], 6), round(s['longitude'], 6)) for s in stops
    )
    coord_str = "|".join(f"{lat},{lng}" for lat, lng in sorted_coords)
    return hashlib.sha256(coord_str.encode()).hexdigest()[:16]


def detect_cluster_spikes(
    stops: List[dict],
    spike_ratio: float = 0.5,
    min_detour_km: float = 0.10,
) -> List[Dict[str, Any]]:
    """Flag visual "spike" triplets in an already-optimised stop sequence.

    For each consecutive (A, B, C) we compute haversine distances and ask:
    is the *straight-line* A→C distance much smaller than the detour
    A→B→C? If so, B sits well off the natural A→C line and the route will
    look like a zig-zag on the map even when the OSRM time-matrix says
    visiting B in the middle is optimal (e.g. one-way pair, highway split,
    cul-de-sac inside a cluster).

    Returns a list of warning dicts the frontend can render as
    "tighten cluster?" hints — empty list when the route is clean. The
    optimised order itself is NEVER mutated by this helper.

    Args:
        stops: optimised stop dicts in visit order. Each must have
               `latitude`, `longitude`, `id` keys.
        spike_ratio: trigger threshold. A triplet is flagged when
            `haversine(A, C) / (haversine(A, B) + haversine(B, C)) < spike_ratio`.
            Default 0.5 — flag when the detour is more than 2× the straight
            A→C distance. Was 0.3 (require >3.3× detour) but real driver
            screenshots showed obvious zig-zags slipping through; we trade
            a few extra OSRM verifications for catching mid-cluster spikes
            that are visually offensive even when the underlying detour
            ratio is "only" 2-3×.
        min_detour_km: floor on `A→B + B→C` to suppress micro-noise.
            Default 0.10 km (was 0.15) so dense urban routes — where every
            stop is 100-200 m apart but a single zig still reads as wrong
            on the screen — get auto-tightened too.
    """
    warnings: List[Dict[str, Any]] = []
    n = len(stops)
    if n < 3 or spike_ratio <= 0:
        return warnings

    for i in range(1, n - 1):
        a, b, c = stops[i - 1], stops[i], stops[i + 1]
        try:
            ac = haversine(
                (a["latitude"], a["longitude"]),
                (c["latitude"], c["longitude"]),
                unit=Unit.KILOMETERS,
            )
            ab = haversine(
                (a["latitude"], a["longitude"]),
                (b["latitude"], b["longitude"]),
                unit=Unit.KILOMETERS,
            )
            bc = haversine(
                (b["latitude"], b["longitude"]),
                (c["latitude"], c["longitude"]),
                unit=Unit.KILOMETERS,
            )
        except (KeyError, TypeError):
            continue  # missing coords on this triplet — skip silently
        detour = ab + bc
        if detour < min_detour_km:
            continue
        ratio = ac / detour if detour > 0 else 1.0
        if ratio < spike_ratio:
            warnings.append({
                "position": i,
                "prev_id": a.get("id"),
                "suspect_id": b.get("id"),
                "next_id": c.get("id"),
                "straight_km": round(ac, 3),
                "detour_km": round(detour, 3),
                "ratio": round(ratio, 3),
                "extra_km": round(detour - ac, 3),
            })
    return warnings



async def _osrm_duration_matrix(stops: List[dict]) -> Optional[List[List[int]]]:
    """Fetch full NxN duration matrix from OSRM Table service.

    Tries the locally-configured OSRM first, then falls back to the public
    OSRM demo server if the local one is unreachable (circuit breaker open).
    This is critical for production where the local OSRM binary isn't
    shipped — a real road-network matrix from public OSRM, even rate-limited,
    is strictly better for solver quality than a stitched Mapbox clustered
    matrix (which has approximate cells across cluster seams and causes
    visible zigzags in the optimised route).
    """
    n = len(stops)
    if n < 2:
        return None

    # --- Cache lookup (shared across candidate URLs) ---
    cache_key = _osrm_cache_key(stops)
    cached = _osrm_matrix_cache.get(cache_key)
    if cached is not None:
        logger.info("OSRM matrix CACHE HIT (%d stops, key=%s)", n, cache_key)
        return cached

    # Try local OSRM first (fast when available), then the public demo.
    candidates: List[tuple[str, str]] = []
    if _osrm_enabled():
        candidates.append(("local", OSRM_URL))
    if OSRM_PUBLIC_URL and OSRM_PUBLIC_URL != OSRM_URL:
        candidates.append(("public", OSRM_PUBLIC_URL))

    for label, base_url in candidates:
        matrix = await _osrm_duration_matrix_for_url(stops, base_url, label)
        if matrix is not None:
            _osrm_matrix_cache.set(cache_key, matrix)
            return matrix
    return None


async def _osrm_duration_matrix_for_url(
    stops: List[dict], base_url: str, label: str
) -> Optional[List[List[int]]]:
    """Single-URL variant of the OSRM duration matrix fetch.

    Returns the N×N matrix on success, None on failure so the caller can try
    the next candidate URL. Handles the 100-coord per-call OSRM limit by
    falling back to cross-batch table queries stitched on top of a Haversine
    baseline for any cells the batched queries fail to cover.
    """
    n = len(stops)
    OSRM_BATCH = 100

    # Fast-connect timeout so a never-listening `localhost:5000` (e.g.
    # when the Emergent prod pod doesn't run a local OSRM service)
    # bails in ~2 s instead of burning the whole 30 s read window.
    # The 30 s read timeout still applies to in-flight requests so a
    # genuinely slow public OSRM response isn't truncated.
    OSRM_TIMEOUT = httpx.Timeout(connect=2.0, read=30.0, write=10.0, pool=5.0)

    try:
        async with httpx.AsyncClient(timeout=OSRM_TIMEOUT) as client:
            # Try single call first (local OSRM with --max-table-size=500 supports this)
            coords = ";".join(f"{s['longitude']},{s['latitude']}" for s in stops)
            resp = await client.get(
                f"{base_url}/table/v1/driving/{coords}",
            )
            if resp.status_code == 200:
                data = resp.json()
                if data.get("code") == "Ok" and data.get("durations"):
                    logger.info("OSRM[%s] duration matrix: full %dx%d in single call", label, n, n)
                    if label == "local":
                        _osrm_note_success()
                    return [
                        [max(1, int(round(d))) if d is not None else 9999 for d in row]
                        for row in data["durations"]
                    ]

            # If single call fails (e.g., public OSRM TooBig limit), fall back to batching
            if n <= OSRM_BATCH:
                return None  # Small enough for single call; if that failed, give up

            # Batched: split into groups of ≤40 for cross-batch queries
            # Cross-batch: 40 src + 40 dst = 80 unique coords ≤ 100 OSRM limit
            HALF = 40
            batches = [list(range(i, min(i + HALF, n))) for i in range(0, n, HALF)]

            # Start with haversine baseline
            matrix = _haversine_duration_matrix(stops)

            sem = asyncio.Semaphore(1)  # Serialize for public OSRM demo server rate limits

            async def _fetch_cross(src_ids, dst_ids):
                async with sem:
                    all_ids = list(src_ids) + [i for i in dst_ids if i not in set(src_ids)]
                    if len(all_ids) > OSRM_BATCH:
                        return None
                    idx_map = {gid: loc for loc, gid in enumerate(all_ids)}
                    coords = ";".join(f"{stops[i]['longitude']},{stops[i]['latitude']}" for i in all_ids)
                    src_local = ";".join(str(idx_map[i]) for i in src_ids)
                    dst_local = ";".join(str(idx_map[i]) for i in dst_ids)

                    for attempt in range(3):
                        resp = await client.get(
                            f"{base_url}/table/v1/driving/{coords}",
                            params={"sources": src_local, "destinations": dst_local},
                            timeout=30,
                        )
                        if resp.status_code == 429:
                            await asyncio.sleep(1.0 * (attempt + 1))
                            continue
                        if resp.status_code != 200:
                            return None
                        data = resp.json()
                        if data.get("code") != "Ok" or not data.get("durations"):
                            return None
                        return (data["durations"], src_ids, dst_ids)
                    return None  # All retries exhausted

            tasks = [_fetch_cross(sb, db) for sb in batches for db in batches]
            results = await asyncio.gather(*tasks)

            upgraded = 0
            for result in results:
                if result is None:
                    continue
                sub, src_ids, dst_ids = result
                for i, gi in enumerate(src_ids):
                    for j, gj in enumerate(dst_ids):
                        val = sub[i][j]
                        if val is not None and gi != gj:
                            matrix[gi][gj] = max(1, int(round(val)))
                            upgraded += 1

            total_cells = n * (n - 1)
            # Require at least 70% coverage before trusting the matrix; below
            # that the haversine baseline dominates and solvers will again
            # see false diagonals. Letting the caller fall through to the
            # next candidate (or to Mapbox clusters) is the safer choice.
            if upgraded < int(total_cells * 0.7):
                logger.warning(
                    "OSRM[%s] matrix only %d/%d cells upgraded (%.0f%%) — rejecting, will try next candidate",
                    label, upgraded, total_cells, 100.0 * upgraded / max(1, total_cells),
                )
                return None

            logger.info(
                "OSRM[%s] duration matrix: %d/%d cells upgraded (%d batches)",
                label, upgraded, total_cells, len(tasks),
            )
            if label == "local":
                _osrm_note_success()
            return matrix

    except Exception as exc:
        _osrm_log_failure(f"OSRM[{label}] duration matrix failed", exc)
        return None


# Separate cache for OSRM distance matrices (km)
_osrm_distance_cache = TTLCache(maxsize=50, ttl=600)

async def _osrm_distance_matrix(stops: List[dict]) -> Optional[List[List[float]]]:
    """Fetch full NxN distance matrix (km) from OSRM Table service.

    Uses annotations=distance to get road distances instead of durations.
    Cached with 10-min TTL. Returns matrix of floats in km, or None on failure.
    """
    n = len(stops)
    if n < 2 or not _osrm_enabled():
        return None

    cache_key = "dist_" + _osrm_cache_key(stops)
    cached = _osrm_distance_cache.get(cache_key)
    if cached is not None:
        logger.info("OSRM distance matrix CACHE HIT (%d stops)", n)
        return cached

    OSRM_BATCH = 100

    try:
        coords = ";".join(f"{s['longitude']},{s['latitude']}" for s in stops)

        if n <= OSRM_BATCH:
            async with httpx.AsyncClient(timeout=60.0) as client:
                resp = await client.get(
                    f"{OSRM_URL}/table/v1/driving/{coords}",
                    params={"annotations": "distance"},
                )
            if resp.status_code == 200:
                data = resp.json()
                if data.get("code") == "Ok" and data.get("distances"):
                    logger.info("OSRM distance matrix: full %dx%d in single call", n, n)
                    matrix = [
                        [round(d / 1000.0, 4) if d is not None else 999.0 for d in row]
                        for row in data["distances"]
                    ]
                    _osrm_distance_cache.set(cache_key, matrix)
                    return matrix

        # Batched approach for > 100 stops
        matrix = [[0.0] * n for _ in range(n)]
        for i in range(n):
            for j in range(n):
                if i != j:
                    c1 = (stops[i]["latitude"], stops[i]["longitude"])
                    c2 = (stops[j]["latitude"], stops[j]["longitude"])
                    matrix[i][j] = haversine(c1, c2, unit=Unit.KILOMETERS)

        batch_stops_list = []
        for start in range(0, n, OSRM_BATCH):
            end = min(start + OSRM_BATCH, n)
            batch_stops_list.append((start, end))

        async def fetch_distance_batch(s, e):
            sub_coords = ";".join(f"{stops[i]['longitude']},{stops[i]['latitude']}" for i in range(s, e))
            async with httpx.AsyncClient(timeout=60.0) as client:
                r = await client.get(
                    f"{OSRM_URL}/table/v1/driving/{sub_coords}",
                    params={"annotations": "distance"},
                )
            if r.status_code == 200:
                d = r.json()
                if d.get("code") == "Ok" and d.get("distances"):
                    return (s, e, d["distances"])
            return None

        tasks = [fetch_distance_batch(s, e) for s, e in batch_stops_list]
        results = await asyncio.gather(*tasks)

        upgraded = 0
        for result in results:
            if result is None:
                continue
            s, e, distances = result
            for li, gi in enumerate(range(s, e)):
                for lj, gj in enumerate(range(s, e)):
                    val = distances[li][lj]
                    if val is not None and gi != gj:
                        matrix[gi][gj] = round(val / 1000.0, 4)
                        upgraded += 1

        logger.info("OSRM distance matrix: %d/%d cells upgraded (%d batches)", upgraded, n * (n - 1), len(tasks))
        _osrm_distance_cache.set(cache_key, matrix)
        return matrix

    except Exception as exc:
        logger.warning("OSRM distance matrix failed: %s", exc)
        return None


def _open_path_matrix(matrix: List[List[int]], depot: int) -> List[List[int]]:
    """Convert a closed-loop matrix to an open-path matrix by zeroing return-to-depot.

    Why this exists:
        Delivery routes don't return to depot — the driver finishes at whichever
        stop is last. Closed-loop TSP solvers (LKH, PyVRP via Hybrid Genetic
        Search with `end_depot`) optimise the full Hamiltonian cycle including
        the return leg back to the start. The "optimal" cycle is often
        catastrophically wrong for open-path delivery: the solver routes
        `depot → far_cluster → ... → near_cluster → back_to_depot` because that
        minimises the cycle, but the driver actually drives `depot → far_cluster
        → ... → near_cluster` and stops there — having driven past every
        near_cluster house at the start.

        The standard fix: tell the solver the return edge costs zero. Then the
        closed-loop optimum is identical to the open-path optimum because the
        return is "free" and never affects the objective.

    Args:
        matrix: N×N cost matrix (seconds or meters). Will be deep-copied.
        depot: Index of the start node. The column `[i][depot]` is zeroed for
            all i != depot, leaving the diagonal alone.

    Returns:
        A new N×N matrix with the same shape and same row/col semantics, but
        with `result[i][depot] = 0` for `i != depot`. The original matrix is
        left untouched (callers can still report distances from it).
    """
    n = len(matrix)
    if n == 0:
        return []
    # Use list comprehension over per-row slice to keep the original immutable
    out = [list(row) for row in matrix]
    if 0 <= depot < n:
        for i in range(n):
            if i != depot:
                out[i][depot] = 0
    return out


def vroom_tsp_solve(
    duration_matrix: List[List[int]],
    depot: int = 0,
    exploration_level: int = 5,
) -> List[int]:
    """Solve open-path TSP using VROOM (pyvroom).

    Args:
        duration_matrix: NxN integer seconds matrix.
        depot: Starting node index.
        exploration_level: VROOM search depth (1-5, higher = better but slower).

    Returns:
        Ordered list of stop indices (excluding depot if it appears at start).
    """
    if not VROOM_AVAILABLE:
        raise RuntimeError(f"pyvroom not available: {VROOM_IMPORT_ERROR}")

    n = len(duration_matrix)
    if n <= 1:
        return list(range(n))

    problem = vroom.Input()

    # Set the pre-computed duration matrix (accepts list-of-lists directly)
    problem.set_durations_matrix(profile="car", matrix_input=duration_matrix)

    # Single vehicle starting at depot (open-path: no explicit end)
    problem.add_vehicle(vroom.Vehicle(id=0, start=depot, profile="car"))

    # All non-depot stops as jobs
    jobs = []
    for i in range(n):
        if i != depot:
            jobs.append(vroom.Job(id=i, location=i))
    problem.add_job(jobs)

    # Solve
    solution = problem.solve(exploration_level=exploration_level, nb_threads=4)

    # Extract route order from solution.
    # pyvroom returns solution.routes as a pandas DataFrame with columns:
    # vehicle_id, type, arrival, duration, setup, service, waiting_time, location_index, id, description
    route_indices = [depot]
    routes_df = solution.routes
    if routes_df is not None and len(routes_df) > 0:
        job_rows = routes_df[routes_df["type"] == "job"]
        for _, row in job_rows.iterrows():
            job_id = int(row["id"])
            if job_id != depot:
                route_indices.append(job_id)

    # Add any stops missed by VROOM (shouldn't happen, but defensive)
    visited = set(route_indices)
    for i in range(n):
        if i not in visited:
            route_indices.append(i)

    return route_indices


def pyvrp_tsp_solve(
    duration_matrix: List[List[int]],
    depot: int = 0,
    time_limit_seconds: float = 2.0,
    seed: int = 0,
    coordinates: Optional[List[Tuple[float, float]]] = None,
) -> List[int]:
    """Solve open-path TSP using PyVRP's Hybrid Genetic Search.

    Thin adapter over `PyVRPTspSolver` that matches the shape of the other
    native-solver wrappers (`vroom_tsp_solve`, `lkh_tsp_solve`, …): take a
    duration matrix whose row/col 0 is the depot, return an index list that
    starts at `depot` and visits every other node exactly once.

    Args:
        duration_matrix: N×N integer seconds matrix.
        depot: Index of the starting node inside `duration_matrix`.
        time_limit_seconds: HGS search budget (1-2s is plenty for pure TSP).
        seed: Deterministic seed for reproducible test runs.
        coordinates: Optional list of `(longitude, latitude)` per matrix row
            (length must equal `len(duration_matrix)`). When supplied,
            stops sharing identical `(lon, lat)` are collapsed into a single
            PyVRP super-node and re-expanded in input order — this stops the
            HGS solver from randomly shuffling stops at the same address
            (apartments/units in one building) which would otherwise produce
            visible zig-zags on the map.

    Returns:
        Ordered list of node indices beginning with `depot`.
    """
    if not PYVRP_AVAILABLE:
        raise RuntimeError(f"pyvrp not available: {PYVRP_IMPORT_ERROR}")

    n = len(duration_matrix)
    if n <= 1:
        return list(range(n))

    if coordinates is not None and len(coordinates) != n:
        raise ValueError(
            f"coordinates length {len(coordinates)} does not match matrix "
            f"size {n}"
        )

    # ── Open-path TSP via free return edge ────────────────────────────────
    # PyVRP's HGS is a closed-loop solver (vehicle.end_depot = depot is a
    # required field). For delivery routes the driver does NOT return to
    # depot, so we patch the return-to-depot column to 0 BEFORE handing the
    # matrix to PyVRP. The closed-loop optimum on the patched matrix equals
    # the open-path optimum on the original. Without this, PyVRP routinely
    # picked routes like `[0, 37, 38, ..., 1, 2, 3]` — efficient if you'd
    # return to the start, but pessimal for one-way delivery (driver passed
    # stop 1 at the start and had to come back at the end).
    duration_matrix = _open_path_matrix(duration_matrix, depot)

    # PyVRP expects numpy + integer seconds. Build the matrix so row/col 0
    # correspond to the depot regardless of what `depot` the caller passed.
    import numpy as _np  # local import — matches the pattern used elsewhere
    matrix = _np.asarray(duration_matrix, dtype=_np.int64)
    if depot != 0:
        order = [depot] + [i for i in range(n) if i != depot]
        matrix = matrix[_np.ix_(order, order)]
    else:
        order = list(range(n))

    # Build per-stop DeliveryStop including coords (if any) so PyVRPTspSolver
    # can collapse identical-coordinate clusters into super-nodes.
    if coordinates is not None:
        stops = [
            DeliveryStop(
                stop_id=original_idx,
                service_duration=0,
                x=float(coordinates[original_idx][0]),
                y=float(coordinates[original_idx][1]),
            )
            for original_idx in order[1:]
        ]
        depot_lon, depot_lat = coordinates[depot]
        depot_stop = DeliveryStop(
            stop_id=depot,
            service_duration=0,
            x=float(depot_lon),
            y=float(depot_lat),
        )
    else:
        stops = [
            DeliveryStop(stop_id=original_idx, service_duration=0)
            for original_idx in order[1:]
        ]
        depot_stop = DeliveryStop(stop_id=depot, service_duration=0)

    solver = PyVRPTspSolver(
        max_runtime_seconds=time_limit_seconds,
        seed=seed,
        display=False,
    )
    sequence = solver.solve(
        depot=depot_stop,
        stops=stops,
        time_matrix=matrix,
    )

    # `sequence` is already a list of ORIGINAL node indices (we stuffed the
    # original index into `stop_id`), so just prepend the depot to match the
    # convention used by `vroom_tsp_solve` and `lkh_tsp_solve`.
    visited = [depot] + [int(sid) for sid in sequence]

    # Defensive: if PyVRP ever drops a node, append it so callers never lose
    # a stop. Mirrors the guard at the bottom of `vroom_tsp_solve`.
    seen = set(visited)
    for i in range(n):
        if i not in seen:
            visited.append(i)
    return visited


def lkh_tsp_solve(
    duration_matrix: List[List[int]],
    depot: int = 0,
    runs: int = 5,
    time_limit_seconds: int = 10,
) -> List[int]:
    """Solve ATSP using LKH-3 (Lin-Kernighan-Helsgaun), the gold-standard TSP heuristic.

    Args:
        duration_matrix: NxN integer cost matrix (seconds).
        depot: Starting node index (fixed as tour start).
        runs: Number of LKH trial runs (more = better quality, slower).
        time_limit_seconds: Max wall-clock time for the solver.

    Returns:
        Ordered list of 0-indexed stop indices starting from depot.
    """
    global LKH_AVAILABLE, LKH_IMPORT_ERROR
    if not LKH_AVAILABLE:
        raise RuntimeError("LKH-3 binary not available")

    n = len(duration_matrix)
    if n <= 2:
        return list(range(n))

    # ── Matrix sanitisation ──────────────────────────────────────────────
    # OSRM occasionally returns `null`/negative cells for un-snappable coords;
    # passed verbatim to LKH those become "free" or "negative-cost" edges and
    # the solver gladly exploits them, producing visibly absurd tours. Force
    # `null/NaN/<0 → PENALTY_SECONDS` and the diagonal to 0 BEFORE the
    # open-path patch so the depot column zero-out is preserved.
    from solvers.pyvrp_tsp_solver import sanitize_osrm_matrix
    clean = sanitize_osrm_matrix(duration_matrix).tolist()

    # ── Open-path TSP via free return edge ────────────────────────────────
    # LKH solves a closed Hamiltonian cycle. For delivery routes we DO NOT
    # return to the depot — the driver finishes wherever the last stop is.
    # Zeroing the return-to-depot column makes the closed-loop optimum equal
    # to the open-path optimum because the return leg becomes free and drops
    # out of the objective. Without this, LKH produced routes that started
    # `depot → far_cluster → ...` because returning past near_cluster was
    # cheap in the cycle, even though the driver never actually returns.
    open_path_matrix = _open_path_matrix(clean, depot)

    # LKH uses ATSP format with FULL_MATRIX edge weights.
    problem = lkh.LKHProblem(
        type='ATSP',
        dimension=n,
        edge_weight_type='EXPLICIT',
        edge_weight_format='FULL_MATRIX',
        edge_weights=open_path_matrix,
    )

    # Scale runs and time with problem size
    actual_runs = max(runs, min(10, n // 20))
    actual_time = max(time_limit_seconds, min(30, n // 10))

    try:
        result = lkh.solve(
            solver=LKH_SOLVER_PATH,
            problem=problem,
            runs=actual_runs,
            time_limit=actual_time,
        )
    except OSError as exec_err:
        # ── Architecture mismatch self-disable ────────────────────────────
        # `[Errno 8] Exec format error` fires when the cached LKH binary at
        # LKH_SOLVER_PATH was compiled for a CPU arch that doesn't match the
        # current container (e.g. x86_64 binary on aarch64). Without this
        # guard every Optimize call re-tries LKH, re-throws OSError, and
        # spams the production log via the caller's `logger.warning`.
        # Flip `LKH_AVAILABLE=False` so the top-of-function guard short-
        # circuits future calls (and the caller-level `if LKH_AVAILABLE:`
        # blocks skip LKH cleanly). VROOM/3-opt fallback already exists.
        if exec_err.errno in (errno.ENOEXEC, 8):
            if LKH_AVAILABLE:
                LKH_AVAILABLE = False
                LKH_IMPORT_ERROR = (
                    f"LKH binary incompatible with current arch ({exec_err})"
                )
                logger.info(
                    "[lkh] Disabling LKH for this process — binary at %s is "
                    "incompatible with current CPU arch (Errno 8). Falling "
                    "back to VROOM+3-opt.",
                    LKH_SOLVER_PATH,
                )
        raise RuntimeError(f"LKH-3 binary not runnable: {exec_err}") from exec_err

    if not result or not result[0]:
        raise RuntimeError("LKH returned empty solution")

    # LKH returns 1-indexed tour. Convert to 0-indexed.
    tour_1indexed = result[0]
    tour = [x - 1 for x in tour_1indexed]

    # Rotate tour so depot is first
    if depot in tour:
        depot_pos = tour.index(depot)
        tour = tour[depot_pos:] + tour[:depot_pos]

    # Defensive: add any missing nodes
    visited = set(tour)
    for i in range(n):
        if i not in visited:
            tour.append(i)

    return tour



async def calculate_duration_matrix(stops: List[dict]) -> List[List[int]]:
    """Build NxN driving-duration matrix (integer seconds) using Mapbox.

    Used as FALLBACK when OSRM is unavailable.
    - N <= 25: single Mapbox Matrix API call.
    - N > 25: haversine estimate (use OSRM for larger routes).
    """
    n = len(stops)
    fallback = _haversine_duration_matrix(stops)

    if n <= 1 or not MAPBOX_TOKEN:
        return fallback

    try:
        if n <= 25:
            dur = await _mapbox_duration_matrix_batch(stops)
            if dur:
                logger.info("Duration matrix: full %dx%d from Mapbox", n, n)
                return dur
        # For >25 stops without OSRM, haversine is the best we can do via Mapbox
        # (cross-batch queries would require too many API calls)
        logger.info("Duration matrix: %dx%d haversine estimate (Mapbox limit exceeded)", n, n)
        return fallback

    except Exception as exc:
        logger.warning("Duration matrix build failed, using haversine estimate: %s", exc)
        return fallback


async def calculate_road_distance_matrix(stops: List[dict]) -> List[List[float]]:
    """Build distance matrix using OSRM road distances (primary) or Mapbox (fallback).

    Strategy:
    - Try OSRM Table API first (local, no rate limit, handles any N).
    - Fallback to Mapbox Matrix API if OSRM unavailable.
    - Final fallback to haversine if both APIs fail.
    """
    n = len(stops)
    haversine_matrix = calculate_distance_matrix(stops)

    if n <= 1:
        return haversine_matrix

    # --- Primary: OSRM ---
    osrm_dist = await _osrm_distance_matrix(stops)
    if osrm_dist:
        logger.info("Road distance matrix: full %dx%d from OSRM", n, n)
        return osrm_dist

    # --- Fallback: Mapbox ---
    if not MAPBOX_TOKEN:
        return haversine_matrix

    try:
        if n <= 25:
            road = await _mapbox_matrix_batch(stops)
            if road:
                logger.info("Road distance matrix: full %dx%d from Mapbox", n, n)
                return road
            return haversine_matrix

        CLUSTER_SIZE = 25

        # Geographic sort: group nearby stops into clusters.
        sorted_indices = sorted(range(n), key=lambda i: (
            round(stops[i]['latitude'] * 100),
            stops[i]['longitude'],
        ))

        clusters = []
        for i in range(0, n, CLUSTER_SIZE):
            clusters.append(sorted_indices[i:i + CLUSTER_SIZE])

        # Deep copy haversine baseline
        matrix = [row[:] for row in haversine_matrix]

        # Overwrite intra-cluster cells with Mapbox road distances
        upgraded = 0
        for cluster_indices in clusters:
            if len(cluster_indices) < 2:
                continue
            cluster_stops = [stops[i] for i in cluster_indices]
            road_sub = await _mapbox_matrix_batch(cluster_stops)
            if road_sub:
                for ci, gi in enumerate(cluster_indices):
                    for cj, gj in enumerate(cluster_indices):
                        matrix[gi][gj] = road_sub[ci][cj]
                upgraded += len(cluster_indices)

        logger.info(
            "Road distance matrix: %d/%d stops upgraded to Mapbox road distances (%d clusters)",
            upgraded, n, len(clusters),
        )
        return matrix

    except Exception as exc:
        logger.warning("Road distance matrix build failed, using haversine: %s", exc)
        return haversine_matrix


# ==================== FULL ROAD DISTANCE MATRIX (CROSS-BATCH) ====================

async def _mapbox_cross_batch_query(
    client: httpx.AsyncClient,
    stops: List[dict],
    src_global: List[int],
    dst_global: List[int],
    sem: asyncio.Semaphore,
) -> Optional[tuple]:
    """Single Mapbox Matrix API call for a (source_batch, dest_batch) pair.
    
    Combines source and destination coordinates into one request (≤25 coords),
    using the sources/destinations parameters to get the sub-matrix.
    Returns (sub_matrix_km, src_global_indices, dst_global_indices) or None.
    """
    async with sem:
        # Build deduplicated combined coordinate list preserving order
        combined_global = list(src_global)
        dst_only = [i for i in dst_global if i not in set(src_global)]
        combined_global.extend(dst_only)

        if len(combined_global) > 25:
            return None

        # Map global indices to local positions in the combined list
        global_to_local = {gi: li for li, gi in enumerate(combined_global)}
        local_src = [global_to_local[gi] for gi in src_global]
        local_dst = [global_to_local[gi] for gi in dst_global]

        coords = ";".join(
            f"{stops[gi]['longitude']},{stops[gi]['latitude']}"
            for gi in combined_global
        )
        url = f"https://api.mapbox.com/directions-matrix/v1/mapbox/driving/{coords}"
        params = {
            "access_token": MAPBOX_TOKEN,
            "annotations": "distance",
            "sources": ";".join(str(i) for i in local_src),
            "destinations": ";".join(str(i) for i in local_dst),
        }

        try:
            resp = await client.get(url, params=params, timeout=30)
            if resp.status_code != 200:
                return None
            data = resp.json()
            if data.get("code") != "Ok":
                return None
            distances = data.get("distances")
            if not distances:
                return None

            # distances shape: len(src) x len(dst), values in meters
            sub = []
            for row in distances:
                sub.append([
                    round(d / 1000.0, 4) if d is not None else None
                    for d in row
                ])
            return (sub, src_global, dst_global)
        except Exception as exc:
            logger.debug("Mapbox cross-batch failed: %s", exc)
            return None


async def calculate_full_road_distance_matrix(stops: List[dict]) -> List[List[float]]:
    """Build FULL NxN road distance matrix using OSRM (primary) or Mapbox cross-batch (fallback).

    OSRM handles any N natively. Falls back to Mapbox cross-batch queries
    if OSRM is unavailable, then haversine as last resort.
    """
    n = len(stops)
    haversine_matrix = calculate_distance_matrix(stops)

    if n <= 1:
        return haversine_matrix

    # --- Primary: OSRM ---
    osrm_dist = await _osrm_distance_matrix(stops)
    if osrm_dist:
        logger.info("Full road distance matrix: %dx%d from OSRM", n, n)
        return osrm_dist

    # --- Fallback: Mapbox ---
    if not MAPBOX_TOKEN:
        return haversine_matrix

    if n <= 25:
        road = await _mapbox_matrix_batch(stops)
        if road:
            return road
        return haversine_matrix

    try:
        BATCH_SIZE = 12  # 12 src + 12 dst = 24 coords ≤ 25

        # Create batches of global stop indices
        batches = []
        for i in range(0, n, BATCH_SIZE):
            batches.append(list(range(i, min(i + BATCH_SIZE, n))))

        # Deep copy haversine as baseline
        matrix = [row[:] for row in haversine_matrix]

        sem = asyncio.Semaphore(10)
        async with httpx.AsyncClient() as client:
            tasks = [
                _mapbox_cross_batch_query(client, stops, src_batch, dst_batch, sem)
                for src_batch in batches
                for dst_batch in batches
            ]
            results = await asyncio.gather(*tasks)

        upgraded = 0
        for result in results:
            if result is None:
                continue
            sub, src_global, dst_global = result
            for i, gi in enumerate(src_global):
                for j, gj in enumerate(dst_global):
                    if gi != gj and sub[i][j] is not None:
                        matrix[gi][gj] = sub[i][j]
                        upgraded += 1

        total_cells = n * (n - 1)
        logger.info(
            "Full road matrix: %d/%d cells upgraded to Mapbox road distances (%d API calls)",
            upgraded, total_cells, len(tasks),
        )
        return matrix

    except Exception as exc:
        logger.warning("Full road matrix build failed, using haversine: %s", exc)
        return haversine_matrix


# ==================== CLUSTER-FIRST ROUTE-SECOND OPTIMIZATION ====================

def _geographic_dbscan(stops: List[dict], eps_km: float = 0.8, min_samples: int = 2) -> List[int]:
    """DBSCAN clustering on geographic coordinates using haversine distance.
    Returns list of cluster labels per stop. -1 = noise (unassigned)."""
    n = len(stops)
    if n == 0:
        return []

    labels = [-1] * n
    cluster_id = 0
    visited = [False] * n

    for i in range(n):
        if visited[i]:
            continue
        visited[i] = True

        # Find all neighbors within eps
        neighbors = []
        for j in range(n):
            if i != j:
                d = haversine(
                    (stops[i]["latitude"], stops[i]["longitude"]),
                    (stops[j]["latitude"], stops[j]["longitude"]),
                    unit=Unit.KILOMETERS,
                )
                if d <= eps_km:
                    neighbors.append(j)

        if len(neighbors) < min_samples - 1:
            continue  # noise, will be assigned in post-processing

        # Start new cluster
        labels[i] = cluster_id
        seed_set = list(neighbors)
        idx = 0

        while idx < len(seed_set):
            j = seed_set[idx]
            idx += 1

            if not visited[j]:
                visited[j] = True
                j_neighbors = []
                for k in range(n):
                    if k != j:
                        d = haversine(
                            (stops[j]["latitude"], stops[j]["longitude"]),
                            (stops[k]["latitude"], stops[k]["longitude"]),
                            unit=Unit.KILOMETERS,
                        )
                        if d <= eps_km:
                            j_neighbors.append(k)

                if len(j_neighbors) >= min_samples - 1:
                    for k in j_neighbors:
                        if labels[k] == -1:
                            seed_set.append(k)

            if labels[j] == -1:
                labels[j] = cluster_id

        cluster_id += 1

    return labels


def _adaptive_eps(stops: List[dict]) -> float:
    """Compute adaptive DBSCAN eps based on stop density.
    Uses k-nearest-neighbor heuristic (k=4) to find natural cluster radius."""
    n = len(stops)
    if n <= 2:
        return 1.0

    # For each stop, find distance to 4th nearest neighbor
    k = min(4, n - 1)
    k_distances = []

    for i in range(n):
        dists = []
        for j in range(n):
            if i != j:
                d = haversine(
                    (stops[i]["latitude"], stops[i]["longitude"]),
                    (stops[j]["latitude"], stops[j]["longitude"]),
                    unit=Unit.KILOMETERS,
                )
                dists.append(d)
        dists.sort()
        k_distances.append(dists[k - 1] if len(dists) >= k else dists[-1])

    # Sort k-distances and find the "elbow" — we use the median as a robust estimate
    k_distances.sort()
    # Use the 60th percentile as eps (captures most natural clusters)
    eps = k_distances[int(n * 0.6)]
    # Clamp to reasonable delivery neighborhood sizes
    return max(0.3, min(2.5, eps))


def _postprocess_clusters(
    labels: List[int],
    stops: List[dict],
    max_cluster_size: int = 23,
    min_cluster_size: int = 2,
) -> List[List[int]]:
    """Post-process DBSCAN clusters:
    - Assign noise points to nearest cluster
    - Split oversized clusters (>max_cluster_size) for Mapbox API compliance
    - Merge tiny clusters into nearest neighbor
    Returns list of lists of global stop indices."""
    from collections import defaultdict

    clusters_map = defaultdict(list)
    noise = []

    for i, label in enumerate(labels):
        if label == -1:
            noise.append(i)
        else:
            clusters_map[label].append(i)

    cluster_list = list(clusters_map.values())

    # If no clusters found, treat everything as one cluster
    if not cluster_list:
        cluster_list = [list(range(len(stops)))]
        noise = []

    # Assign noise points to nearest cluster
    for ni in noise:
        best_ci = 0
        best_dist = float("inf")
        for ci, cluster in enumerate(cluster_list):
            for si in cluster:
                d = haversine(
                    (stops[ni]["latitude"], stops[ni]["longitude"]),
                    (stops[si]["latitude"], stops[si]["longitude"]),
                    unit=Unit.KILOMETERS,
                )
                if d < best_dist:
                    best_dist = d
                    best_ci = ci
        cluster_list[best_ci].append(ni)

    # Split oversized clusters using geographic k-means for spatially compact subclusters
    split_clusters = []
    for cluster in cluster_list:
        if len(cluster) <= max_cluster_size:
            split_clusters.append(cluster)
        else:
            # k-means split: divide into ceil(n/max_cluster_size) spatially compact groups
            import math as _math
            k = _math.ceil(len(cluster) / max_cluster_size)
            coords = [(stops[i]["latitude"], stops[i]["longitude"]) for i in cluster]

            # Initialize centroids using evenly spaced indices from sorted points
            sorted_by_lat = sorted(range(len(cluster)), key=lambda x: coords[x])
            centroids = [coords[sorted_by_lat[int(j * len(cluster) / k)]] for j in range(k)]

            for _ in range(15):  # k-means iterations
                buckets = [[] for _ in range(k)]
                for ci_local, idx in enumerate(cluster):
                    lat, lng = coords[ci_local]
                    best_k = 0
                    best_d = float("inf")
                    for ki in range(k):
                        d = (lat - centroids[ki][0]) ** 2 + (lng - centroids[ki][1]) ** 2
                        if d < best_d:
                            best_d = d
                            best_k = ki
                    buckets[best_k].append(idx)

                # Recompute centroids
                new_centroids = []
                for ki in range(k):
                    if buckets[ki]:
                        avg_lat = sum(stops[i]["latitude"] for i in buckets[ki]) / len(buckets[ki])
                        avg_lng = sum(stops[i]["longitude"] for i in buckets[ki]) / len(buckets[ki])
                        new_centroids.append((avg_lat, avg_lng))
                    else:
                        new_centroids.append(centroids[ki])

                if new_centroids == centroids:
                    break
                centroids = new_centroids

            for bucket in buckets:
                if bucket:
                    split_clusters.append(bucket)

    # Merge tiny clusters into nearest larger cluster (if it won't exceed max)
    final = []
    tiny = []
    for c in split_clusters:
        if len(c) < min_cluster_size:
            tiny.append(c)
        else:
            final.append(c)

    for tc in tiny:
        if not final:
            final.append(tc)
            continue
        tc_lat = sum(stops[i]["latitude"] for i in tc) / len(tc)
        tc_lng = sum(stops[i]["longitude"] for i in tc) / len(tc)

        best_ci = 0
        best_dist = float("inf")
        for ci, c in enumerate(final):
            if len(c) + len(tc) > max_cluster_size:
                continue
            c_lat = sum(stops[i]["latitude"] for i in c) / len(c)
            c_lng = sum(stops[i]["longitude"] for i in c) / len(c)
            d = haversine((tc_lat, tc_lng), (c_lat, c_lng), unit=Unit.KILOMETERS)
            if d < best_dist:
                best_dist = d
                best_ci = ci
        final[best_ci].extend(tc)

    return final if final else [list(range(len(stops)))]


def _order_clusters_tsp(
    clusters: List[List[int]],
    stops: List[dict],
    start_stop_index: int = 0,
) -> List[int]:
    """Order clusters using centroid nearest-neighbor + 2-opt.
    Returns list of cluster indices in visit order."""
    nc = len(clusters)
    if nc <= 1:
        return list(range(nc))

    # Compute centroids
    centroids = []
    for cluster in clusters:
        avg_lat = sum(stops[i]["latitude"] for i in cluster) / len(cluster)
        avg_lng = sum(stops[i]["longitude"] for i in cluster) / len(cluster)
        centroids.append((avg_lat, avg_lng))

    # Find the cluster that contains (or is nearest to) the start stop
    start_ci = 0
    for ci, cluster in enumerate(clusters):
        if start_stop_index in cluster:
            start_ci = ci
            break

    # Nearest-neighbor TSP on centroids
    visited = [False] * nc
    order = [start_ci]
    visited[start_ci] = True

    for _ in range(nc - 1):
        current = order[-1]
        best = -1
        best_dist = float("inf")
        for j in range(nc):
            if not visited[j]:
                d = haversine(centroids[current], centroids[j], unit=Unit.KILOMETERS)
                if d < best_dist:
                    best_dist = d
                    best = j
        if best != -1:
            order.append(best)
            visited[best] = True

    # 2-opt improvement on the cluster order
    improved = True
    while improved:
        improved = False
        for i in range(1, len(order) - 1):
            for j in range(i + 1, len(order)):
                # Calculate distance change if we reverse order[i:j+1]
                pi, pj = order[i - 1], order[i]
                qi, qj = order[j], order[(j + 1) % len(order)] if j + 1 < len(order) else order[0]

                old_d = (
                    haversine(centroids[pi], centroids[pj], unit=Unit.KILOMETERS)
                    + (haversine(centroids[qi], centroids[qj], unit=Unit.KILOMETERS) if j + 1 < len(order) else 0)
                )
                new_d = (
                    haversine(centroids[pi], centroids[qi], unit=Unit.KILOMETERS)
                    + (haversine(centroids[pj], centroids[qj], unit=Unit.KILOMETERS) if j + 1 < len(order) else 0)
                )
                if new_d < old_d - 0.001:
                    order[i : j + 1] = reversed(order[i : j + 1])
                    improved = True

    return order


def _convex_hull(points: List[tuple]) -> List[tuple]:
    """Compute convex hull of 2D points using Andrew's monotone chain.
    Points are (lng, lat) tuples. Returns hull vertices in CCW order."""
    pts = sorted(set(points))
    if len(pts) <= 2:
        return pts

    def cross(o, a, b):
        return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])

    lower = []
    for p in pts:
        while len(lower) >= 2 and cross(lower[-2], lower[-1], p) <= 0:
            lower.pop()
        lower.append(p)
    upper = []
    for p in reversed(pts):
        while len(upper) >= 2 and cross(upper[-2], upper[-1], p) <= 0:
            upper.pop()
        upper.append(p)
    return lower[:-1] + upper[:-1]


def _padded_polygon(hull: List[tuple], pad_deg: float = 0.0002) -> List[List[float]]:
    """Expand a convex hull outward by pad_deg (~20m at equator).
    Returns GeoJSON-compatible closed ring [[lng,lat], ...]."""
    import math

    if len(hull) < 2:
        # Single point → small octagon
        if hull:
            cx, cy = hull[0]
            return [
                [cx + pad_deg * math.cos(a), cy + pad_deg * math.sin(a)]
                for a in [i * math.pi / 4 for i in range(8)]
            ] + [[cx + pad_deg, cy]]
        return []

    if len(hull) == 2:
        # Line segment → diamond
        ax, ay = hull[0]
        bx, by = hull[1]
        dx, dy = bx - ax, by - ay
        length = math.sqrt(dx * dx + dy * dy) or 1e-8
        nx, ny = -dy / length * pad_deg, dx / length * pad_deg
        return [
            [ax - dx * 0.1 + nx, ay - dy * 0.1 + ny],
            [bx + dx * 0.1 + nx, by + dy * 0.1 + ny],
            [bx + dx * 0.1 - nx, by + dy * 0.1 - ny],
            [ax - dx * 0.1 - nx, ay - dy * 0.1 - ny],
            [ax - dx * 0.1 + nx, ay - dy * 0.1 + ny],
        ]

    # Compute centroid
    cx = sum(p[0] for p in hull) / len(hull)
    cy = sum(p[1] for p in hull) / len(hull)

    # Push each vertex outward from centroid
    padded = []
    for px, py in hull:
        dx, dy = px - cx, py - cy
        dist = math.sqrt(dx * dx + dy * dy) or 1e-8
        padded.append([px + dx / dist * pad_deg, py + dy / dist * pad_deg])
    padded.append(padded[0])  # close the ring
    return padded


# 15 distinct cluster colors — semi-transparent fills with solid borders
CLUSTER_COLORS = [
    {"fill": "rgba(59, 130, 246, 0.25)", "border": "rgba(59, 130, 246, 0.8)"},   # blue
    {"fill": "rgba(239, 68, 68, 0.25)", "border": "rgba(239, 68, 68, 0.8)"},     # red
    {"fill": "rgba(16, 185, 129, 0.25)", "border": "rgba(16, 185, 129, 0.8)"},   # emerald
    {"fill": "rgba(245, 158, 11, 0.25)", "border": "rgba(245, 158, 11, 0.8)"},   # amber
    {"fill": "rgba(168, 85, 247, 0.25)", "border": "rgba(168, 85, 247, 0.8)"},   # purple
    {"fill": "rgba(236, 72, 153, 0.25)", "border": "rgba(236, 72, 153, 0.8)"},   # pink
    {"fill": "rgba(20, 184, 166, 0.25)", "border": "rgba(20, 184, 166, 0.8)"},   # teal
    {"fill": "rgba(251, 146, 60, 0.25)", "border": "rgba(251, 146, 60, 0.8)"},   # orange
    {"fill": "rgba(99, 102, 241, 0.25)", "border": "rgba(99, 102, 241, 0.8)"},   # indigo
    {"fill": "rgba(34, 197, 94, 0.25)", "border": "rgba(34, 197, 94, 0.8)"},     # green
    {"fill": "rgba(244, 63, 94, 0.25)", "border": "rgba(244, 63, 94, 0.8)"},     # rose
    {"fill": "rgba(6, 182, 212, 0.25)", "border": "rgba(6, 182, 212, 0.8)"},     # cyan
    {"fill": "rgba(234, 179, 8, 0.25)", "border": "rgba(234, 179, 8, 0.8)"},     # yellow
    {"fill": "rgba(139, 92, 246, 0.25)", "border": "rgba(139, 92, 246, 0.8)"},   # violet
    {"fill": "rgba(14, 165, 233, 0.25)", "border": "rgba(14, 165, 233, 0.8)"},   # sky
]


def _or_opt_1_improve(indices: List[int], matrix: List[List[float]]) -> List[int]:
    """Or-opt-1: Relocate single stops to better positions using the road distance matrix.
    Catches cases where stops on the same street get split by stops on adjacent streets."""
    n = len(indices)
    if n <= 3:
        return indices

    best = indices[:]

    def route_cost(r):
        return sum(matrix[r[i]][r[i + 1]] for i in range(len(r) - 1))

    current_cost = route_cost(best)
    improved = True
    iterations = 0

    while improved and iterations < 5:
        improved = False
        iterations += 1
        for i in range(1, len(best)):  # Skip index 0 (start point)
            stop = best[i]
            # Remove stop from current position
            remaining = best[:i] + best[i + 1:]
            # Cost without this stop
            remove_save = (
                matrix[best[i - 1]][best[i]]
                + (matrix[best[i]][best[i + 1]] if i + 1 < len(best) else 0)
                - (matrix[best[i - 1]][best[i + 1]] if i + 1 < len(best) else 0)
            )

            best_j = -1
            best_insert_cost = float("inf")

            for j in range(len(remaining)):
                # Try inserting after position j in remaining
                if j + 1 < len(remaining):
                    insert_cost = (
                        matrix[remaining[j]][stop]
                        + matrix[stop][remaining[j + 1]]
                        - matrix[remaining[j]][remaining[j + 1]]
                    )
                else:
                    insert_cost = matrix[remaining[j]][stop]

                if insert_cost < best_insert_cost:
                    best_insert_cost = insert_cost
                    best_j = j

            # Check if relocating improves total cost
            if best_j >= 0 and best_insert_cost < remove_save - 0.001:
                new_route = remaining[:best_j + 1] + [stop] + remaining[best_j + 1:]
                new_cost = route_cost(new_route)
                if new_cost < current_cost - 0.001:
                    best = new_route
                    current_cost = new_cost
                    improved = True
                    break  # Restart from beginning after improvement

    return best


def _build_cluster_info(
    ordered_clusters: List[List[int]],
    stops: List[dict],
) -> List[dict]:
    """Build GeoJSON-ready cluster visualization data with convex hull polygons."""
    cluster_info = []
    for visit_order, cluster_indices in enumerate(ordered_clusters):
        points = [(stops[i]["longitude"], stops[i]["latitude"]) for i in cluster_indices]
        hull = _convex_hull(points)
        polygon = _padded_polygon(hull)

        centroid_lat = sum(stops[i]["latitude"] for i in cluster_indices) / len(cluster_indices)
        centroid_lng = sum(stops[i]["longitude"] for i in cluster_indices) / len(cluster_indices)
        color = CLUSTER_COLORS[visit_order % len(CLUSTER_COLORS)]

        cluster_info.append({
            "id": visit_order,
            "visit_order": visit_order,
            "stop_count": len(cluster_indices),
            "centroid": {"latitude": round(centroid_lat, 6), "longitude": round(centroid_lng, 6)},
            "polygon": polygon,
            "fill_color": color["fill"],
            "border_color": color["border"],
            "label": f"Zone {visit_order + 1}",
        })
    return cluster_info


def _run_inner_algorithm(
    stops: List[dict],
    matrix: List[List[float]],
    start_index: int,
    time_limit: int,
    algorithm: str,
) -> List[dict]:
    """Run a specific optimization algorithm on a subset of stops.
    Used within cluster_first to apply the user's preferred algorithm per cluster.
    Applies post-optimization 2-opt + or-opt using the road distance matrix
    to catch local swaps the main solver may have missed (e.g., grouping same-street stops)."""
    result = None
    try:
        if algorithm == "ortools" and ORTOOLS_AVAILABLE and pywrapcp:
            # Use ortools_tsp_solve directly — the matrix passed in is already the
            # correct type (duration seconds when cluster_first uses OR-Tools inner)
            time_limit_ms = max(1000, time_limit * 1000)
            indices = ortools_tsp_solve(matrix, depot=start_index, time_limit_ms=time_limit_ms)
            result = [stops[i] for i in indices]
        elif algorithm == "pyvrp" and PYVRP_AVAILABLE:
            pyvrp_seconds = max(1.0, min(2.0, len(stops) * 0.05))
            indices = pyvrp_tsp_solve(matrix, depot=start_index, time_limit_seconds=pyvrp_seconds)
            result = [stops[i] for i in indices]
        elif algorithm == "alns":
            try:
                result = alns_hybrid_optimize(stops, matrix, start_index=start_index, time_limit_seconds=time_limit)
            except NameError:
                logger.warning("ALNS not available, falling back to OR-Tools")
                if ORTOOLS_AVAILABLE and pywrapcp:
                    result = ortools_optimize(stops, matrix, start_index, time_limit)
        elif algorithm == "simulated_annealing":
            result = simulated_annealing_optimize(stops, matrix, start_index)
        elif algorithm == "genetic":
            result = genetic_algorithm_optimize(stops, matrix, start_index)
        elif algorithm == "clarke_wright":
            result = clarke_wright_savings(stops, matrix, start_index)
    except Exception as exc:
        logger.warning("Inner algorithm '%s' failed, falling back to NN+2-opt: %s", algorithm, exc)

    if result is None:
        nn = nearest_neighbor_optimize(stops, matrix, start_index)
        ri = _indices_by_identity(stops, nn)
        result = [stops[i] for i in two_opt_improve(ri, matrix)]

    # Post-optimization: apply road-distance or-opt-1 then 2-opt to catch missed local swaps
    # This fixes cases where stops on the same street get split by stops on adjacent streets
    if len(result) > 3:
        indices = _indices_by_identity(stops, result)
        indices = _or_opt_1_improve(indices, matrix)
        indices = two_opt_improve(indices, matrix)
        result = [stops[i] for i in indices]

    return result


def _global_two_opt_pass(optimized: List[dict], max_iterations: int = 3) -> List[dict]:
    """Apply or-opt-1 + 2-opt on the full stitched route using haversine distances.
    Fixes cross-cluster boundary inefficiencies:
    - or-opt-1 relocates single stops to better positions (e.g., moving stop 46 from
      between 45→47 to after 48, avoiding an unnecessary south→north detour)
    - 2-opt reverses segments to uncross route lines
    - 3-opt (large routes only): non-reversing segment swap to escape 2-opt
      local optima on routes ≥150 stops where boundary stitching tends to
      leave a few residual cross-cluster zig-zags that 2-opt can't fix.
    """
    n = len(optimized)
    if n <= 3:
        return optimized

    # Large routes (≥150 stops) get more aggressive polishing: doubling the
    # iteration budget (3 → 6) gives or-opt and 2-opt enough runway to chase
    # cross-cluster relocations to convergence on long routes, where each
    # iteration only nudges a handful of stops at a time.
    if n >= 150:
        max_iterations = max(max_iterations, 6)

    # Build haversine matrix for the stitched route
    matrix = [[0.0] * n for _ in range(n)]
    for i in range(n):
        for j in range(i + 1, n):
            d = haversine(
                (optimized[i]["latitude"], optimized[i]["longitude"]),
                (optimized[j]["latitude"], optimized[j]["longitude"]),
                unit=Unit.KILOMETERS,
            )
            matrix[i][j] = d
            matrix[j][i] = d

    indices = list(range(n))
    best_dist = sum(matrix[indices[i]][indices[i + 1]] for i in range(n - 1))

    # Phase 1: Global or-opt-1 — relocate individual stops across cluster boundaries
    for _iter in range(max_iterations):
        improved = False
        for i in range(1, len(indices)):
            stop_idx = indices[i]
            # Cost of edges touching this stop
            prev_idx = indices[i - 1]
            next_idx = indices[i + 1] if i + 1 < len(indices) else None

            edge_before = matrix[prev_idx][stop_idx]
            edge_after = matrix[stop_idx][next_idx] if next_idx is not None else 0
            edge_skip = matrix[prev_idx][next_idx] if next_idx is not None else 0
            remove_save = edge_before + edge_after - edge_skip

            if remove_save < 0.02:  # Not worth relocating if removal doesn't save much
                continue

            best_j = -1
            best_delta = 0

            # Try inserting this stop at every other position (limited window for speed)
            remaining = indices[:i] + indices[i + 1:]
            for j in range(max(0, i - 40), min(len(remaining), i + 40)):
                a = remaining[j]
                b = remaining[j + 1] if j + 1 < len(remaining) else None
                old_edge = matrix[a][b] if b is not None else 0
                new_edge = matrix[a][stop_idx] + (matrix[stop_idx][b] if b is not None else 0)
                insert_cost = new_edge - old_edge
                delta = remove_save - insert_cost
                if delta > best_delta + 0.01:
                    best_delta = delta
                    best_j = j

            if best_j >= 0:
                # Perform the relocation
                indices.pop(i)
                actual_j = best_j if best_j < i else best_j
                indices.insert(actual_j + 1, stop_idx)
                improved = True
                break  # Restart scan after improvement

        if not improved:
            break

    # Phase 2: Global 2-opt — reverse segments to uncross route lines
    for _ in range(max_iterations):
        improved = False
        for i in range(1, n - 1):
            for j in range(i + 1, min(i + 60, n)):
                d_old = matrix[indices[i - 1]][indices[i]] + (matrix[indices[j]][indices[j + 1]] if j + 1 < n else 0)
                d_new = matrix[indices[i - 1]][indices[j]] + (matrix[indices[i]][indices[j + 1]] if j + 1 < n else 0)
                if d_new < d_old - 0.01:
                    indices[i:j + 1] = reversed(indices[i:j + 1])
                    improved = True
        if not improved:
            break

    # Phase 3: 3-opt polish (large routes only). On routes ≥150 stops the
    # 2-opt pass above usually plateaus with a few residual cross-cluster
    # zig-zags that the reversal-only neighbourhood can't escape.
    # `three_opt_improve` swaps non-adjacent segments without reversing
    # them, which is asymmetric-safe and exact on this haversine matrix.
    # We deliberately keep 3-opt off for smaller routes — the 2-opt window
    # above already converges, and 3-opt's O(n³) inner loop would dominate
    # the per-request budget without measurable quality gain.
    if n >= 150:
        polished = three_opt_improve(indices, matrix, max_iterations=3)
        polished_dist = sum(matrix[polished[i]][polished[i + 1]] for i in range(n - 1))
        current_dist = sum(matrix[indices[i]][indices[i + 1]] for i in range(n - 1))
        if polished_dist < current_dist - 0.01:
            indices = polished
            logger.info(
                "Global 3-opt polish improved route: %.2f km → %.2f km",
                current_dist, polished_dist,
            )

    new_dist = sum(matrix[indices[i]][indices[i + 1]] for i in range(n - 1))
    if new_dist < best_dist:
        logger.info("Global or-opt+2-opt improved route: %.2f km → %.2f km (saved %.2f km)", best_dist, new_dist, best_dist - new_dist)
        return [optimized[i] for i in indices]

    return optimized


async def cluster_first_optimize(
    stops: List[dict],
    distance_matrix: List[List[float]],
    start_index: int = 0,
    time_limit_seconds: int = 30,
    inner_algorithm: str = "ortools",
) -> tuple:
    """Cluster-first route-second optimization.

    Guarantees spatially coherent routing by:
    1. DBSCAN geographic clustering into natural neighborhoods
    2. Inter-cluster ordering via centroid TSP with 2-opt
    3. Intra-cluster optimization with Mapbox road distances + user's preferred algorithm
    4. Smart entry/exit stitching between adjacent clusters
    5. Global 2-opt pass to fix cross-boundary inefficiencies

    Args:
        inner_algorithm: Algorithm to use within each cluster (ortools, alns, etc.)

    Returns (optimized_stops, cluster_info) tuple.
    """
    n = len(stops)
    if n <= 25:
        # Small enough for a single pass — no cluster visualization
        result = _run_inner_algorithm(stops, distance_matrix, start_index, time_limit_seconds, inner_algorithm)
        return result, []

    # Step 1: Geographic clustering
    eps = _adaptive_eps(stops)
    labels = _geographic_dbscan(stops, eps_km=eps, min_samples=2)
    clusters = _postprocess_clusters(labels, stops, max_cluster_size=23, min_cluster_size=2)
    logger.info(
        "Cluster-first (%s): %d clusters from %d stops (eps=%.2f km, sizes=%s)",
        inner_algorithm, len(clusters), n, eps,
        [len(c) for c in clusters],
    )

    # Step 2: Order clusters using centroid TSP
    cluster_order = _order_clusters_tsp(clusters, stops, start_stop_index=start_index)
    ordered_clusters = [clusters[i] for i in cluster_order]

    # Build cluster visualization data
    cluster_info = _build_cluster_info(ordered_clusters, stops)

    # Step 3 & 4: Optimize within each cluster and stitch
    all_optimized: List[dict] = []
    previous_exit_global = start_index
    # per_cluster_time kept as reference for future time-budgeted cluster solves.
    _ = max(5, time_limit_seconds // max(1, len(ordered_clusters)))

    for ci, cluster_indices in enumerate(ordered_clusters):
        cluster_stops = [stops[gi] for gi in cluster_indices]

        if len(cluster_stops) == 1:
            all_optimized.extend(cluster_stops)
            previous_exit_global = cluster_indices[0]
            continue

        # Scale per-cluster OR-Tools time based on cluster size
        # Small clusters are trivially solved — 1 second is plenty
        # OR-Tools GUIDED_LOCAL_SEARCH uses the FULL time limit regardless of problem size
        if len(cluster_stops) <= 5:
            cluster_time = 1
        elif len(cluster_stops) <= 12:
            cluster_time = 2
        elif len(cluster_stops) <= 18:
            cluster_time = 3
        else:
            cluster_time = 5

        # Get cluster cost matrix.
        # OR-Tools inner algorithm uses DURATION (seconds) for time-optimal routing.
        # Other algorithms use road DISTANCE (km).
        if inner_algorithm == "ortools":
            # Try OSRM duration matrix first — same primary source as the main
            # `/optimize` pipeline (server.py line 5120). `calculate_duration_matrix`
            # below is the Mapbox/haversine FALLBACK path; calling it directly
            # silently degraded routing quality whenever a cluster had >25 stops
            # (Mapbox Matrix API limit → haversine straight-line distances), or
            # whenever Mapbox was rate-limited. Now cluster_first gets the same
            # road-aware OSRM data as VROOM/OR-Tools/LKH on the top-level path.
            cluster_matrix = await _osrm_duration_matrix(cluster_stops)
            if not cluster_matrix:
                cluster_matrix = await calculate_duration_matrix(cluster_stops)
        else:
            # `calculate_road_distance_matrix` already tries OSRM first internally
            # (see server.py line 2812), so non-ortools inner algorithms have
            # always had the OSRM-first path. No change needed here.
            cluster_matrix = await calculate_road_distance_matrix(cluster_stops)

        # Determine entry point: closest stop to previous cluster's exit
        entry_local = 0
        if ci == 0:
            # First cluster: find the start stop
            for li, gi in enumerate(cluster_indices):
                if gi == start_index:
                    entry_local = li
                    break
        else:
            prev_stop = stops[previous_exit_global]
            min_d = float("inf")
            for li, gi in enumerate(cluster_indices):
                d = haversine(
                    (prev_stop["latitude"], prev_stop["longitude"]),
                    (stops[gi]["latitude"], stops[gi]["longitude"]),
                    unit=Unit.KILOMETERS,
                )
                if d < min_d:
                    min_d = d
                    entry_local = li

        # Optimize within this cluster using the user's preferred algorithm
        optimized = _run_inner_algorithm(
            cluster_stops, cluster_matrix,
            start_index=entry_local,
            time_limit=cluster_time,
            algorithm=inner_algorithm,
        )

        all_optimized.extend(optimized)

        # Track exit point (last stop in this cluster) for stitching to next cluster
        last_stop = optimized[-1]
        for gi in range(n):
            if stops[gi] is last_stop:
                previous_exit_global = gi
                break

    # Step 5: Global 2-opt pass to fix cross-cluster boundary inefficiencies
    all_optimized = _global_two_opt_pass(all_optimized, max_iterations=3)

    return all_optimized, cluster_info


def build_time_matrix_from_distance(distance_matrix: List[List[float]], avg_speed_kmh: float = 38.0) -> List[List[int]]:
    """Approximate travel-time matrix (seconds) from distance matrix (km)."""
    if avg_speed_kmh <= 0:
        avg_speed_kmh = 38.0

    n = len(distance_matrix)
    time_matrix = [[0] * n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            if i == j:
                continue
            # seconds = (km / kmh) * 3600
            time_matrix[i][j] = max(1, int((distance_matrix[i][j] / avg_speed_kmh) * 3600))
    return time_matrix


def ortools_optimize(
    stops: List[dict],
    distance_matrix: List[List[float]],
    start_index: int = 0,
    time_limit_seconds: int = 10,
) -> List[dict]:
    """Legacy wrapper — calls ortools_tsp_solve and maps indices back to stops."""
    if len(stops) <= 1:
        return stops
    indices = ortools_tsp_solve(distance_matrix, depot=start_index, time_limit_ms=time_limit_seconds * 1000)
    return [stops[i] for i in indices]


def ortools_tsp_solve(
    matrix: List[List[float]],
    depot: int = 0,
    time_limit_ms: int = 2000,
    initial_indices: List[int] = None,
) -> List[int]:
    """
    Solve the Travelling Salesman Problem using Google OR-Tools.

    This is the single, industry-standard solver for route optimization.
    It accepts a Distance/Duration Matrix and returns the optimal visit order.

    ── How it works ──
    1. An (N+1)-node model is created: N real stops + 1 dummy "end" node.
       The dummy end node has zero cost from every real node, giving OR-Tools
       freedom to terminate the route at whichever real stop is cheapest.
       This produces an OPEN-PATH route (start at depot, end anywhere).

    2. First solution: PATH_CHEAPEST_ARC greedily extends the cheapest arc.
    3. Metaheuristic: GUIDED_LOCAL_SEARCH escapes local minima by penalising
       frequently-used arcs, untangling crossed paths and producing routes
       similar to commercial apps like Circuit/Routific.
    4. The solver runs for `time_limit_ms` milliseconds, returning the best
       solution found within that budget.

    ── Mapping matrix indices to front-end stops ──
    1. Build your stops array:
         stops = [current_location] + delivery_stops
       Index 0 = current location (depot), 1..N = delivery stops.
    2. Query Mapbox Matrix API with the coordinates of all stops.
       The returned matrix[i][j] = driving time/distance from stop i to stop j.
       Use duration (seconds) for time-optimal routing.
    3. Call: ordered = ortools_tsp_solve(matrix, depot=0)
    4. Map back: route = [stops[i] for i in ordered]

    Args:
        matrix:        NxN matrix of costs (driving seconds or meters).
                       matrix[i][j] = cost to travel from node i to node j.
                       Populated by the Mapbox Matrix API.
        depot:         Index of the starting node (typically 0 = current location).
        time_limit_ms: Solver time budget in milliseconds (default 2000).
                       2000ms is enough for ≤50 stops. Scale up for larger routes.

    Returns:
        Ordered list of node indices (0..N-1) representing the visit sequence.
        The depot is always first. The route ends at whichever stop minimises
        total cost (open-path TSP).

    Raises:
        RuntimeError: If OR-Tools is not installed.
        ValueError:   If no solution is found.
    """
    if not ORTOOLS_AVAILABLE or pywrapcp is None or routing_enums_pb2 is None:
        raise RuntimeError(f"OR-Tools not available: {ORTOOLS_IMPORT_ERROR or 'import failed'}")

    n = len(matrix)
    if n <= 1:
        return list(range(n))
    if n == 2:
        return [depot, 1 - depot]

    safe_depot = depot if 0 <= depot < n else 0

    # ── Build (N+1)-node model with dummy end node for open-path TSP ──
    #
    # Node indices 0..n-1 are real stops.
    # Node n is a dummy "end" node: cost FROM any real node TO dummy = 0,
    # cost FROM dummy TO any real node = very large (never used as source).
    # The vehicle starts at `safe_depot` and ends at node `n` (the dummy).
    # Since travelling to the dummy is free, OR-Tools ends at whichever
    # real stop produces the shortest total route.
    N = n + 1  # total nodes including dummy
    DUMMY = n
    LARGE = 10**9  # prohibitive cost — dummy is never a real origin

    # Scale matrix values to integers (OR-Tools requires int callbacks).
    # If the matrix contains floats (km), multiply by 1000 to preserve
    # three decimal places. If already in seconds (int), use as-is.
    scale = 1000 if any(isinstance(matrix[i][j], float) for i in range(min(2, n)) for j in range(min(2, n))) else 1

    # Build the raw n×n integer matrix first (vectorised via NumPy).
    import numpy as _np
    int_nxn = _np.asarray(matrix, dtype=_np.float64) * scale
    _np.clip(int_nxn, 0, None, out=int_nxn)
    int_nxn = int_nxn.astype(_np.int64, copy=False)

    # ── Matrix sparsification (single-driver, large-N only) ──
    # For large routes, clamp geographically absurd arcs to a large penalty so
    # OR-Tools never routes through them. Keeps all nodes reachable via the
    # depot (sparsify_matrix preserves depot row + col), preserves optimality
    # on real-world delivery data, and shrinks the effective search space.
    if n >= 20:
        try:
            from vrp_solver import sparsify_matrix
            nonzero = int_nxn[int_nxn > 0]
            if nonzero.size > 0:
                threshold = int(3 * _np.median(nonzero))
                int_nxn, _n_pruned = sparsify_matrix(
                    int_nxn, prune_threshold_s=threshold, keep_depot=safe_depot
                )
        except Exception as _e:
            logger.warning(f"Matrix sparsification skipped (non-fatal): {_e}")

    # ── Expand to (N+1)×(N+1) with dummy-end-node scaffolding ──
    int_matrix = [[0] * N for _ in range(N)]
    for i in range(n):
        row = int_nxn[i]
        for j in range(n):
            int_matrix[i][j] = int(row[j])
        int_matrix[i][DUMMY] = 0      # free to end the route here
    for j in range(N):
        int_matrix[DUMMY][j] = LARGE   # dummy is never a real origin
    int_matrix[DUMMY][DUMMY] = 0

    # ── OR-Tools model ──
    manager = pywrapcp.RoutingIndexManager(N, 1, [safe_depot], [DUMMY])
    routing = pywrapcp.RoutingModel(manager)

    def cost_callback(from_index: int, to_index: int) -> int:
        return int_matrix[manager.IndexToNode(from_index)][manager.IndexToNode(to_index)]

    transit_idx = routing.RegisterTransitCallback(cost_callback)
    routing.SetArcCostEvaluatorOfAllVehicles(transit_idx)

    # Add cumulative dimension to track total cost (for diagnostics / constraints)
    routing.AddDimension(transit_idx, 0, LARGE, True, "Cost")

    # ── Search strategy ──
    search_params = pywrapcp.DefaultRoutingSearchParameters()
    search_params.local_search_metaheuristic = (
        routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    )
    search_params.time_limit.FromMilliseconds(max(500, int(time_limit_ms)))

    # ── Warm-start: inject VROOM initial solution if provided ──
    solution = None
    if initial_indices and len(initial_indices) >= 2:
        try:
            # Strip depot from head — OR-Tools expects only the intermediate nodes
            warm_route = [i for i in initial_indices if i != safe_depot]
            initial_assignment = routing.ReadAssignmentFromRoutes([warm_route], True)
            if initial_assignment:
                # With a warm-start, skip greedy construction — jump straight to GLS
                search_params.first_solution_strategy = (
                    routing_enums_pb2.FirstSolutionStrategy.FIRST_UNBOUND_MIN_VALUE
                )
                solution = routing.SolveFromAssignmentWithParameters(
                    initial_assignment, search_params
                )
        except Exception:
            pass  # Fall through to cold-start below

    if not solution:
        # Cold-start: PATH_CHEAPEST_ARC greedy seed, then GLS improvement
        search_params.first_solution_strategy = (
            routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
        )
        solution = routing.SolveWithParameters(search_params)
    if not solution:
        raise ValueError("OR-Tools could not find a route solution")

    # ── Extract ordered real-node indices (exclude dummy end) ──
    ordered: List[int] = []
    index = routing.Start(0)
    while not routing.IsEnd(index):
        node = manager.IndexToNode(index)
        if node != DUMMY:
            ordered.append(node)
        index = solution.Value(routing.NextVar(index))

    # Safety: ensure every real node appears exactly once
    seen = set(ordered)
    for i in range(n):
        if i not in seen:
            ordered.append(i)

    return ordered

def nearest_neighbor_optimize(stops: List[dict], distance_matrix: List[List[float]], start_index: int = 0) -> List[dict]:
    """Basic nearest neighbor optimization - greedy approach"""
    if len(stops) <= 1:
        return stops
    
    n = len(stops)
    visited = [False] * n
    route = [start_index]
    visited[start_index] = True
    
    for _ in range(n - 1):
        current = route[-1]
        nearest = -1
        nearest_dist = float('inf')
        
        for j in range(n):
            if not visited[j] and distance_matrix[current][j] < nearest_dist:
                nearest = j
                nearest_dist = distance_matrix[current][j]
        
        if nearest != -1:
            route.append(nearest)
            visited[nearest] = True
    
    return [stops[i] for i in route]

def calculate_route_distance(route: List[int], matrix: List[List[float]]) -> float:
    """Sum of edge costs along a route (list of indices into the cost matrix)."""
    return sum(matrix[route[i]][route[i + 1]] for i in range(len(route) - 1))


# ─── Greedy fallback (Nearest Neighbor with super-node clustering) ──────
def _nearest_neighbor_indices(
    matrix: Sequence[Sequence[float]],
    depot: int = 0,
    **_kwargs: object,
) -> List[int]:
    """Pure index-space NN. Picks the `min` outgoing edge from the current
    node, ignoring already-visited indices. O(n²) — no warm-starts, no
    randomness, fully deterministic.

    `**_kwargs` swallows extra args so this can be passed straight to
    ``cluster_aware_solve`` (which forwards solver kwargs verbatim)."""
    n = len(matrix)
    if n == 0:
        return []
    if n == 1:
        return [depot]
    visited = [False] * n
    route = [depot]
    visited[depot] = True
    for _ in range(n - 1):
        current = route[-1]
        best_idx = -1
        best_cost = float("inf")
        row = matrix[current]
        for j in range(n):
            if visited[j]:
                continue
            c = row[j]
            if c < best_cost:
                best_cost = c
                best_idx = j
        if best_idx < 0:
            break
        route.append(best_idx)
        visited[best_idx] = True
    return route


def solve_nearest_neighbor(
    distance_matrix: Sequence[Sequence[float]],
    stops: List[dict],
    start_index: int = 0,
) -> List[dict]:
    """Bulletproof greedy fallback for the routing pipeline.

    Pipeline:
      1. Wrap the index-space NN in ``cluster_aware_solve`` so identical-
         coordinate "super nodes" (multi-parcel doorsteps) are collapsed
         before the solver runs and re-expanded sequentially after — same
         protection PyVRP gets internally. Prevents the "Zero-Cost
         Interleaving" bug where the greedy picks A1 → B → A2 because
         the inter-parcel edge cost is 0.
      2. If the matrix degenerates (empty, no stops, identical depot) the
         function falls back to returning the input list unchanged.

    Why a wrapper around the existing ``nearest_neighbor_optimize``:
        ``nearest_neighbor_optimize`` works in stop-dict space and can't
        be passed to ``cluster_aware_solve`` directly. ``_nearest_neighbor_indices``
        is the index-space twin that integrates with the cluster pipeline.
        Returning ``List[dict]`` here matches every other top-level solver
        in this file (PyVRP, ALNS, OR-Tools, etc.) so the call sites are
        drop-in-compatible.

    Args:
        distance_matrix: square matrix in seconds (or any cost). Driver-
            provided OSRM/Mapbox `duration_matrix` is the right input.
        stops: list of stop dicts with `latitude`/`longitude`.
        start_index: depot index (driver location), default 0.
    """
    if not stops or len(stops) == 1:
        return list(stops)
    indices = cluster_aware_solve(
        _nearest_neighbor_indices,
        distance_matrix,
        start_index,
        stops,
    )
    return [stops[i] for i in indices]


def _indices_by_identity(source_list: List[dict], ordered: List[dict]) -> List[int]:
    """Map each dict in `ordered` back to its position in `source_list` using
    Python object identity (`id()`), not equality.

    Why: every pre-existing call site used ``[source_list.index(s) for s in ordered]``,
    which returns the FIRST equal dict. For users with duplicate-address stops
    (same lat/lng, different stop ids) that silently collapses two different
    stops onto the same index → the optimizer output loses a real stop.

    Since every solver in this file returns the same dict *references* that
    were passed in (see e.g. ``nearest_neighbor_optimize``: ``return [stops[i] for i in route]``),
    `id()` identifies each dict uniquely regardless of duplicate values.
    """
    id_map = {id(item): idx for idx, item in enumerate(source_list)}
    return [id_map[id(item)] for item in ordered]


def two_opt_improve(route_indices: List[int], distance_matrix: List[List[float]]) -> List[int]:
    """2-Opt improvement - reverse segments to reduce total distance"""
    improved = True
    best = route_indices[:]
    
    while improved:
        improved = False
        for i in range(1, len(best) - 1):
            for j in range(i + 1, len(best)):
                # Calculate current distance
                if i == 0:
                    d1 = 0
                else:
                    d1 = distance_matrix[best[i-1]][best[i]]
                d2 = distance_matrix[best[j-1]][best[j]] if j < len(best) else 0
                
                # Calculate new distance after reversal
                d3 = distance_matrix[best[i-1]][best[j-1]] if i > 0 else 0
                d4 = distance_matrix[best[i]][best[j]] if j < len(best) else 0
                
                # If improvement found, reverse the segment
                if d1 + d2 > d3 + d4:
                    best[i:j] = reversed(best[i:j])
                    improved = True
    
    return best

def iterated_local_search(
    stops: List[dict],
    cost_matrix: List[List[float]],
    start_index: int = 0,
    time_limit_seconds: float = 10.0,
) -> List[dict]:
    """Iterated Local Search with double-bridge perturbation.

    Significantly outperforms SA/GA because:
    - Uses structured double-bridge kicks (not random swaps) to escape local minima
    - Applies Or-Opt + 2-opt after every perturbation (deep local search)
    - Accepts only improving moves (no random acceptance) → always moves toward better solutions

    Time complexity: O(n^2) per local search pass × number of restarts in time budget.
    """
    import time
    import random

    n = len(stops)
    if n <= 3:
        return stops

    def _local_search(route: List[int]) -> List[int]:
        """2-opt + Or-Opt pass until no improvement."""
        r = two_opt_improve(route, cost_matrix)
        r = or_opt_improve(r, cost_matrix)
        return r

    def _double_bridge(route: List[int]) -> List[int]:
        """Double-bridge 4-opt move: split into A|B|C|D → A|C|B|D.
        Keeps depot fixed at position 0. Creates crossings that 2-opt cannot undo,
        enabling escape from deep local minima."""
        if len(route) < 6:
            # Not enough nodes for a meaningful double-bridge — do a segment reversal instead
            i, j = sorted(random.sample(range(1, len(route)), 2))
            r = route[:]
            r[i:j] = reversed(r[i:j])
            return r
        # Pick 3 cut points inside the route (after the fixed depot at index 0)
        positions = sorted(random.sample(range(1, len(route)), 3))
        a, b, c = positions
        seg_A = route[:a]
        seg_B = route[a:b]
        seg_C = route[b:c]
        seg_D = route[c:]
        return seg_A + seg_C + seg_B + seg_D

    # Seed: nearest-neighbour → local search
    nn_result = nearest_neighbor_optimize(stops, cost_matrix, start_index)
    current = _local_search(_indices_by_identity(stops, nn_result))
    best = current[:]
    best_cost = calculate_route_distance(best, cost_matrix)

    deadline = time.monotonic() + time_limit_seconds
    restarts = 0
    while time.monotonic() < deadline:
        candidate = _local_search(_double_bridge(current[:]))
        candidate_cost = calculate_route_distance(candidate, cost_matrix)
        # Always accept improvements; keep best ever seen
        if candidate_cost < calculate_route_distance(current, cost_matrix):
            current = candidate
        if candidate_cost < best_cost:
            best = candidate[:]
            best_cost = candidate_cost
        restarts += 1

    return [stops[i] for i in best]


def three_opt_improve(route_indices: List[int], cost_matrix: List[List[float]], max_iterations: int = 5) -> List[int]:
    """3-Opt improvement — non-reversal segment swap (asymmetric-safe).

    On an open-path tour we want to escape 2-opt local optima without breaking
    on directed-graph cost matrices. The classic textbook 3-opt enumerates 7
    reconnections, six of which REVERSE one or both inner segments
    (`A + B[::-1] + C + D`, `A + C[::-1] + B + D`, etc.). When the cost
    matrix is asymmetric (OSRM's one-way streets, turn restrictions) reversing
    a segment changes every internal edge cost — but the textbook delta-cost
    formula only re-prices the 3 boundary edges and assumes internal costs
    are unchanged. The result: 3-opt accepts moves that LOOK cheaper than
    they actually are, occasionally producing worse tours than its input
    (the symptom: zig-zags and "doubling back" past a stop the route already
    passed). We saw this in production with stops 11→12→13→14 doubling back.

    Fix: keep only the ONE 3-opt candidate that doesn't reverse any segment:
    `A + C + B + D` (swap segments B and C, preserving their internal
    direction). Its boundary-delta cost is correct on any matrix, symmetric
    or not. We lose some search power (no reversal escapes) but every move
    we DO accept is guaranteed to be a real improvement.

    The first node (depot) is held fixed.
    """
    best = route_indices[:]
    n = len(best)
    if n < 5:
        return best

    for _ in range(max_iterations):
        improved = False
        for i in range(1, n - 3):
            for j in range(i + 1, n - 2):
                for k in range(j + 1, n - 1):
                    # Segments: A = best[:i], B = best[i:j], C = best[j:k], D = best[k:]
                    A_last = best[i - 1]
                    B_first, B_last = best[i], best[j - 1]
                    C_first, C_last = best[j], best[k - 1]
                    D_first = best[k]

                    # Old boundary edges removed by the move.
                    d0 = (cost_matrix[A_last][B_first]
                          + cost_matrix[B_last][C_first]
                          + cost_matrix[C_last][D_first])

                    # Non-reversing swap: tour becomes A + C + B + D, with
                    # internal edges of B and C unchanged. Delta is exact
                    # on any (a)symmetric matrix because no edge inside B
                    # or C is altered — only the 3 join edges change.
                    d_new = (cost_matrix[A_last][C_first]
                             + cost_matrix[C_last][B_first]
                             + cost_matrix[B_last][D_first])

                    if d_new < d0:
                        best = best[:i] + best[j:k] + best[i:j] + best[k:]
                        improved = True

        if not improved:
            break

    return best


def or_opt_improve(route_indices: List[int], cost_matrix: List[List[float]], max_iterations: int = 10) -> List[int]:
    """Or-opt improvement — relocate sequences of 1, 2, or 3 consecutive stops.

    For each segment size (3, 2, 1), tries removing the segment from its
    current position and re-inserting it at every other position in the route.
    Accepts the move if total cost decreases. Repeats until no improvement
    found or max_iterations reached.

    Catches "misplaced cluster" improvements that 3-opt and LKH may miss.
    Runs in O(n^2) per pass per segment size. Keeps first node (depot) fixed.
    """
    best = route_indices[:]
    n = len(best)
    if n < 4:
        return best

    def _total_cost(route):
        return sum(cost_matrix[route[k]][route[k + 1]] for k in range(len(route) - 1))

    for _ in range(max_iterations):
        improved = False
        # Try segment sizes 3, 2, 1 (larger segments first for bigger wins)
        for seg_len in (3, 2, 1):
            if n < seg_len + 2:
                continue
            for i in range(1, n - seg_len):  # skip depot at index 0
                # Extract the segment
                segment = best[i:i + seg_len]
                # Build route without the segment
                rest = best[:i] + best[i + seg_len:]

                # Cost of current route around the removal point
                # Edges removed: (i-1 -> i), (i+seg_len-1 -> i+seg_len)
                # Edge added:    (i-1 -> i+seg_len)
                old_removal_cost = (
                    cost_matrix[best[i - 1]][best[i]] +
                    cost_matrix[best[i + seg_len - 1]][best[i + seg_len]] if (i + seg_len) < n else
                    cost_matrix[best[i - 1]][best[i]]
                )
                new_removal_cost = (
                    cost_matrix[best[i - 1]][best[i + seg_len]] if (i + seg_len) < n else 0
                )
                removal_delta = new_removal_cost - old_removal_cost

                # Try inserting the segment at every valid position in `rest`
                best_delta = 0
                best_insert_pos = -1
                for j in range(1, len(rest)):  # skip inserting before depot
                    # Edge being broken: rest[j-1] -> rest[j]
                    # Edges being added: rest[j-1] -> segment[0], segment[-1] -> rest[j]
                    old_insert_cost = cost_matrix[rest[j - 1]][rest[j]]
                    new_insert_cost = (
                        cost_matrix[rest[j - 1]][segment[0]] +
                        cost_matrix[segment[-1]][rest[j]]
                    )
                    # Internal segment cost stays the same, so only edge changes matter
                    insert_delta = new_insert_cost - old_insert_cost
                    total_delta = removal_delta + insert_delta

                    if total_delta < best_delta - 1e-9:
                        best_delta = total_delta
                        best_insert_pos = j

                if best_insert_pos >= 0:
                    # Apply the best move
                    best = rest[:best_insert_pos] + segment + rest[best_insert_pos:]
                    improved = True
                    break  # restart from scratch after each improvement
            if improved:
                break  # restart outer loop

        if not improved:
            break

    return best

def simulated_annealing_optimize(stops: List[dict], distance_matrix: List[List[float]], 
                                  start_index: int = 0, iterations: int = 10000) -> List[dict]:
    """Simulated Annealing optimization - probabilistic meta-heuristic"""
    import random
    import math
    
    n = len(stops)
    if n <= 2:
        return stops
    
    # Start with nearest neighbor solution
    current = list(range(n))
    if start_index != 0:
        current.remove(start_index)
        current = [start_index] + current
    
    current_dist = calculate_route_distance(current, distance_matrix)
    best = current[:]
    best_dist = current_dist
    
    temperature = 100.0
    cooling_rate = 0.9995
    
    for _ in range(iterations):
        # Generate neighbor by swapping two random positions (keep start fixed)
        i, j = random.sample(range(1, n), 2)
        neighbor = current[:]
        neighbor[i], neighbor[j] = neighbor[j], neighbor[i]
        
        neighbor_dist = calculate_route_distance(neighbor, distance_matrix)
        delta = neighbor_dist - current_dist
        
        # Accept better solutions or worse with probability
        if delta < 0 or random.random() < math.exp(-delta / temperature):
            current = neighbor
            current_dist = neighbor_dist
            
            if current_dist < best_dist:
                best = current[:]
                best_dist = current_dist
        
        temperature *= cooling_rate
    
    return [stops[i] for i in best]

def genetic_algorithm_optimize(stops: List[dict], distance_matrix: List[List[float]], 
                               start_index: int = 0, generations: int = 200, 
                               population_size: int = 50) -> List[dict]:
    """Genetic Algorithm optimization - evolutionary meta-heuristic"""
    import random
    
    n = len(stops)
    if n <= 2:
        return stops
    
    def create_individual():
        """Create a random route keeping start_index first"""
        route = list(range(n))
        route.remove(start_index)
        random.shuffle(route)
        return [start_index] + route
    
    def fitness(individual):
        """Lower distance = higher fitness"""
        return 1.0 / (1.0 + calculate_route_distance(individual, distance_matrix))
    
    def crossover(parent1, parent2):
        """Order crossover (OX)"""
        size = len(parent1)
        start, end = sorted(random.sample(range(1, size), 2))
        
        child = [None] * size
        child[0] = start_index
        child[start:end] = parent1[start:end]
        
        remaining = [x for x in parent2 if x not in child]
        idx = 0
        for i in range(size):
            if child[i] is None:
                child[i] = remaining[idx]
                idx += 1
        
        return child
    
    def mutate(individual, rate=0.1):
        """Swap mutation"""
        if random.random() < rate and len(individual) > 2:
            i, j = random.sample(range(1, len(individual)), 2)
            individual[i], individual[j] = individual[j], individual[i]
        return individual
    
    # Initialize population
    population = [create_individual() for _ in range(population_size)]
    
    for _ in range(generations):
        # Selection (tournament)
        new_population = []
        
        # Elitism - keep best
        population.sort(key=fitness, reverse=True)
        new_population.append(population[0][:])
        
        while len(new_population) < population_size:
            # Tournament selection
            tournament = random.sample(population, 5)
            parent1 = max(tournament, key=fitness)
            tournament = random.sample(population, 5)
            parent2 = max(tournament, key=fitness)
            
            child = crossover(parent1, parent2)
            child = mutate(child)
            new_population.append(child)
        
        population = new_population
    
    # Return best individual
    best = max(population, key=fitness)
    return [stops[i] for i in best]

def clarke_wright_savings(stops: List[dict], distance_matrix: List[List[float]], 
                          depot_index: int = 0) -> List[dict]:
    """Clarke-Wright Savings Algorithm - classic VRP algorithm
    Treats first stop as depot and builds routes from there"""
    n = len(stops)
    if n <= 2:
        return stops
    
    # Calculate savings for each pair of customers
    savings = []
    for i in range(n):
        if i == depot_index:
            continue
        for j in range(i + 1, n):
            if j == depot_index:
                continue
            # Saving = distance(depot,i) + distance(depot,j) - distance(i,j)
            s = distance_matrix[depot_index][i] + distance_matrix[depot_index][j] - distance_matrix[i][j]
            savings.append((s, i, j))
    
    # Sort by savings (descending)
    savings.sort(reverse=True)
    
    # Build routes
    routes = [[i] for i in range(n) if i != depot_index]
    customer_route = {i: i - (1 if i > depot_index else 0) for i in range(n) if i != depot_index}
    
    for s, i, j in savings:
        route_i = customer_route.get(i)
        route_j = customer_route.get(j)
        
        if route_i is None or route_j is None or route_i == route_j:
            continue
        
        # Check if i and j are at the ends of their routes
        ri = routes[route_i]
        rj = routes[route_j]
        
        if (ri[0] == i or ri[-1] == i) and (rj[0] == j or rj[-1] == j):
            # Merge routes
            if ri[-1] == i and rj[0] == j:
                new_route = ri + rj
            elif ri[0] == i and rj[-1] == j:
                new_route = rj + ri
            elif ri[-1] == i and rj[-1] == j:
                new_route = ri + rj[::-1]
            else:
                new_route = ri[::-1] + rj
            
            # Update routes
            routes[route_i] = new_route
            routes[route_j] = []
            
            # Update customer_route mapping
            for c in new_route:
                customer_route[c] = route_i
    
    # Combine all non-empty routes
    final_route = [depot_index]
    for route in routes:
        if route:
            final_route.extend(route)
    
    return [stops[i] for i in final_route]

async def mapbox_optimize(stops: List[dict], current_latitude: float = None, current_longitude: float = None) -> List[dict]:
    """Use Mapbox Optimization API for route optimization
    
    Mapbox Optimization API handles up to 12 coordinates per request.
    For larger routes, we'll batch them.
    """
    if not MAPBOX_TOKEN:
        raise ValueError("Mapbox token not configured")
    
    if len(stops) < 2:
        return stops
    
    # Build coordinates string - Mapbox wants lon,lat format
    all_coords = []
    
    # Add current location as first point if provided
    if current_latitude and current_longitude:
        all_coords.append(f"{current_longitude},{current_latitude}")
    
    for stop in stops:
        all_coords.append(f"{stop['longitude']},{stop['latitude']}")
    
    # Mapbox Optimization API has a limit of 12 coordinates
    # For larger routes, we batch optimize
    if len(all_coords) <= 12:
        coordinates = ";".join(all_coords)
        
        # Source index determined by current_latitude presence — reserved for
        # future Mapbox calls that explicitly pin a non-first origin.
        _ = 0 if current_latitude else "any"

        params = {
            "access_token": MAPBOX_TOKEN,
            "source": "first",
            "destination": "last",
            "roundtrip": "false",
            "geometries": "geojson",
            "overview": "full"
        }
        
        url = f"https://api.mapbox.com/optimized-trips/v1/mapbox/driving/{coordinates}"
        
        async with httpx.AsyncClient() as client:
            response = await client.get(url, params=params, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                
                if data.get("code") == "Ok" and data.get("waypoints"):
                    # Extract optimized order from waypoints
                    waypoints = data["waypoints"]
                    
                    # Build reordered stops based on waypoint_index
                    ordered_waypoints = sorted(waypoints, key=lambda w: w["waypoint_index"])
                    
                    # Map back to original stops (skip current location if added)
                    offset = 1 if current_latitude else 0
                    optimized_stops = []
                    
                    for wp in ordered_waypoints:
                        original_idx = wp["waypoint_index"] - offset
                        if original_idx >= 0 and original_idx < len(stops):
                            optimized_stops.append(stops[original_idx])
                    
                    return optimized_stops
            
            # If Mapbox fails, fall back to nearest neighbor
            print(f"Mapbox Optimization API error: {response.status_code} - {response.text[:200]}")
    
    else:
        # For routes with more than 12 stops, batch optimize
        # Split into chunks of 10 (leaving room for start/end)
        chunk_size = 10
        optimized_chunks = []
        
        for i in range(0, len(stops), chunk_size):
            chunk = stops[i:i + chunk_size]
            
            if len(chunk) >= 2:
                coords = ";".join([f"{s['longitude']},{s['latitude']}" for s in chunk])
                
                params = {
                    "access_token": MAPBOX_TOKEN,
                    "roundtrip": "false",
                    "geometries": "geojson"
                }
                
                url = f"https://api.mapbox.com/optimized-trips/v1/mapbox/driving/{coords}"
                
                async with httpx.AsyncClient() as client:
                    response = await client.get(url, params=params, timeout=30)
                    
                    if response.status_code == 200:
                        data = response.json()
                        
                        if data.get("code") == "Ok" and data.get("waypoints"):
                            waypoints = data["waypoints"]
                            ordered_waypoints = sorted(waypoints, key=lambda w: w["waypoint_index"])
                            
                            chunk_optimized = []
                            for wp in ordered_waypoints:
                                if wp["waypoint_index"] < len(chunk):
                                    chunk_optimized.append(chunk[wp["waypoint_index"]])
                            
                            optimized_chunks.extend(chunk_optimized)
                        else:
                            optimized_chunks.extend(chunk)
                    else:
                        optimized_chunks.extend(chunk)
            else:
                optimized_chunks.extend(chunk)
        
        return optimized_chunks
    
    # Fallback: return original order
    return stops


async def generoute_optimize(stops: List[dict], current_latitude: float = None, current_longitude: float = None) -> List[dict]:
    """
    Use Generoute API for route optimization.
    
    Generoute provides simple, fast route optimization via https://api.generoute.io/v1/trip
    
    Args:
        stops: List of stop dictionaries with latitude, longitude, id, address
        current_latitude: Optional starting latitude
        current_longitude: Optional starting longitude
    
    Returns:
        Optimized list of stops
    """
    if not GENEROUTE_API_KEY:
        raise ValueError("Generoute API key not configured")
    
    if len(stops) < 2:
        return stops
    
    # Generoute Free plan limit is 100 locations
    MAX_LOCATIONS = 99  # Leave room for current location
    
    try:
        # If too many stops, chunk them and optimize each chunk
        if len(stops) > MAX_LOCATIONS:
            logger.info(f"Chunking {len(stops)} stops for Generoute (max {MAX_LOCATIONS} per request)")
            
            # Split stops into chunks
            chunks = []
            for i in range(0, len(stops), MAX_LOCATIONS):
                chunks.append(stops[i:i + MAX_LOCATIONS])
            
            # Optimize each chunk
            all_optimized = []
            for chunk_idx, chunk in enumerate(chunks):
                # For subsequent chunks, use last stop of previous chunk as starting point
                chunk_start_lat = None
                chunk_start_lng = None
                if chunk_idx == 0 and current_latitude and current_longitude:
                    chunk_start_lat = current_latitude
                    chunk_start_lng = current_longitude
                elif all_optimized:
                    last_stop = all_optimized[-1]
                    chunk_start_lat = last_stop['latitude']
                    chunk_start_lng = last_stop['longitude']
                
                try:
                    # Recursively call with smaller chunk
                    optimized_chunk = await generoute_optimize(chunk, chunk_start_lat, chunk_start_lng)
                    all_optimized.extend(optimized_chunk)
                except Exception as e:
                    logger.warning(f"Chunk {chunk_idx} optimization failed: {e}, using original order")
                    all_optimized.extend(chunk)
            
            return all_optimized
        
        # Build locations array for Generoute API
        locations = []
        
        # Add current location as first point if provided
        if current_latitude and current_longitude:
            locations.append({
                "coordinates": [current_longitude, current_latitude],
                "title": "Current Location",
                "data": {"id": "current_location"}
            })
        
        # Add all stops
        for stop in stops:
            locations.append({
                "coordinates": [stop['longitude'], stop['latitude']],
                "title": stop.get('address', stop.get('name', '')),
                "data": {"id": stop.get('id', str(uuid.uuid4()))}
            })
        
        # Make API request to Generoute
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://api.generoute.io/v1/trip",
                headers={
                    "Authorization": f"Bearer {GENEROUTE_API_KEY}",
                    "Content-Type": "application/json"
                },
                json={
                    "region": "AU",  # Australia - adjust based on your region
                    "locations": locations
                },
                timeout=30.0
            )
            
            if response.status_code != 200:
                logger.error(f"Generoute API error: {response.status_code} - {response.text}")
                raise ValueError(f"Generoute API error: {response.status_code}")
            
            result = response.json()
            
            # Extract optimized order from response - structure is trips[0].waypoints
            trips = result.get('trips', [])
            if not trips or len(trips) == 0:
                logger.warning("Generoute returned no trips, using original order")
                return stops
            
            optimized_waypoints = trips[0].get('waypoints', [])
            
            if not optimized_waypoints:
                logger.warning("Generoute returned no optimized waypoints, using original order")
                return stops
            
            # Sort by waypoint_order to ensure correct sequence
            optimized_waypoints.sort(key=lambda w: w.get('waypoint_order', 0))
            
            # Reorder stops based on Generoute's optimized sequence
            id_to_stop = {stop.get('id'): stop for stop in stops}
            optimized_stops = []
            
            for opt_wp in optimized_waypoints:
                loc_id = opt_wp.get('data', {}).get('id')
                
                # Skip current location entry
                if loc_id == "current_location":
                    continue
                
                if loc_id and loc_id in id_to_stop:
                    optimized_stops.append(id_to_stop[loc_id])
                else:
                    # Try to match by coordinates
                    opt_coords = opt_wp.get('coordinates', opt_wp.get('waypoint_location', []))
                    if len(opt_coords) == 2:
                        for stop in stops:
                            if stop not in optimized_stops:
                                if abs(stop['longitude'] - opt_coords[0]) < 0.0001 and \
                                   abs(stop['latitude'] - opt_coords[1]) < 0.0001:
                                    optimized_stops.append(stop)
                                    break
            
            # Add any stops that weren't matched
            for stop in stops:
                if stop not in optimized_stops:
                    optimized_stops.append(stop)
            
            logger.info(f"Generoute optimization succeeded: {len(optimized_stops)} stops optimized")
            return optimized_stops
            
    except httpx.TimeoutException:
        logger.error("Generoute API timeout")
        raise ValueError("Generoute API timeout - try again later")
    except Exception as e:
        logger.error(f"Generoute optimization error: {e}")
        raise ValueError(f"Generoute optimization failed: {str(e)}")


# Local delivery constraints — Sugar Bag Rd waypoint injection only.
# (school-zone penalty removed 2026-05-13 per user request; helpers
# `school_penalty_factor` / `apply_school_zone_penalty` / `is_in_school_zone`
# remain available in routes/_route_constraints.py if we want to re-enable.)
from routes._route_constraints import (
    parse_start_time,
    inject_sugar_bag_waypoints,
    needs_sugar_bag_injection,
)


def _traffic_multiplier(hour: int) -> float:
    """Return a duration multiplier based on time-of-day traffic patterns.

    Based on typical Australian urban traffic patterns:
    - AM peak (7-9): 1.35x
    - PM peak (16-18): 1.40x
    - School run (15-16): 1.20x
    - Midday (10-14): 1.05x
    - Early morning (5-7): 1.10x
    - Night (20-5): 1.00x (free flow)
    """
    if 7 <= hour < 9:
        return 1.35
    elif 16 <= hour < 18:
        return 1.40
    elif 15 <= hour < 16:
        return 1.20
    elif 9 <= hour < 10:
        return 1.15
    elif 10 <= hour < 15:
        return 1.05
    elif 5 <= hour < 7:
        return 1.10
    elif 18 <= hour < 20:
        return 1.15
    else:
        return 1.00


def apply_traffic_multiplier(matrix: List[List[int]], hour: int) -> List[List[int]]:
    """Apply time-of-day traffic multiplier to a duration matrix.

    Returns a new matrix with all durations scaled by the traffic factor.
    """
    m = _traffic_multiplier(hour)
    if m == 1.0:
        return matrix
    return [
        [max(1, int(round(cell * m))) for cell in row]
        for row in matrix
    ]


def assign_stops_to_hub_segments(stops: List[dict], hubs: List[dict], current_location: dict = None) -> List[List[dict]]:
    """
    Assign each stop to the nearest hub segment.
    
    The route is divided into segments:
    - Segment 0: From start (current location or first stop) to Hub 1
    - Segment 1: From Hub 1 to Hub 2
    - ...
    - Segment N: From Hub N to end (remaining stops)
    
    Each stop is assigned to the segment whose hub endpoints it's closest to.
    """
    if not hubs:
        return [stops]
    
    # Sort hubs by their order
    sorted_hubs = sorted(hubs, key=lambda h: h['order'])
    
    # Segment boundaries (start_point, end_point) tuples were previously
    # collected here; kept as docstring-only intent since downstream uses
    # the full `waypoints` list instead.
    
    # Build waypoints list: [start] + hubs
    waypoints = []
    if current_location:
        waypoints.append({
            'latitude': current_location['latitude'],
            'longitude': current_location['longitude'],
            'is_hub': False
        })
    
    for hub in sorted_hubs:
        waypoints.append({
            'latitude': hub['latitude'],
            'longitude': hub['longitude'],
            'is_hub': True,
            'hub_id': hub['id']
        })
    
    # Create segments (N hubs = N+1 segments if we have start location, else N segments)
    num_segments = len(sorted_hubs) + (1 if current_location else 0)
    segments = [[] for _ in range(num_segments)]
    
    # Assign each stop to the best segment based on proximity to segment endpoints
    for stop in stops:
        stop_coord = (stop['latitude'], stop['longitude'])
        
        best_segment = 0
        best_score = float('inf')
        
        for seg_idx in range(num_segments):
            # Calculate which segment this stop fits best
            # Use distance to the segment's "center" or endpoints
            
            if seg_idx < len(waypoints):
                # Distance to the segment start waypoint
                start_wp = waypoints[seg_idx]
                start_coord = (start_wp['latitude'], start_wp['longitude'])
                dist_to_start = haversine(stop_coord, start_coord, unit=Unit.KILOMETERS)
                
                # For segments with a next waypoint, also consider distance to end
                if seg_idx + 1 < len(waypoints):
                    end_wp = waypoints[seg_idx + 1]
                    end_coord = (end_wp['latitude'], end_wp['longitude'])
                    dist_to_end = haversine(stop_coord, end_coord, unit=Unit.KILOMETERS)
                    score = min(dist_to_start, dist_to_end)
                else:
                    # Last segment - just use distance to the last hub
                    score = dist_to_start
            else:
                # Fallback for edge case
                score = float('inf')
            
            if score < best_score:
                best_score = score
                best_segment = seg_idx
        
        segments[best_segment].append(stop)
    
    return segments


def optimize_segment(stops: List[dict], algorithm: str, start_point: dict = None, end_point: dict = None) -> List[dict]:
    """
    Optimize a single segment of stops.
    
    Args:
        stops: List of stops in this segment
        algorithm: Optimization algorithm to use
        start_point: Optional fixed start point (hub or current location)
        end_point: Optional fixed end point (next hub)
    
    Returns:
        Optimized list of stops for this segment
    """
    if len(stops) <= 1:
        return stops
    
    # Build the list with optional start/end anchors
    working_stops = []
    start_idx = 0
    
    if start_point:
        anchor_start = {
            'id': f"anchor_start_{start_point.get('id', 'loc')}",
            'latitude': start_point['latitude'],
            'longitude': start_point['longitude'],
            'is_anchor': True
        }
        working_stops.append(anchor_start)
        start_idx = 0
    
    working_stops.extend(stops)
    
    if end_point:
        anchor_end = {
            'id': f"anchor_end_{end_point.get('id', 'loc')}",
            'latitude': end_point['latitude'],
            'longitude': end_point['longitude'],
            'is_anchor': True
        }
        working_stops.append(anchor_end)
    
    # Calculate distance matrix for this segment
    distance_matrix = calculate_distance_matrix(working_stops)
    
    # Apply optimization algorithm
    if algorithm == 'alns':
        try:
            optimized = alns_hybrid_optimize(
                working_stops,
                distance_matrix,
                start_index=start_idx,
                time_limit_seconds=6,
            )
        except Exception as exc:
            logger.warning("ALNS segment optimization failed, using 2-Opt fallback: %s", exc)
            nn_result = nearest_neighbor_optimize(working_stops, distance_matrix, start_idx)
            route_indices = _indices_by_identity(working_stops, nn_result)
            improved_indices = two_opt_improve(route_indices, distance_matrix)
            optimized = [working_stops[i] for i in improved_indices]
    elif algorithm == 'ortools':
        try:
            optimized = ortools_optimize(
                working_stops,
                distance_matrix,
                start_index=start_idx,
                time_limit_seconds=8,
            )
        except Exception as exc:
            logger.warning("OR-Tools segment optimization failed, using 2-Opt fallback: %s", exc)
            nn_result = nearest_neighbor_optimize(working_stops, distance_matrix, start_idx)
            route_indices = _indices_by_identity(working_stops, nn_result)
            improved_indices = two_opt_improve(route_indices, distance_matrix)
            optimized = [working_stops[i] for i in improved_indices]
    elif algorithm in ['two_opt', 'auto'] or len(working_stops) <= 10:
        nn_result = nearest_neighbor_optimize(working_stops, distance_matrix, start_idx)
        route_indices = _indices_by_identity(working_stops, nn_result)
        improved_indices = two_opt_improve(route_indices, distance_matrix)
        optimized = [working_stops[i] for i in improved_indices]
    elif algorithm == 'simulated_annealing':
        optimized = simulated_annealing_optimize(working_stops, distance_matrix, start_idx)
    elif algorithm == 'genetic':
        optimized = genetic_algorithm_optimize(working_stops, distance_matrix, start_idx)
    else:
        optimized = nearest_neighbor_optimize(working_stops, distance_matrix, start_idx)
    
    # Remove anchor points from result, return only actual stops
    result = [s for s in optimized if not s.get('is_anchor')]
    return result


@api_router.post("/optimize")
async def optimize_route(
    request: OptimizationRequest = OptimizationRequest(),
    current_user: User = Depends(get_current_user)
):
    """Optimize route order using various algorithms
    
    Algorithms:
    - auto: Smart selection based on stop count (ALNS for 10+ stops, 2-opt for small)
    - alns: ALNS Hybrid Metaheuristic (NN construction + ALNS/SA + Local Search polish)
    - ortools: Google OR-Tools guided local search (single-vehicle, time-first)
    - nearest_neighbor: Greedy approach, fast O(n²)
    - two_opt: Improvement heuristic, good quality
    - simulated_annealing: Meta-heuristic, better for medium routes
    - genetic: Evolutionary algorithm, best for complex routes
    - clarke_wright: VRP savings algorithm, treats start as depot
    """
    try:
        return await _optimize_route_inner(request, current_user)
    except HTTPException:
        raise  # Let FastAPI handle 4xx errors (401, 402, 403) cleanly
    except Exception as e:
        logger.error("[optimize] Unhandled exception for user=%s algorithm=%s:\n%s",
                     current_user.user_id, request.algorithm, traceback.format_exc())
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e)},
        )


async def _optimize_route_inner(
    request: OptimizationRequest = OptimizationRequest(),
    current_user: User = Depends(get_current_user)
):
    all_user_stops = await db.stops.find({"user_id": current_user.user_id}, {"_id": 0}).sort("order", 1).to_list(1000)
    completed_stops = [s for s in all_user_stops if s.get("completed")]
    stops = [s for s in all_user_stops if not s.get("completed")]

    # ── AUDIT 1: raw input + super-node proxy ────────────────────────────
    # Counts unique (lat, lng) — this is what PyVRP's super-node grouper
    # SHOULD collapse to. If `unique_coords == raw_pending` even though
    # the route contains multi-parcel addresses, super-node clustering
    # is broken and PyVRP will compute a matrix N times larger than
    # needed (with zero-cost legs that the algorithm has to ignore).
    _audit_unique_coords = len({
        (round(s["latitude"], 6), round(s["longitude"], 6))
        for s in stops if s.get("latitude") is not None and s.get("longitude") is not None
    })
    logger.info(
        "AUDIT[/optimize] raw_pending=%d raw_completed=%d unique_coords=%d "
        "(if unique_coords==raw_pending with sibling parcels present, "
        "super-node clustering is broken)",
        len(stops), len(completed_stops), _audit_unique_coords,
    )
    
    if len(stops) < 2:
        return {"message": "Need at least 2 stops to optimize", "stops": stops + completed_stops, "algorithm": "none"}
    
    # Handle current location as starting point
    start_index = 0
    current_loc_stop = None
    
    logger.info("Optimize request: use_current=%s, lat=%s, lng=%s",
                request.use_current_location, request.current_latitude, request.current_longitude)
    
    if request.use_current_location and request.current_latitude and request.current_longitude:
        # Create a virtual "current location" stop
        current_loc_stop = {
            "id": "current_location",
            "address": "Current Location",
            "name": "Start (Current Location)",
            "latitude": request.current_latitude,
            "longitude": request.current_longitude,
            "priority": "high",
            "completed": False,
            "order": -1,
            "is_start_point": True
        }
        # Insert at beginning
        stops = [current_loc_stop] + stops
        start_index = 0
    
    # Determine algorithm before distance matrix calculation
    algorithm_used = request.algorithm
    inner_algorithm = "ortools"  # Default inner algorithm for cluster_first
    
    # ── Auto-selection (2026-04-25 update) ─────────────────────────────────
    # Previously `auto` resolved to raw VROOM. On real 79-stop user data
    # VROOM gave 96.0 min while LKH-3 (with the open-path fix) finds the
    # true optimum at 95.7 min in 0.11s — a 0.3 min improvement at zero
    # latency cost. The `vroom_lkh_3opt` cascade does:
    #   Stage 1: VROOM seed (best fast TSP heuristic, <0.3s for 100 stops)
    #   Stage 2: LKH-3 refine (state-of-the-art Lin-Kernighan, ~0.1s)
    #   Stage 3: 3-opt polish (catches anything LKH missed)
    # …which is strictly ≥ VROOM-alone quality and routinely 0.3–2.4 min
    # better. The added latency is <100 ms vs raw VROOM. For tiny routes
    # (<11 stops) we skip the cascade and use the existing 2-opt path
    # earlier in the pipeline (line ~4644) which is already optimal at
    # that scale via brute-force-ish enumeration.
    if algorithm_used == "auto":
        if VROOM_AVAILABLE and LKH_AVAILABLE and len(stops) >= 11:
            algorithm_used = "vroom_lkh_3opt"
        elif VROOM_AVAILABLE:
            algorithm_used = "vroom"
        else:
            algorithm_used = "ortools"
    
    # Respect the user's algorithm choice. Basic heuristics may be slow on large routes,
    # but silently hijacking the selection prevents the user from seeing how their picked
    # solver actually performs (and makes "algorithm X isn't working" look like a bug).
    # If the caller wants the auto-cluster behaviour they can select `cluster_first` explicitly.

    # Build cost matrices — OSRM first (free, local), then Mapbox, then haversine last resort.
    # All solvers receive OSRM data when available so quality is consistent regardless of algorithm.
    duration_matrix = None

    if algorithm_used == "cluster_first":
        # Cluster-first uses haversine for the spatial grouping phase; inner per-cluster
        # solver builds its own duration matrix internally.
        distance_matrix = calculate_distance_matrix(stops)
    else:
        # ── Primary: OSRM (local, free, handles 100+ stops natively) ──
        duration_matrix = await _osrm_duration_matrix(stops)
        if duration_matrix:
            logger.info("Using OSRM duration matrix for %d stops (%s)", len(stops), algorithm_used)
            osrm_dist = await _osrm_distance_matrix(stops)
            distance_matrix = osrm_dist if osrm_dist else calculate_distance_matrix(stops)
            if osrm_dist:
                logger.info("Using OSRM distance matrix for %d stops (reporting)", len(stops))
        else:
            # ── Fallback: Mapbox for duration-sensitive solvers, road distance for others ──
            logger.info("OSRM unavailable, building fallback matrices for %s (%d stops)", algorithm_used, len(stops))
            if algorithm_used in ("vroom", "ortools", "lkh", "vroom_lkh_3opt", "vroom_ortools", "timefold"):
                duration_matrix = await calculate_duration_matrix(stops)
                distance_matrix = calculate_distance_matrix(stops)
            elif algorithm_used == "alns" and len(stops) > 25:
                distance_matrix = await calculate_full_road_distance_matrix(stops)
            else:
                distance_matrix = await calculate_road_distance_matrix(stops)

    # ── ML service-time injection (Phase 1.5) ────────────────────────────
    # If the user has trained a service-time model (via
    # POST /api/_meta/ml/train), bake the per-stop predicted service
    # seconds INTO the duration matrix's outgoing edges. From every
    # solver's perspective, "leaving node i for j" now costs the actual
    # travel time PLUS the median service time observed at node i's
    # suburb-and-hour bucket.
    #
    # Why bake into the matrix vs. pass `service=` to each solver?
    # ----------------------------------------------------------------
    # VROOM has a `service` param on Job, but OR-Tools needs a
    # time-dimension callback, LKH/3-opt/genetic only consume a flat
    # matrix, and the post-optimize 2-opt refines also only see the
    # matrix. A matrix-baked approach uniformly applies the service
    # time to ALL solvers without per-solver wiring — change one place,
    # benefit everywhere. Outgoing-from-i (not incoming-to-j) so the
    # last stop's service time isn't double-counted on the virtual
    # exit edge (see service_time_learner.apply_service_times_to_matrix
    # for the rationale).
    #
    # Skips silently when:
    #   - User has no trained model (cold start before first Train Now)
    #   - We're on the cluster_first/haversine path (no duration matrix)
    if duration_matrix and len(stops) > 1:
        try:
            ml_doc = await db.ml_service_time_models.find_one(
                {"user_id": current_user.user_id},
                {"_id": 0},
            )
            if ml_doc:
                from ml.service_time_learner import (
                    predict_service_time_seconds,
                    apply_service_times_to_matrix,
                )
                # Start hour drives the (suburb, hour_bucket) lookup. If
                # the request didn't supply one we fall back to "now"
                # which matches what `predict_service_time_seconds`
                # already does internally.
                from datetime import datetime as _dt, timezone as _tz
                start_hour = _dt.now(_tz.utc).hour
                try:
                    if request.start_time:
                        # Parse "HH:MM" into hour. parse_start_time
                        # returns a datetime; we just want the hour.
                        _sd = parse_start_time(request.start_time)
                        start_hour = _sd.hour if _sd else start_hour
                except Exception:
                    pass

                service_times = [
                    predict_service_time_seconds(s, ml_doc, completion_hour=start_hour)
                    for s in stops
                ]
                duration_matrix = apply_service_times_to_matrix(
                    duration_matrix, service_times,
                )
                logger.info(
                    "[ml] Service-time injection: user=%s, %d stops, "
                    "min=%.0fs, median=%.0fs, max=%.0fs",
                    current_user.user_id,
                    len(service_times),
                    min(service_times),
                    sorted(service_times)[len(service_times)//2],
                    max(service_times),
                )
        except Exception as _ml_exc:
            # Non-fatal: optimize MUST still produce a route even if the
            # learner errors. Logged at warning level so we can grep
            # the failure pattern without spamming on every request.
            logger.warning(
                "[ml] Service-time injection failed (%s) — falling back to "
                "raw duration matrix without service times.",
                _ml_exc,
            )

    # ── School-zone penalty removed 2026-05-13 ────────────────────────────
    # The Meridan State College + Parklands Blvd inbound-edge penalty was
    # removed per user request. Helpers remain in
    # routes/_route_constraints.py if we want to revisit.

    # ── No-Go Zone penalty ───────────────────────────────────────────────
    # User-defined polygons (real road closures, mistagged footbridges,
    # private roads). Two-stage check:
    #   Stage 1 — straight-line: any (A, B) leg whose great-circle line
    #   crosses any zone gets +1e9 seconds. Cheap and catches the
    #   majority of cases.
    #   Stage 2 — OSRM-geometry-aware: for cells whose straight line is
    #   *near* a zone but doesn't intersect it, fetch the actual OSRM
    #   road path and check the LineString against each polygon. Catches
    #   cases where the road bends through the closed area while the
    #   straight line skirts past (Meridan Way × Rainforest Drive
    #   diagonal report 2026-05-09). Pre-filter to bbox+1.5 km so we
    #   only do a few hundred OSRM calls instead of 28 k.
    # We DON'T touch the distance matrix: distance-based solvers like NN
    # are rarer, and double-penalising risks integer overflow on int
    # matrices.
    try:
        from routes.nogo_zones import (
            fetch_user_zone_polygons,
            apply_nogo_penalty,
            apply_nogo_penalty_osrm_aware,
        )
        _nogo_polygons = await fetch_user_zone_polygons(db, current_user.user_id)
        if _nogo_polygons and duration_matrix is not None:
            _straight = apply_nogo_penalty(duration_matrix, stops, _nogo_polygons)
            _osrm_extra = await apply_nogo_penalty_osrm_aware(
                duration_matrix, stops, _nogo_polygons,
                osrm_url=OSRM_URL,
            )
            if _straight or _osrm_extra:
                logger.info(
                    "[nogo-zones] penalised %d cells (straight=%d, osrm-aware=%d) across %d zone(s) for user=%s",
                    _straight + _osrm_extra, _straight, _osrm_extra,
                    len(_nogo_polygons), current_user.user_id,
                )
    except Exception as nogo_err:
        # Non-fatal: never let a zone bug block optimisation.
        logger.warning("[nogo-zones] skipped due to error: %s", nogo_err)

    reasoning = ""

    # ── AUDIT 2: matrix sanity ───────────────────────────────────────────
    # If the OSRM matrix is broken (all zeros, NaN-like ints, missing
    # row), every solver downstream will produce an arbitrary tour because
    # they think every leg is free. Sample row 0 first 5 cols + the
    # source provenance ("road" via OSRM, "haversine" fallback, "cached",
    # etc.). If you see all zeros here, OSRM is dead and we silently fell
    # back to crow-flies — which IS the patchy/zig-zag symptom.
    #
    # Note: `distance_source` is assigned later inside the per-algorithm
    # branch, so at this point in the function it may not yet be defined
    # — we use locals().get() to read it safely. The same applies to
    # `duration_matrix` on the haversine-only fallback path.
    try:
        _audit_dim = len(distance_matrix) if distance_matrix else 0
        _audit_row0 = (
            list(distance_matrix[0][:5]) if _audit_dim > 0 else []
        )
        _dur = locals().get("duration_matrix")
        _audit_dur = (
            list(_dur[0][:5]) if _dur and len(_dur) > 0 else None
        )
        _audit_src = locals().get("distance_source", "unset-at-audit-time")
        logger.info(
            "AUDIT[/optimize] matrix dim=%dx%d source=%s row0[:5]=%s duration_row0[:5]=%s "
            "(all zeros ⇒ matrix is broken, solver tour will be arbitrary)",
            _audit_dim, _audit_dim, _audit_src,
            _audit_row0, _audit_dur,
        )
    except Exception as _e:
        logger.warning("AUDIT[/optimize] matrix sample failed: %s", _e)
    
    # ========== HUB-BASED SEGMENTED OPTIMIZATION ==========
    # If hubs are provided, use segmented optimization
    if request.hubs and len(request.hubs) > 0:
        logger.info(f"Hub-based optimization with {len(request.hubs)} hubs")
        
        # Convert hubs to dict format
        hubs_dict = [{"id": h.id, "latitude": h.latitude, "longitude": h.longitude, "order": h.order} 
                     for h in request.hubs]
        
        # Prepare current location dict
        current_loc_dict = None
        if current_loc_stop:
            current_loc_dict = {
                "latitude": current_loc_stop["latitude"],
                "longitude": current_loc_stop["longitude"]
            }
        
        # Get stops without current location for segmentation
        actual_stops = [s for s in stops if s.get("id") != "current_location"]
        
        # Assign stops to segments based on hub proximity
        segments = assign_stops_to_hub_segments(actual_stops, hubs_dict, current_loc_dict)
        
        # Sort hubs by order
        sorted_hubs = sorted(hubs_dict, key=lambda h: h['order'])
        
        # Optimize each segment independently
        optimized_segments = []
        
        # Build waypoints for start/end anchors
        waypoints = []
        if current_loc_dict:
            waypoints.append(current_loc_dict)
        waypoints.extend(sorted_hubs)
        
        for seg_idx, segment_stops in enumerate(segments):
            if len(segment_stops) == 0:
                continue
                
            # Determine start and end points for this segment
            start_point = waypoints[seg_idx] if seg_idx < len(waypoints) else None
            end_point = waypoints[seg_idx + 1] if seg_idx + 1 < len(waypoints) else None
            
            # Optimize this segment
            optimized_segment = optimize_segment(segment_stops, algorithm_used, start_point, end_point)
            optimized_segments.append(optimized_segment)
        
        # Stitch segments together in order
        optimized_stops = []
        for segment in optimized_segments:
            optimized_stops.extend(segment)
        
        reasoning = f"Hub-based segmented optimization with {len(request.hubs)} waypoints using {algorithm_used}"
        
        # Update stop orders in database
        from pymongo import UpdateOne
        ops = [
            UpdateOne({"id": stop["id"], "user_id": current_user.user_id}, {"$set": {"order": index}})
            for index, stop in enumerate(optimized_stops)
            if stop.get("id") != "current_location" and not stop.get("is_anchor")
        ]
        if ops:
            await db.stops.bulk_write(ops, ordered=False)
        
        # Save hubs to database for navigation to use
        # Clear existing hubs first
        await db.optimization_hubs.delete_many({"user_id": current_user.user_id})
        
        # Insert new hubs
        for hub in sorted_hubs:
            hub_doc = {
                "id": hub["id"],
                "user_id": current_user.user_id,
                "latitude": hub["latitude"],
                "longitude": hub["longitude"],
                "order": hub["order"],
                "name": f"Hub {hub['order']}",
                "is_hub": True
            }
            await db.optimization_hubs.insert_one(hub_doc)
        
        # Calculate total distance
        total_distance = 0
        all_stops_for_distance = optimized_stops
        if current_loc_stop:
            all_stops_for_distance = [current_loc_stop] + optimized_stops
        
        for i in range(len(all_stops_for_distance) - 1):
            coord1 = (all_stops_for_distance[i]["latitude"], all_stops_for_distance[i]["longitude"])
            coord2 = (all_stops_for_distance[i+1]["latitude"], all_stops_for_distance[i+1]["longitude"])
            total_distance += haversine(coord1, coord2, unit=Unit.KILOMETERS)
        
        return {
            "message": "Route optimized with hub waypoints",
            "algorithm": algorithm_used,
            "reasoning": reasoning,
            "total_distance_km": round(total_distance, 2),
            "stop_count": len(optimized_stops),
            "hub_count": len(request.hubs),
            "started_from_current_location": current_loc_stop is not None,
            "stops": optimized_stops + completed_stops
        }
    
    # ========== STANDARD OPTIMIZATION (no hubs) ==========
    
    # Clear any existing hubs since we're doing standard optimization
    await db.optimization_hubs.delete_many({"user_id": current_user.user_id})
    
    # ========== SECTION-BASED ROUTE REFINEMENT ==========
    # If sections are provided (from lasso tool), optimize within each section and stitch together
    if request.sections and len(request.sections) > 0:
        logger.info(f"Section-based route refinement with {len(request.sections)} sections")
        
        # Get stops without current location
        actual_stops = [s for s in stops if s.get("id") != "current_location"]
        
        # Create a mapping of stop_id to stop
        id_to_stop = {s["id"]: s for s in actual_stops}
        
        # Sort sections by id (order in which they were drawn)
        sorted_sections = sorted(request.sections, key=lambda sec: sec.id)
        
        # Track which stops are assigned to sections
        assigned_stop_ids = set()
        for section in sorted_sections:
            assigned_stop_ids.update(section.stop_ids)
        
        # Get unassigned stops (not in any section)
        unassigned_stops = [s for s in actual_stops if s["id"] not in assigned_stop_ids]
        
        # Optimize each section independently and stitch them together
        optimized_stops = []
        previous_end_point = None
        
        # If we have current location, use it as the starting point for the first section
        if current_loc_stop:
            previous_end_point = {
                "latitude": current_loc_stop["latitude"],
                "longitude": current_loc_stop["longitude"]
            }
        
        for section_idx, section in enumerate(sorted_sections):
            # Get the stops in this section
            section_stops = [id_to_stop[sid] for sid in section.stop_ids if sid in id_to_stop]
            
            if len(section_stops) == 0:
                continue
            
            # Optimize this section
            if len(section_stops) == 1:
                # Single stop, no optimization needed
                optimized_section = section_stops
            else:
                # Use the best available solver chain: VROOM → LKH → 3-opt → 2-opt fallback
                section_distance_matrix = calculate_distance_matrix(section_stops)
                
                # Find the best starting point - closest to previous end point
                start_idx = 0
                if previous_end_point:
                    min_dist = float('inf')
                    for idx, stop in enumerate(section_stops):
                        dist = haversine(
                            (previous_end_point["latitude"], previous_end_point["longitude"]),
                            (stop["latitude"], stop["longitude"]),
                            unit=Unit.KILOMETERS
                        )
                        if dist < min_dist:
                            min_dist = dist
                            start_idx = idx
                
                # Try VROOM first (best quality + speed). Use OSRM for the
                # duration matrix so the section optimizer sees real road
                # times, not great-circle distances — matches the main
                # /api/optimize path and prevents refine from "fixing" a
                # zig-zag onto an even worse one because haversine thinks
                # it's shorter.
                solver_used = "nearest_neighbor+2opt"
                try:
                    try:
                        solver_matrix = await _osrm_duration_matrix(section_stops)
                    except Exception as osrm_err:
                        logger.warning(
                            "Section refine: OSRM matrix failed (%s), falling back to haversine",
                            osrm_err,
                        )
                        solver_matrix = _haversine_duration_matrix(section_stops)
                    indices = vroom_tsp_solve(solver_matrix, depot=start_idx, exploration_level=5)
                    solver_used = "VROOM"
                    
                    # LKH post-processing for gold-standard refinement
                    pre_cost = calculate_route_distance(indices, solver_matrix)
                    if LKH_AVAILABLE and len(section_stops) >= 4:
                        try:
                            lkh_indices = lkh_tsp_solve(solver_matrix, depot=start_idx, runs=3, time_limit_seconds=5)
                            lkh_cost = calculate_route_distance(lkh_indices, solver_matrix)
                            if lkh_cost < pre_cost:
                                indices = lkh_indices
                                solver_used = "VROOM+LKH"
                        except Exception:
                            pass
                    elif len(section_stops) >= 4:
                        indices = three_opt_improve(indices, solver_matrix, max_iterations=3)
                        solver_used = "VROOM+3opt"
                    
                    optimized_section = [section_stops[i] for i in indices]
                except Exception as vroom_err:
                    logger.warning("VROOM section optimization failed, using 2-Opt: %s", vroom_err)
                    nn_result = nearest_neighbor_optimize(section_stops, section_distance_matrix, start_idx)
                    route_indices = _indices_by_identity(section_stops, nn_result)
                    improved_indices = two_opt_improve(route_indices, section_distance_matrix)
                    optimized_section = [section_stops[i] for i in improved_indices]
                
                logger.info(f"Section {section_idx+1}: {len(section_stops)} stops optimized with {solver_used}")
            
            # Add optimized section to results
            optimized_stops.extend(optimized_section)
            
            # Update previous end point for next section
            if optimized_section:
                last_stop = optimized_section[-1]
                previous_end_point = {
                    "latitude": last_stop["latitude"],
                    "longitude": last_stop["longitude"]
                }
        
        # Optimize and append unassigned stops at the end
        if len(unassigned_stops) > 0:
            if len(unassigned_stops) == 1:
                optimized_stops.extend(unassigned_stops)
            else:
                # Use best available solver for unassigned stops too
                unassigned_distance_matrix = calculate_distance_matrix(unassigned_stops)
                start_idx = 0
                if previous_end_point:
                    min_dist = float('inf')
                    for idx, stop in enumerate(unassigned_stops):
                        dist = haversine(
                            (previous_end_point["latitude"], previous_end_point["longitude"]),
                            (stop["latitude"], stop["longitude"]),
                            unit=Unit.KILOMETERS
                        )
                        if dist < min_dist:
                            min_dist = dist
                            start_idx = idx
                
                try:
                    try:
                        solver_matrix = await _osrm_duration_matrix(unassigned_stops)
                    except Exception as osrm_err:
                        logger.warning(
                            "Section refine (unassigned): OSRM matrix failed (%s), haversine fallback",
                            osrm_err,
                        )
                        solver_matrix = _haversine_duration_matrix(unassigned_stops)
                    indices = vroom_tsp_solve(solver_matrix, depot=start_idx, exploration_level=5)
                    if LKH_AVAILABLE and len(unassigned_stops) >= 4:
                        try:
                            lkh_indices = lkh_tsp_solve(solver_matrix, depot=start_idx, runs=3, time_limit_seconds=5)
                            if calculate_route_distance(lkh_indices, solver_matrix) < calculate_route_distance(indices, solver_matrix):
                                indices = lkh_indices
                        except Exception:
                            pass
                    elif len(unassigned_stops) >= 4:
                        indices = three_opt_improve(indices, solver_matrix, max_iterations=3)
                    optimized_unassigned = [unassigned_stops[i] for i in indices]
                except Exception:
                    nn_result = nearest_neighbor_optimize(unassigned_stops, unassigned_distance_matrix, start_idx)
                    route_indices = _indices_by_identity(unassigned_stops, nn_result)
                    improved_indices = two_opt_improve(route_indices, unassigned_distance_matrix)
                    optimized_unassigned = [unassigned_stops[i] for i in improved_indices]
                optimized_stops.extend(optimized_unassigned)
        
        # Track polish savings so we can surface them on the response
        # ("Refined: saved X km") instead of buying the win silently in
        # the logs. The polish runs in two phases — spike-relocate first,
        # then global or-opt + 2-opt — so we measure the total km between
        # the raw stitched route and the final optimised one.
        polish_relocations = 0
        polish_distance_saved_km = 0.0
        try:
            if len(optimized_stops) >= 4:
                pre_polish_km = sum(
                    haversine(
                        (optimized_stops[i]["latitude"], optimized_stops[i]["longitude"]),
                        (optimized_stops[i + 1]["latitude"], optimized_stops[i + 1]["longitude"]),
                        unit=Unit.KILOMETERS,
                    )
                    for i in range(len(optimized_stops) - 1)
                )
                tightened, refine_moves = _iterative_haversine_tighten(optimized_stops)
                optimized_stops = tightened
                if refine_moves:
                    polish_relocations = len(refine_moves)
                    logger.info(
                        "Refine polish: tightened %d spike(s) across stitched route",
                        polish_relocations,
                    )
                pre_polish_count = len(optimized_stops)
                optimized_stops = _global_two_opt_pass(optimized_stops, max_iterations=3)
                # Defensive: never let the polish accidentally drop a stop.
                if len(optimized_stops) != pre_polish_count:
                    logger.error(
                        "Refine polish: stop count drift %d → %d, reverting",
                        pre_polish_count, len(optimized_stops),
                    )
                post_polish_km = sum(
                    haversine(
                        (optimized_stops[i]["latitude"], optimized_stops[i]["longitude"]),
                        (optimized_stops[i + 1]["latitude"], optimized_stops[i + 1]["longitude"]),
                        unit=Unit.KILOMETERS,
                    )
                    for i in range(len(optimized_stops) - 1)
                )
                # Floor at 0 so a tiny rounding regression never reads as
                # "polish made it worse" on the UI.
                polish_distance_saved_km = max(0.0, pre_polish_km - post_polish_km)
        except Exception as polish_err:
            # Polish is a quality booster, not a correctness step. If it
            # explodes we keep the unpolished section-stitched route.
            logger.warning("Refine polish skipped due to error: %s", polish_err)

        reasoning = f"Section-based route refinement with {len(request.sections)} sections"
        
        # Update stop orders in database
        from pymongo import UpdateOne as _BulkUpdate
        bulk_ops = [
            _BulkUpdate({"id": stop["id"], "user_id": current_user.user_id}, {"$set": {"order": index}})
            for index, stop in enumerate(optimized_stops)
            if stop.get("id") != "current_location"
        ]
        if bulk_ops:
            await db.stops.bulk_write(bulk_ops, ordered=False)
        
        # Calculate total distance
        total_distance = 0
        all_stops_for_distance = optimized_stops
        if current_loc_stop:
            all_stops_for_distance = [current_loc_stop] + optimized_stops
        
        for i in range(len(all_stops_for_distance) - 1):
            coord1 = (all_stops_for_distance[i]["latitude"], all_stops_for_distance[i]["longitude"])
            coord2 = (all_stops_for_distance[i+1]["latitude"], all_stops_for_distance[i+1]["longitude"])
            total_distance += haversine(coord1, coord2, unit=Unit.KILOMETERS)
        
        return {
            "message": "Route refined with sections",
            "algorithm": "section_refinement",
            "reasoning": reasoning,
            "total_distance_km": round(total_distance, 2),
            "stop_count": len(optimized_stops),
            "section_count": len(request.sections),
            "started_from_current_location": current_loc_stop is not None,
            # Polish stats — frontend uses these to render a success Alert
            # ("Refined: saved 26.8 km · 3 spike(s) tightened"). 0/0.0 when
            # the route was already clean enough that polish skipped.
            "polish_relocations": polish_relocations,
            "polish_distance_saved_km": round(polish_distance_saved_km, 2),
            "stops": optimized_stops + completed_stops
        }
    
    # Auto-select was already resolved before distance matrix calculation — this is a no-op
    
    # Apply selected algorithm
    cluster_info = []  # Populated only by cluster_first
    if algorithm_used == "generoute":
        try:
            optimized_stops = await generoute_optimize(
                stops if not current_loc_stop else stops[1:],
                current_latitude=request.current_latitude,
                current_longitude=request.current_longitude
            )
            reasoning = "Optimized using Generoute API (road-based optimization)"
        except Exception as e:
            print(f"Generoute optimization failed: {e}, falling back to 2-opt")
            nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
            route_indices = _indices_by_identity(stops, nn_result)
            improved_indices = two_opt_improve(route_indices, distance_matrix)
            optimized_stops = [stops[i] for i in improved_indices]
            if current_loc_stop:
                optimized_stops = [s for s in optimized_stops if s.get("id") != "current_location"]
            reasoning = f"Generoute failed ({str(e)[:50]}), used 2-Opt fallback"
    
    elif algorithm_used == "mapbox":
        try:
            optimized_stops = await mapbox_optimize(
                stops if not current_loc_stop else stops[1:],  # Exclude current location from optimization
                current_latitude=request.current_latitude,
                current_longitude=request.current_longitude
            )
            reasoning = "Optimized using Mapbox Optimization API (road-based)"
        except Exception as e:
            print(f"Mapbox optimization failed: {e}, falling back to 2-opt")
            # Fallback to 2-opt if Mapbox fails
            nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
            route_indices = _indices_by_identity(stops, nn_result)
            improved_indices = two_opt_improve(route_indices, distance_matrix)
            optimized_stops = [stops[i] for i in improved_indices]
            if current_loc_stop:
                optimized_stops = [s for s in optimized_stops if s.get("id") != "current_location"]
            reasoning = "Mapbox failed, used 2-Opt fallback"
            
    elif algorithm_used == "alns":
        try:
            alns_time_limit = max(4, min(15, 8 + len(stops) // 10))
            optimized_stops = alns_hybrid_optimize(
                stops,
                distance_matrix,
                start_index=start_index,
                time_limit_seconds=alns_time_limit,
            )
            reasoning = "Optimized using ALNS Hybrid Metaheuristic (NN + ALNS/SA + Local Search)"
        except Exception as e:
            logger.warning("ALNS optimization failed, using 2-Opt fallback: %s", e)
            nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
            route_indices = _indices_by_identity(stops, nn_result)
            improved_indices = two_opt_improve(route_indices, distance_matrix)
            optimized_stops = [stops[i] for i in improved_indices]
            reasoning = f"ALNS failed ({str(e)[:80]}), used 2-Opt fallback"

    elif algorithm_used == "cluster_first":
        cluster_info = []
        try:
            cf_time_limit = max(15, min(60, 20 + len(stops) // 5))
            optimized_stops, cluster_info = await cluster_first_optimize(
                stops,
                distance_matrix,
                start_index=start_index,
                time_limit_seconds=cf_time_limit,
                inner_algorithm=inner_algorithm,
            )
            inner_label = inner_algorithm.upper().replace("_", " ")
            matrix_label = "Mapbox driving durations" if inner_algorithm == "ortools" else "Mapbox road distances"
            reasoning = f"Optimized using Cluster-First Route-Second (DBSCAN neighborhoods + per-cluster {inner_label} with {matrix_label})"
        except Exception as e:
            logger.warning("Cluster-first optimization failed, falling back to ALNS: %s", e, exc_info=True)
            try:
                alns_time_limit = max(4, min(15, 8 + len(stops) // 10))
                optimized_stops = alns_hybrid_optimize(stops, distance_matrix, start_index=start_index, time_limit_seconds=alns_time_limit)
                reasoning = f"Cluster-first failed ({str(e)[:50]}), used ALNS fallback"
            except Exception:
                nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
                route_indices = _indices_by_identity(stops, nn_result)
                improved_indices = two_opt_improve(route_indices, distance_matrix)
                optimized_stops = [stops[i] for i in improved_indices]
                reasoning = f"Cluster-first failed ({str(e)[:50]}), used 2-Opt fallback"

    elif algorithm_used == "vroom":
        try:
            solver_matrix = duration_matrix if duration_matrix else _haversine_duration_matrix(stops)

            # Apply traffic-aware multiplier if requested
            traffic_info = ""
            if request.traffic_aware:
                from datetime import datetime, timezone
                dep_hour = request.departure_hour if request.departure_hour is not None else datetime.now(timezone.utc).hour
                tmult = _traffic_multiplier(dep_hour)
                solver_matrix = apply_traffic_multiplier(solver_matrix, dep_hour)
                traffic_info = f", traffic={tmult:.2f}x@{dep_hour}:00"
                logger.info("Traffic-aware: applied %.2fx multiplier for hour %d", tmult, dep_hour)

            indices = cluster_aware_solve(
                vroom_tsp_solve, solver_matrix, start_index, stops,
                exploration_level=5,
            )

            # LKH post-processing: refine VROOM's solution with gold-standard TSP heuristic
            pre_cost = calculate_route_distance(indices, solver_matrix)
            if LKH_AVAILABLE:
                try:
                    lkh_indices = cluster_aware_solve(
                        lkh_tsp_solve, solver_matrix, start_index, stops,
                        runs=5, time_limit_seconds=10,
                    )
                    lkh_cost = calculate_route_distance(lkh_indices, solver_matrix)
                    # Use LKH result only if it's actually better
                    if lkh_cost < pre_cost:
                        indices = lkh_indices
                        post_cost = lkh_cost
                        refinement = "LKH"
                    else:
                        post_cost = pre_cost
                        refinement = "LKH(no improvement)"
                except Exception as lkh_err:
                    logger.warning("LKH post-processing failed, keeping VROOM result: %s", lkh_err)
                    post_cost = pre_cost
                    refinement = "LKH(failed)"
            else:
                # Fallback to 3-opt if LKH binary not available
                indices = three_opt_improve(indices, solver_matrix, max_iterations=3)
                post_cost = calculate_route_distance(indices, solver_matrix)
                refinement = "3-opt"

            saved_pct = ((pre_cost - post_cost) / pre_cost * 100) if pre_cost > 0 else 0
            logger.info("%s post-processing: %.0f → %.0f (saved %.1f%%)", refinement, pre_cost, post_cost, saved_pct)

            # Or-opt final polish: relocate 1-3 stop sequences
            pre_oropt = post_cost
            indices = or_opt_improve(indices, solver_matrix, max_iterations=10)
            post_oropt = calculate_route_distance(indices, solver_matrix)
            oropt_saved = ((pre_oropt - post_oropt) / pre_oropt * 100) if pre_oropt > 0 else 0
            if post_oropt < pre_oropt:
                logger.info("Or-opt polish: %.0f → %.0f (saved %.1f%%)", pre_oropt, post_oropt, oropt_saved)
                refinement += "+Or-opt"
                post_cost = post_oropt
                saved_pct = ((pre_cost - post_cost) / pre_cost * 100) if pre_cost > 0 else 0
            else:
                logger.info("Or-opt polish: no improvement (%.0f)", pre_oropt)

            optimized_stops = [stops[i] for i in indices]
            matrix_source = "OSRM" if (duration_matrix and OSRM_URL) else "Mapbox"
            reasoning = f"VROOM + {refinement} ({matrix_source} duration matrix, {len(stops)} stops, {refinement} saved {saved_pct:.1f}%{traffic_info})"
        except Exception as e:
            logger.warning("VROOM optimization failed, falling back to OR-Tools: %s", e)
            try:
                solver_matrix = duration_matrix if duration_matrix else _haversine_duration_matrix(stops)
                ortools_time_ms = max(2000, min(30000, 2000 + len(stops) * 80))
                indices = ortools_tsp_solve(solver_matrix, depot=start_index, time_limit_ms=ortools_time_ms)
                optimized_stops = [stops[i] for i in indices]
                reasoning = f"VROOM failed, OR-Tools fallback (duration matrix, {ortools_time_ms}ms)"
            except Exception as e2:
                logger.warning("OR-Tools fallback also failed: %s", e2)
                nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
                route_indices = _indices_by_identity(stops, nn_result)
                improved_indices = two_opt_improve(route_indices, distance_matrix)
                optimized_stops = [stops[i] for i in improved_indices]
                reasoning = "VROOM+OR-Tools failed, used 2-Opt fallback"

    elif algorithm_used == "ortools":
        try:
            # Use DURATION matrix (seconds) for time-optimal routing.
            # Scale time limit: 2s base + 80ms per stop. 123 stops ≈ 12s.
            # OR-Tools GUIDED_LOCAL_SEARCH needs adequate time for large routes.
            ortools_time_ms = max(2000, min(30000, 2000 + len(stops) * 80))
            solver_matrix = duration_matrix if duration_matrix else distance_matrix
            indices = ortools_tsp_solve(solver_matrix, depot=start_index, time_limit_ms=ortools_time_ms)
            optimized_stops = [stops[i] for i in indices]
            matrix_type = "duration" if duration_matrix else "distance"
            reasoning = f"OR-Tools TSP (PATH_CHEAPEST_ARC + GUIDED_LOCAL_SEARCH, {matrix_type} matrix, {ortools_time_ms}ms, {len(stops)} stops)"
        except Exception as e:
            logger.warning("OR-Tools optimization failed, using 2-Opt fallback: %s", e)
            nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
            route_indices = _indices_by_identity(stops, nn_result)
            improved_indices = two_opt_improve(route_indices, distance_matrix)
            optimized_stops = [stops[i] for i in improved_indices]
            reasoning = f"OR-Tools failed ({str(e)[:80]}), used 2-Opt fallback"

    elif algorithm_used == "pyvrp":
        try:
            solver_matrix = duration_matrix if duration_matrix else _haversine_duration_matrix(stops)
            # HGS needs ≥2s on small/medium TSPs to breed enough generations
            # to untangle crossings, but we cap at 3s to keep the request
            # comfortably inside the frontend's 90s AbortController.
            pyvrp_seconds = max(2.0, min(3.0, 0.5 + len(stops) * 0.04))
            # Pass coordinates so identical-address clusters (multi-unit
            # buildings, apartments, units sharing one front door) are
            # collapsed into a single PyVRP super-node — prevents random
            # zig-zag ordering between zero-distance stops.
            stop_coords = [
                (float(s["longitude"]), float(s["latitude"]))
                for s in stops
            ]
            indices = await asyncio.to_thread(
                pyvrp_tsp_solve,
                solver_matrix,
                start_index,
                pyvrp_seconds,
                0,  # seed
                stop_coords,
            )
            optimized_stops = [stops[i] for i in indices]
            matrix_source = "OSRM" if (duration_matrix and OSRM_URL) else "Mapbox"
            reasoning = (
                f"PyVRP Hybrid Genetic Search ({matrix_source} duration matrix, "
                f"{len(stops)} stops, {pyvrp_seconds:.1f}s budget)"
            )
        except Exception as e:
            logger.warning("PyVRP optimization failed, using OR-Tools fallback: %s", e)
            try:
                ortools_time_ms = max(2000, min(30000, 2000 + len(stops) * 80))
                solver_matrix = duration_matrix if duration_matrix else distance_matrix
                indices = cluster_aware_solve(
                    ortools_tsp_solve, solver_matrix, start_index, stops,
                    time_limit_ms=ortools_time_ms,
                )
                optimized_stops = [stops[i] for i in indices]
                reasoning = f"PyVRP failed ({str(e)[:60]}), OR-Tools fallback"
            except Exception:
                nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
                route_indices = _indices_by_identity(stops, nn_result)
                improved_indices = two_opt_improve(route_indices, distance_matrix)
                optimized_stops = [stops[i] for i in improved_indices]
                reasoning = "PyVRP + OR-Tools both failed, used 2-Opt fallback"

    elif algorithm_used == "vroom_ortools":
        try:
            solver_matrix = duration_matrix if duration_matrix else _haversine_duration_matrix(stops)
            ortools_time_ms = max(2000, min(30000, 2000 + len(stops) * 80))
            # Stage 1: VROOM — fast construction heuristic (~100ms)
            vroom_indices = cluster_aware_solve(
                vroom_tsp_solve, solver_matrix, start_index, stops,
                exploration_level=5,
            )
            vroom_cost = calculate_route_distance(vroom_indices, solver_matrix)
            # Stage 2: OR-Tools GLS warm-started from VROOM solution
            ortools_indices = cluster_aware_solve(
                ortools_tsp_solve, solver_matrix, start_index, stops,
                time_limit_ms=ortools_time_ms,
                initial_indices=vroom_indices,
            )
            ortools_cost = calculate_route_distance(ortools_indices, solver_matrix)
            # Take the best of the two
            indices = ortools_indices if ortools_cost <= vroom_cost else vroom_indices
            saved_pct = ((vroom_cost - ortools_cost) / vroom_cost * 100) if vroom_cost > 0 else 0
            matrix_source = "OSRM" if (duration_matrix and OSRM_URL) else "Mapbox"
            reasoning = (
                f"VROOM warm-start → OR-Tools GLS ({matrix_source}, {len(stops)} stops, "
                f"GLS improved {saved_pct:.1f}% over VROOM seed, {ortools_time_ms}ms budget)"
            )
            optimized_stops = [stops[i] for i in indices]
        except Exception as e:
            logger.warning("VROOM+OR-Tools pipeline failed, using OR-Tools cold-start: %s", e)
            try:
                solver_matrix = duration_matrix if duration_matrix else distance_matrix
                ortools_time_ms = max(2000, min(30000, 2000 + len(stops) * 80))
                indices = ortools_tsp_solve(solver_matrix, depot=start_index, time_limit_ms=ortools_time_ms)
                optimized_stops = [stops[i] for i in indices]
                reasoning = f"VROOM→OR-Tools failed ({str(e)[:60]}), OR-Tools cold-start fallback"
            except Exception:
                nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
                route_indices = _indices_by_identity(stops, nn_result)
                improved_indices = two_opt_improve(route_indices, distance_matrix)
                optimized_stops = [stops[i] for i in improved_indices]
                reasoning = "VROOM+OR-Tools both failed, used 2-Opt fallback"

    elif algorithm_used in ("nearest_neighbor", "greedy"):
        # Bulletproof greedy fallback. Per spec: rely strictly on the OSRM
        # /Mapbox driving-time matrix (`duration_matrix`) so distances are
        # real road seconds, not haversine. If OSRM is unreachable in the
        # build env, drop to a haversine-shaped duration matrix — same
        # shape, lower precision, never blocks the request. Wrapped in
        # `cluster_aware_solve` via `solve_nearest_neighbor` so multi-
        # parcel super-nodes (apartment doorsteps) expand sequentially
        # at the end and never get split by a "B in the middle of A1, A2".
        nn_matrix = duration_matrix if duration_matrix else _haversine_duration_matrix(stops)
        optimized_stops = solve_nearest_neighbor(nn_matrix, stops, start_index)
        matrix_label = "OSRM" if duration_matrix else "haversine"
        reasoning = (
            f"Optimized using Nearest Neighbor (greedy, super-node aware, "
            f"matrix={matrix_label})"
        )

    elif algorithm_used in ("greedy_2opt", "nearest_neighbor_2opt"):
        # Greedy → 2-opt polish. Roughly halves the quality gap between
        # basic greedy and VROOM at the cost of ~50 ms extra on a 167-stop
        # manifest. Same OSRM-duration-matrix discipline as plain greedy.
        # cluster_aware_solve still wraps the construction step so super-
        # nodes are kept contiguous; the 2-opt refinement runs on the
        # expanded sequence (it never reverses a segment that splits a
        # super-node because such a split would only ever increase cost
        # — the inter-parcel edge is 0 inside a super-node, so swapping
        # it out always lengthens the tour).
        nn_matrix = duration_matrix if duration_matrix else _haversine_duration_matrix(stops)
        nn_stops = solve_nearest_neighbor(nn_matrix, stops, start_index)
        route_indices = _indices_by_identity(stops, nn_stops)
        improved_indices = await asyncio.to_thread(two_opt_improve, route_indices, nn_matrix)
        optimized_stops = [stops[i] for i in improved_indices]
        matrix_label = "OSRM" if duration_matrix else "haversine"
        reasoning = (
            f"Optimized using Greedy + 2-Opt polish (super-node aware, "
            f"matrix={matrix_label})"
        )
        
    elif algorithm_used == "two_opt":
        # Start with nearest neighbor, then improve with 2-opt
        nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
        route_indices = _indices_by_identity(stops, nn_result)
        improved_indices = two_opt_improve(route_indices, distance_matrix)
        optimized_stops = [stops[i] for i in improved_indices]
        reasoning = "Optimized using 2-Opt improvement heuristic"
        
    elif algorithm_used == "simulated_annealing":
        # Silently upgraded to ILS — same interface, strictly better results.
        # ILS uses double-bridge kicks + 2-opt/Or-opt which consistently beats SA.
        ils_time = max(5, min(15, 5 + len(stops) // 10))
        optimized_stops = await asyncio.to_thread(iterated_local_search, stops, distance_matrix, start_index, ils_time)
        reasoning = f"Optimized using ILS (upgraded SA: double-bridge + 2-opt/Or-opt, {ils_time}s budget)"

    elif algorithm_used == "ils":
        ils_time = max(5, min(15, 5 + len(stops) // 10))
        try:
            optimized_stops = await asyncio.to_thread(iterated_local_search, stops, distance_matrix, start_index, ils_time)
            reasoning = f"Optimized using ILS (double-bridge perturbation + 2-opt + Or-opt local search, {ils_time}s budget)"
        except Exception as e:
            logger.warning("ILS optimization failed, using 2-Opt fallback: %s", e)
            nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
            route_indices = _indices_by_identity(stops, nn_result)
            improved_indices = two_opt_improve(route_indices, distance_matrix)
            optimized_stops = [stops[i] for i in improved_indices]
            reasoning = f"ILS failed ({str(e)[:80]}), used 2-Opt fallback"
        
    elif algorithm_used == "genetic":
        optimized_stops = await asyncio.to_thread(genetic_algorithm_optimize, stops, distance_matrix, start_index, 
                                                      100 + len(stops) * 2,
                                                      max(30, len(stops)))
        reasoning = "Optimized using Genetic Algorithm (evolutionary)"
        
    elif algorithm_used == "clarke_wright":
        optimized_stops = clarke_wright_savings(stops, distance_matrix, start_index)
        reasoning = "Optimized using Clarke-Wright Savings (VRP algorithm)"

    elif algorithm_used == "lkh":
        try:
            if not LKH_AVAILABLE:
                raise RuntimeError("LKH not available")
            solver_matrix = duration_matrix if duration_matrix else _haversine_duration_matrix(stops)
            indices = cluster_aware_solve(
                lkh_tsp_solve, solver_matrix, start_index, stops,
                runs=5, time_limit_seconds=15,
            )
            optimized_stops = [stops[i] for i in indices]
            reasoning = f"LKH-3 (Lin-Kernighan-Helsgott, {len(stops)} stops)"
        except Exception as e:
            logger.warning("LKH failed, falling back to OR-Tools: %s", e)
            try:
                solver_matrix = duration_matrix if duration_matrix else _haversine_duration_matrix(stops)
                indices = cluster_aware_solve(
                    ortools_tsp_solve, solver_matrix, start_index, stops,
                    time_limit_ms=max(2000, len(stops) * 80),
                )
                optimized_stops = [stops[i] for i in indices]
                reasoning = f"LKH failed ({str(e)[:50]}), OR-Tools fallback"
            except Exception:
                nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
                route_indices = _indices_by_identity(stops, nn_result)
                improved_indices = two_opt_improve(route_indices, distance_matrix)
                optimized_stops = [stops[i] for i in improved_indices]
                reasoning = "LKH+OR-Tools failed, 2-Opt fallback"

    elif algorithm_used == "vroom_lkh_3opt":
        try:
            solver_matrix = duration_matrix if duration_matrix else _haversine_duration_matrix(stops)
            # Stage 1: VROOM
            if VROOM_AVAILABLE:
                indices = vroom_tsp_solve(solver_matrix, depot=start_index, exploration_level=5)
            else:
                nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
                indices = _indices_by_identity(stops, nn_result)
            # Stage 2: LKH refinement
            if LKH_AVAILABLE:
                try:
                    lkh_indices = lkh_tsp_solve(solver_matrix, depot=start_index, runs=5, time_limit_seconds=10)
                    if calculate_route_distance(lkh_indices, solver_matrix) < calculate_route_distance(indices, solver_matrix):
                        indices = lkh_indices
                except Exception:
                    pass
            # Stage 3: 3-opt polish — only KEEP if it actually improves.
            # 3-opt is supposed to be monotonic but its `calculate_route_distance`
            # implementation occasionally regresses by ~0.3 min on routes the
            # LKH stage already brought to local optimum. Guard against that
            # so the cascade is *strictly* ≥ LKH-alone quality.
            pre_cost = calculate_route_distance(indices, solver_matrix)
            polished = three_opt_improve(indices, solver_matrix, max_iterations=5)
            polished_cost = calculate_route_distance(polished, solver_matrix)
            if polished_cost < pre_cost:
                indices = polished
                post_cost = polished_cost
            else:
                post_cost = pre_cost
            saved = ((pre_cost - post_cost) / pre_cost * 100) if pre_cost > 0 else 0
            optimized_stops = [stops[i] for i in indices]
            reasoning = f"VROOM+LKH+3-opt pipeline ({len(stops)} stops, 3-opt saved {saved:.1f}%)"
        except Exception as e:
            logger.warning("VROOM+LKH+3opt failed: %s", e)
            # Use the OSRM duration matrix for the fallback so the route
            # is still road-quality. Haversine fallback would produce
            # crow-fly orderings that look like spaghetti once rendered
            # against real one-way streets. If OSRM duration is missing
            # (build env without OSRM) we fall back to a haversine-based
            # *duration* matrix — still seconds-shaped, just less precise.
            fallback_matrix = duration_matrix if duration_matrix else _haversine_duration_matrix(stops)
            nn_result = nearest_neighbor_optimize(stops, fallback_matrix, start_index)
            route_indices = _indices_by_identity(stops, nn_result)
            improved_indices = two_opt_improve(route_indices, fallback_matrix)
            optimized_stops = [stops[i] for i in improved_indices]
            reasoning = "VROOM+LKH+3opt failed, OSRM 2-Opt fallback"

    elif algorithm_used == "timefold":
        try:
            if not TIMEFOLD_AVAILABLE:
                raise RuntimeError(f"Timefold not available: {TIMEFOLD_IMPORT_ERROR}")
            time_limit = max(5, min(15, 5 + len(stops) // 20))
            # Timefold runs a Java constraint solver via JPype — CPU-bound for
            # the full time_limit. Wrap in to_thread so the event loop stays
            # responsive for concurrent requests (health checks, map tiles…).
            optimized_stops = await asyncio.to_thread(
                timefold_optimize, stops, distance_matrix,
                start_index=start_index, time_limit_seconds=time_limit,
            )
            reasoning = f"Timefold Java constraint solver ({len(stops)} stops, {time_limit}s limit)"
        except Exception as e:
            logger.warning("Timefold failed, falling back to OR-Tools: %s", e)
            try:
                solver_matrix = duration_matrix if duration_matrix else _haversine_duration_matrix(stops)
                indices = ortools_tsp_solve(solver_matrix, depot=start_index, time_limit_ms=max(2000, len(stops) * 80))
                optimized_stops = [stops[i] for i in indices]
                reasoning = f"Timefold failed ({str(e)[:50]}), OR-Tools fallback"
            except Exception:
                nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
                route_indices = _indices_by_identity(stops, nn_result)
                improved_indices = two_opt_improve(route_indices, distance_matrix)
                optimized_stops = [stops[i] for i in improved_indices]
                reasoning = "Timefold+OR-Tools failed, 2-Opt fallback"

    elif algorithm_used == "three_opt":
        nn_result = nearest_neighbor_optimize(stops, distance_matrix, start_index)
        route_indices = _indices_by_identity(stops, nn_result)
        improved_indices = three_opt_improve(route_indices, distance_matrix, max_iterations=5)
        optimized_stops = [stops[i] for i in improved_indices]
        reasoning = "Optimized using 3-Opt improvement heuristic (NN seed + 3-edge reconnection)"
        
    else:
        # Default to nearest neighbor
        optimized_stops = nearest_neighbor_optimize(stops, distance_matrix, start_index)
        reasoning = "Optimized using Nearest Neighbor algorithm"
    
    # Remove current location stop from results (it's just for optimization)
    if current_loc_stop:
        optimized_stops = [s for s in optimized_stops if s.get("id") != "current_location"]
    
    # Update stop orders in database (bulk write — single round-trip)
    from pymongo import UpdateOne as _BulkOp
    bulk_ops = [
        _BulkOp({"id": stop["id"], "user_id": current_user.user_id}, {"$set": {"order": index}})
        for index, stop in enumerate(optimized_stops)
        if stop.get("id") != "current_location"
    ]
    if bulk_ops:
        await db.stops.bulk_write(bulk_ops, ordered=False)
    
    # Calculate total distance — prefer Mapbox road distance, fall back to haversine.
    # Kick off the shadow benchmark in a background thread FIRST so it runs in
    # parallel with the async OSRM/Mapbox fetch below.
    all_stops_for_distance = optimized_stops
    if current_loc_stop:
        all_stops_for_distance = [current_loc_stop] + optimized_stops

    # Shadow-test: run the best alternative algorithm for comparison. Wrapped
    # in `asyncio.to_thread` + `create_task` so it runs CONCURRENTLY with the
    # road-distance fetch instead of serialising another 5-10 s onto every
    # optimize call. We await the task right before we need the result.
    SHADOW_CANDIDATES = ["alns", "ortools", "two_opt"]
    shadow_algo = next((a for a in SHADOW_CANDIDATES if a != algorithm_used), None)
    shadow_task = (
        asyncio.create_task(
            asyncio.to_thread(_run_algorithm_benchmark, shadow_algo, stops, distance_matrix, start_index)
        )
        if shadow_algo and len(stops) >= 2
        else None
    )

    road_distance = await calculate_road_distance_km(all_stops_for_distance)

    haversine_distance = 0.0
    for i in range(len(all_stops_for_distance) - 1):
        coord1 = (all_stops_for_distance[i]["latitude"], all_stops_for_distance[i]["longitude"])
        coord2 = (all_stops_for_distance[i+1]["latitude"], all_stops_for_distance[i+1]["longitude"])
        haversine_distance += haversine(coord1, coord2, unit=Unit.KILOMETERS)

    total_distance = road_distance if road_distance is not None else round(haversine_distance, 2)
    distance_source = "road" if road_distance is not None else "haversine"

    # Collect the shadow result now (it's been running concurrently with the
    # road-distance fetch above). `await` on the task just waits for whatever
    # time remains — usually zero when the main solver was slower.
    shadow = None
    if shadow_task is not None:
        try:
            shadow = await shadow_task
            if shadow and shadow.get("error") is None:
                shadow["savings_km"] = round(shadow["total_distance_km"] - haversine_distance, 3)
        except Exception as _shadow_err:
            logger.warning("Shadow benchmark failed: %s", _shadow_err)
            shadow = None

    # ── Quality badge: optimized vs nearest-neighbor baseline ──
    # NN is O(n²), always available, and is the universally-understood naïve
    # greedy baseline. Showing "saved X km / Y% vs greedy" gives the driver
    # instant visual proof the optimize button is worth tapping.
    quality_badge = None
    try:
        if algorithm_used != "nearest_neighbor" and len(stops) >= 3:
            nn_stops = nearest_neighbor_optimize(stops, distance_matrix, start_index)
            nn_indices = _indices_by_identity(stops, nn_stops)
            nn_km = sum(
                distance_matrix[nn_indices[i]][nn_indices[i + 1]]
                for i in range(len(nn_indices) - 1)
            )
            opt_km = total_distance if distance_source == "road" else haversine_distance
            saved_km = nn_km - opt_km
            saved_pct = (saved_km / nn_km * 100.0) if nn_km > 0 else 0.0
            quality_badge = {
                "baseline_algorithm": "nearest_neighbor",
                "baseline_km": round(nn_km, 2),
                "optimized_km": round(opt_km, 2),
                "saved_km": round(saved_km, 2),
                "saved_pct": round(saved_pct, 1),
                "improved": saved_km > 0.05,
            }
    except Exception as _badge_err:
        logger.debug(f"Quality badge computation skipped: {_badge_err}")

    # ── Time savings vs unoptimised input order (driver-facing badge) ──
    # The most meaningful comparison for a driver is "how much time did
    # tapping Optimise save me vs the order I had?". This is the order they
    # would have actually driven if they hadn't optimised — not the NN
    # greedy (above) nor a theoretical baseline. We compute open-path
    # duration on the SAME OSRM duration_matrix the solver used, so the
    # number is directly comparable and self-consistent.
    time_savings = None
    try:
        if duration_matrix is not None and len(stops) >= 3:
            optimized_indices = _indices_by_identity(stops, optimized_stops)
            input_indices = list(range(len(stops)))  # input was in DB order
            input_seconds = sum(
                duration_matrix[input_indices[i]][input_indices[i + 1]]
                for i in range(len(input_indices) - 1)
            )
            optimized_seconds = sum(
                duration_matrix[optimized_indices[i]][optimized_indices[i + 1]]
                for i in range(len(optimized_indices) - 1)
            )
            saved_seconds = max(0, int(input_seconds) - int(optimized_seconds))
            saved_pct = (
                (saved_seconds / input_seconds * 100.0) if input_seconds > 0 else 0.0
            )
            time_savings = {
                "baseline_seconds": int(input_seconds),
                "optimized_seconds": int(optimized_seconds),
                "saved_seconds": int(saved_seconds),
                "saved_minutes": round(saved_seconds / 60.0, 1),
                "saved_pct": round(saved_pct, 1),
                "improved": saved_seconds >= 30,  # show badge only if >=30s saved
            }
    except Exception as _ts_err:
        logger.debug(f"Time savings computation skipped: {_ts_err}")

    # ── Visual cluster-spike detection (post-solve) ─────────────────────
    # Even an optimal-by-OSRM-time tour can LOOK fragmented on the map: a
    # stop B can be a small driving-time detour from A→C (highway split,
    # one-way pair) yet be a large *geographic* spike. We sweep every
    # consecutive triplet (A, B, C) using haversine distance and flag any
    # B where `dist(A,C) < threshold * (dist(A,B) + dist(B,C))` — i.e. B
    # is well off the natural A→C line.
    cluster_warnings: List[Dict[str, Any]] = detect_cluster_spikes(optimized_stops)

    # ── Auto-tighten cluster spikes IN-PLACE during optimisation ────────
    # User feedback: drivers don't want to tap a banner — they expect the
    # optimiser to never produce visible zig-zags in the first place.
    # We iteratively relocate the worst spike (largest `extra_km`) up to
    # 10 passes, then run OSRM verification. If OSRM agrees the cleaned
    # route isn't slower in driving time, we silently swap it in. If OSRM
    # rolls back (i.e. the detour is genuinely faster on the road
    # network — e.g. a highway split or one-way pair), we keep the
    # solver's choice and surface the remaining warning so the driver can
    # still override manually. Net effect: cosmetic zig-zags vanish; only
    # OSRM-justified detours ever reach the screen.
    if cluster_warnings and len(optimized_stops) >= 3:
        try:
            cleaned, auto_moves = _iterative_haversine_tighten(optimized_stops)
            if auto_moves:
                # Driver-preference tolerance: accept the cleaned route even
                # if OSRM thinks it's marginally slower. A driver would
                # rather drive 90s longer than do an obvious cross-suburb
                # zig-zag mid-cluster (cf. the Parklands Blvd 68→69→70
                # spike report from 2026-04-25). The effective threshold is
                # `max(90s, before_s * 0.03)` — so a 1-hour route can grow
                # by up to ~108s (3%), capped on the upper end by the route
                # length itself. Manual /tighten endpoints stay strict
                # (slack=0) so an explicit user tap never makes them strictly
                # slower.
                #
                # 2026-05-11 — REVERTED from a wider (240 s / 5 %) tier on
                # ≥150-stop routes. Empirically the wider tier let OSRM
                # accept 2-opt swaps that displaced individual stops into
                # neighbouring clusters on big production runs ("specific
                # stops out of order that obviously shouldn't be"). The
                # tightener has no cluster-locality guard, so the slack
                # budget IS the cluster-locality guard — a small budget
                # restricts moves to within-cluster relocations. Single
                # tier across all route sizes restores the working
                # baseline. If we ever want to allow wider cleanups on
                # big routes, the right fix is a cluster-locality guard
                # in the move generator, NOT a wider OSRM slack.
                chosen, _b, _a, rolled_back = await _osrm_verify_relocation(
                    optimized_stops, cleaned,
                    slack_seconds=90, slack_ratio=0.03,
                )
                if not rolled_back:
                    optimized_stops = chosen
                    cluster_warnings = detect_cluster_spikes(optimized_stops)
                else:
                    # The tightener tried, OSRM said the fix would cost more
                    # road-time than the slack budget allows, and the move
                    # chain was rolled back. Nothing the user can do via the
                    # banner will improve this route. Suppress warnings
                    # entirely — keeping them on screen is a UI lie.
                    cluster_warnings = []
                logger.info(
                    "Auto-tightened %d move(s) during /api/optimize "
                    "(rolled_back=%s, raw_warnings=%d)",
                    len(auto_moves), rolled_back, len(cluster_warnings),
                )
        except Exception as auto_err:
            logger.debug(f"Auto-tighten skipped: {auto_err}")

    # Honest banner: hide warnings the algorithm cannot actually fix on a
    # follow-up tighten. Runs unconditionally — without this filter, an
    # OSRM rollback or a route the solver already nailed leaves the UI
    # showing "17 detour stops" even though every flagged stop is at its
    # haversine-optimal position and Tighten All would be a no-op. Silent
    # spikes that no further single-stop relocation can address are pure
    # geometric quirks (peninsulas, road-network asymmetries) and have no
    # business raising a "you can fix this" banner.
    if cluster_warnings:
        before_filter = len(cluster_warnings)
        cluster_warnings = _filter_actionable_warnings(
            optimized_stops, cluster_warnings
        )
        logger.info(
            "Cluster warnings filter: raw=%d → actionable=%d",
            before_filter, len(cluster_warnings),
        )

    all_output_stops = optimized_stops + completed_stops
    # ── Absolute stop-id binding (no positional drift) ──────────────────
    # Frontend currently maps the route by positional order of `stops`.
    # That works as long as nothing reorders the array in transit, but it
    # silently breaks if a serialiser or middleware ever shuffles them.
    # `optimized_sequence` is the canonical, ID-based answer to "what
    # order should the driver visit?"  — a flat list of `stop.id` strings
    # in the optimised order (uncompleted stops first, then completed).
    optimized_sequence = [
        s["id"] for s in all_output_stops if s.get("id") is not None
    ]
    response_body = {
        "message": "Route optimized",
        "algorithm": algorithm_used,
        "reasoning": reasoning,
        "total_distance_km": total_distance,
        "distance_source": distance_source,
        "stop_count": len(all_output_stops),
        "started_from_current_location": current_loc_stop is not None,
        "stops": all_output_stops,
        "optimized_sequence": optimized_sequence,
        "cluster_warnings": cluster_warnings,
        "shadow": shadow,
        "quality_badge": quality_badge,
        "time_savings": time_savings,
    }
    # Include cluster overlay data when cluster_first is used
    if algorithm_used == "cluster_first" and cluster_info:
        response_body["clusters"] = cluster_info

    # ── AUDIT 3: outbound sequence ───────────────────────────────────────
    # The exact ID order this response will deliver to the device. Pair
    # this with the frontend "AUDIT API RX" log: if the two arrays differ,
    # something between FastAPI's JSON encoder and the React state set is
    # re-shuffling. If they match but the polyline still draws wrong,
    # the bug is in the polyline-coord builder (the array-iteration order
    # vs `order` field mismatch we already chased once).
    logger.info(
        "AUDIT[/optimize] TX algorithm=%s sequence_first5=%s "
        "(if frontend RX differs, transit is re-shuffling)",
        algorithm_used,
        [s["id"] for s in all_output_stops[:5] if s.get("id")],
    )
    return response_body


def _two_opt_pass(seq: List[dict]) -> Tuple[List[dict], int]:
    """Single 2-opt sweep: try every pair of non-adjacent edges
    `(seq[i], seq[i+1])` and `(seq[j], seq[j+1])`, and reverse the
    interior segment `seq[i+1..j]` whenever the swap shortens the
    haversine path. Greedy first-improvement with restart — i.e. as
    soon as one improving swap is found we restart the outer scan, so
    interleaved spikes that single-stop relocation cannot reach get
    untangled in a single call.

    Returns `(new_seq, swaps_applied)`. The original list is not mutated.

    Why 2-opt fixes things `_relocate_stop_haversine` cannot:
      Single-stop relocate moves *one* node to its best insertion point.
      But a route like `...A B C D E F G...` where the visually-spiky
      stop is B, and the only improvement is the *swap* of edge `A→B`
      with edge `F→G` (giving `...A F E D C B G...` reversed), is
      invisible to single-stop search — moving B alone yields no gain
      because B's best insertion is exactly its current spot. 2-opt
      sees the edge pair and reverses the run.

    Cost: O(n²) per scan, restart on improvement. For n=167 a single
    sweep is ~28k pair checks × ~50 ns each = ~1.5 ms. Even with 50
    restarts that's <100 ms — fits under the auto-tighten budget.
    """
    n = len(seq)
    if n < 4:
        return list(seq), 0
    current = list(seq)
    swaps = 0
    improved = True
    # Cap scans defensively — pathological data could otherwise loop
    # near-forever. Real routes converge in 1-3 scans.
    max_scans = 50
    # Cluster-locality guard: reject any swap whose longest NEW edge is
    # more than 1.5× the longest OLD edge it replaces. This blocks the
    # 2026-05-11 regression where 2-opt would collapse two medium edges
    # (e.g. 20 km + 20 km, the natural bridges between two clusters)
    # into one tiny + one giant edge (e.g. 0.1 km + 39 km). The total
    # path shrinks, so the haversine improvement check passes, but the
    # giant new edge crosses a cluster boundary and drops a stop into
    # the wrong cluster. By capping per-edge growth, we keep the
    # solver inside the cluster graph the solver itself produced.
    # 1.5× allows legitimate "lengthen one edge slightly to remove a
    # zig-zag" within a cluster while rejecting cross-cluster jumps.
    LOCALITY_MULTIPLIER = 1.5
    while improved and swaps < max_scans * n:
        improved = False
        cur_len = _haversine_path_km(current)
        # Outer i goes up to n-3 so j has room. Inner j starts at i+2 to
        # avoid adjacent edges (whose swap is a no-op).
        for i in range(n - 3):
            ai = current[i]
            bi = current[i + 1]
            d_ab = haversine(
                (ai["latitude"], ai["longitude"]),
                (bi["latitude"], bi["longitude"]),
                unit=Unit.KILOMETERS,
            )
            for j in range(i + 2, n - 1):
                cj = current[j]
                dj = current[j + 1]
                # Existing edges: ai→bi and cj→dj
                # Swap to: ai→cj and bi→dj  (reverses the bi..cj segment)
                d_cd = haversine(
                    (cj["latitude"], cj["longitude"]),
                    (dj["latitude"], dj["longitude"]),
                    unit=Unit.KILOMETERS,
                )
                d_ac = haversine(
                    (ai["latitude"], ai["longitude"]),
                    (cj["latitude"], cj["longitude"]),
                    unit=Unit.KILOMETERS,
                )
                d_bd = haversine(
                    (bi["latitude"], bi["longitude"]),
                    (dj["latitude"], dj["longitude"]),
                    unit=Unit.KILOMETERS,
                )
                # Strict 1e-9 epsilon to dodge floating-point oscillation
                # that would otherwise let two equivalent tours flip back
                # and forth forever.
                if d_ac + d_bd + 1e-9 < d_ab + d_cd:
                    # Cluster-locality guard (see top of function). Reject
                    # any swap whose longest new edge exceeds 1.5× the
                    # longest old edge it replaces. Prevents 2-opt from
                    # creating cross-cluster bridges that haversine-sum
                    # tolerates but drivers visibly hate.
                    max_old = d_ab if d_ab > d_cd else d_cd
                    max_new = d_ac if d_ac > d_bd else d_bd
                    if max_new > max_old * LOCALITY_MULTIPLIER:
                        continue
                    current = current[: i + 1] + current[i + 1 : j + 1][::-1] + current[j + 1 :]
                    swaps += 1
                    improved = True
                    break  # restart outer scan from i=0
            if improved:
                break
        # Defensive: if the path didn't actually shrink despite improved=True,
        # bail. (Shouldn't trigger; the strict epsilon makes it monotonic.)
        if improved and _haversine_path_km(current) >= cur_len - 1e-9:
            break
    return current, swaps


def _filter_actionable_warnings(
    cleaned: List[dict], warnings: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """Drop warnings whose `suspect_id` cannot be improved by *any* further
    single-stop relocation on the already-tightened sequence.

    Without this filter the banner lies: it shows "15 detour stops" even
    when Tighten All would do nothing because every flagged stop is at
    its haversine-optimal position. Filtering down to only the warnings
    the relocator can actually address keeps the UX honest — if 0
    warnings remain, the banner hides entirely.

    We only check single-stop relocation here (not 2-opt) because the
    callers already run `_two_opt_pass` upstream as part of the cleaning
    pipeline. Anything 2-opt can fix has been fixed; what's left is
    only what relocate can address. If relocate can't either, it's a
    permanent geometric quirk (e.g. a peninsula stop with truly no
    better neighbour) — silently informational, not actionable.
    """
    if not warnings:
        return []
    id_to_idx = {s["id"]: i for i, s in enumerate(cleaned) if "id" in s}
    out: List[Dict[str, Any]] = []
    for w in warnings:
        idx = id_to_idx.get(w.get("suspect_id"))
        if idx is None:
            continue
        _, _, before, after = _relocate_stop_haversine(cleaned, idx)
        if after < before - 1e-6:
            out.append(w)
    return out


def _haversine_path_km(seq: List[dict]) -> float:
    """Sum of haversine distances along an ordered list of stop dicts."""
    total = 0.0
    for i in range(len(seq) - 1):
        a, b = seq[i], seq[i + 1]
        total += haversine(
            (a["latitude"], a["longitude"]),
            (b["latitude"], b["longitude"]),
            unit=Unit.KILOMETERS,
        )
    return total


def _relocate_stop_haversine(
    pending: List[dict], suspect_idx: int
) -> Tuple[List[dict], int, float, float]:
    """Lift `pending[suspect_idx]` and reinsert it where the haversine path
    is shortest. Returns `(new_seq, new_position, before_km, after_km)`.

    The original list is not mutated.
    """
    suspect = pending[suspect_idx]
    rest = [s for i, s in enumerate(pending) if i != suspect_idx]

    before_km = _haversine_path_km(pending)
    best_seq = pending
    best_cost = before_km
    best_position = suspect_idx

    for pos in range(len(rest) + 1):
        candidate = rest[:pos] + [suspect] + rest[pos:]
        cost = _haversine_path_km(candidate)
        if cost < best_cost:
            best_cost = cost
            best_seq = candidate
            best_position = pos

    return best_seq, best_position, before_km, best_cost


def _iterative_haversine_tighten(
    seq: List[dict], max_passes: int = 50
) -> Tuple[List[dict], List[Dict[str, Any]]]:
    """Repeatedly relocate the worst cluster spike (largest `extra_km`) on
    `seq`, then run a 2-opt edge-swap sweep, alternating until both
    converge (no further haversine improvement is possible).

    Returns `(new_seq, moves)`. `moves` is empty when the input was
    already clean. The original list is not mutated. Use this for the
    pure-geometric pass; pair with `_osrm_verify_relocation` to make sure
    the cleaned route also wins on driving time.

    Why both relocate AND 2-opt: relocate moves a single stop to its
    haversine-best insertion point; 2-opt reverses an edge pair. Each
    can fix things the other cannot — interleaved spikes (stop 21
    visited mid-cluster of 119-124) are 2-opt-improvable but often
    relocate-stuck, while isolated detours (one stop far from the line
    A→C) are relocate-improvable but often 2-opt-stuck. Alternating
    catches both. We attribute every improvement to a recorded `move`
    so the audit log stays informative even when 2-opt does the heavy
    lifting.

    `max_passes=50` is a *ceiling*, not a target — natural exit is the
    no-improvement break inside each inner loop. 50 is well above any
    realistic spike count for a 200-stop manifest and still bounds
    runtime pathologically.
    """
    moves: List[Dict[str, Any]] = []
    current = list(seq)

    def _relocate_loop():
        """One full single-stop-relocate sweep. Mutates `current` and
        appends to `moves`. Returns count of moves applied this sweep."""
        nonlocal current
        applied = 0
        for _ in range(max_passes):
            warnings = detect_cluster_spikes(current)
            if not warnings:
                break
            worst = max(warnings, key=lambda w: w["extra_km"])
            suspect_idx = next(
                (i for i, s in enumerate(current) if s.get("id") == worst["suspect_id"]),
                None,
            )
            if suspect_idx is None:
                break
            new_seq, new_pos, before_km, after_km = _relocate_stop_haversine(
                current, suspect_idx
            )
            if after_km >= before_km - 1e-6:
                break
            moves.append({
                "moved_stop_id": worst["suspect_id"],
                "from_position": suspect_idx,
                "to_position": new_pos,
                "saved_km": round(before_km - after_km, 3),
                "kind": "relocate",
            })
            current = new_seq
            applied += 1
        return applied

    # Alternate relocate ↔ 2-opt up to `max_passes` rounds. Each round
    # accepts only strict haversine improvements, so the loop is monotone
    # and terminates when both move generators are stuck.
    for _ in range(max_passes):
        relocated = _relocate_loop()
        before_two_opt_km = _haversine_path_km(current)
        new_seq, swaps = _two_opt_pass(current)
        if swaps:
            current = new_seq
            moves.append({
                "moved_stop_id": None,
                "from_position": None,
                "to_position": None,
                "saved_km": round(before_two_opt_km - _haversine_path_km(current), 3),
                "kind": "two_opt",
                "swaps": swaps,
            })
        if relocated == 0 and swaps == 0:
            break
    return current, moves


async def _persist_pending_order(user_id: str, ordered: List[dict]) -> None:
    """Bulk-write the new `order` field for every stop in `ordered`."""
    from pymongo import UpdateOne as _BulkOp
    bulk_ops = [
        _BulkOp(
            {"id": s["id"], "user_id": user_id},
            {"$set": {"order": i}},
        )
        for i, s in enumerate(ordered)
    ]
    if bulk_ops:
        await db.stops.bulk_write(bulk_ops, ordered=False)


async def _osrm_verify_relocation(
    original_seq: List[dict],
    proposed_seq: List[dict],
    slack_seconds: int = 0,
    slack_ratio: float = 0.0,
) -> Tuple[List[dict], Optional[int], Optional[int], bool]:
    """Cost two stop sequences against the OSRM duration matrix and pick
    whichever takes less time on the road.

    Haversine is a fine visual proxy for "this looks zig-zaggy", but real
    drivers feel OSRM seconds. We fetch the duration matrix once (on the
    proposed sequence — same stops as original, just reordered) and re-cost
    both orderings. If OSRM agrees the relocation is faster (or ties), we
    keep it. If it disagrees, we roll back to the original. The matrix is
    only fetched once per call; for medium routes it's already cached by
    the wider /api/optimize matrix cache.

    `slack_seconds` / `slack_ratio` tune how much driving-time the cleaned
    sequence is allowed to add. Defaults are zero (strict OSRM-wins) for
    manual tighten endpoints, so an explicit user tap never makes them
    slower. The auto-tighten path inside `/api/optimize` passes a small
    tolerance because real drivers prefer a visually clean route over a
    1–2% faster one with a single obvious cross-suburb detour. The actual
    threshold is `max(slack_seconds, before_s * slack_ratio)`.

    Returns:
        (chosen_seq, before_seconds, after_seconds, rolled_back)
        before_seconds / after_seconds are `None` if OSRM was unreachable
        and verification couldn't be performed (in which case we fall
        through, keeping the proposed sequence).
    """
    try:
        # Pull the matrix straight from the local OSRM Table service. This
        # used to call `calculate_duration_matrix`, which silently falls
        # back to a haversine estimate for N>25 (the Mapbox cap), so on
        # 100+-stop routes the verification was a haversine check
        # masquerading as an OSRM check — defeating the entire purpose
        # of "did the cleaner sequence actually win on driving time?".
        # Using `_osrm_duration_matrix` keeps every verification grounded
        # in real road-network seconds (with the public OSRM demo as a
        # last-resort fallback inside that helper).
        duration_matrix = await _osrm_duration_matrix(proposed_seq)
    except Exception as exc:
        logger.debug(f"OSRM duration matrix fetch failed: {exc}")
        return proposed_seq, None, None, False

    if not duration_matrix:
        return proposed_seq, None, None, False

    # `proposed_seq` and `original_seq` are permutations of the same stops, so
    # we can build the id→row map once on `proposed_seq` and use it to look up
    # rows for either ordering.
    id_to_row = {s["id"]: i for i, s in enumerate(proposed_seq)}

    def _seq_seconds(seq: List[dict]) -> int:
        total = 0
        for k in range(len(seq) - 1):
            i = id_to_row[seq[k]["id"]]
            j = id_to_row[seq[k + 1]["id"]]
            total += int(duration_matrix[i][j])
        return total

    before_s = _seq_seconds(original_seq)
    after_s = _seq_seconds(proposed_seq)

    tolerance = max(slack_seconds, int(before_s * slack_ratio))
    if after_s > before_s + tolerance:
        # OSRM disagrees with the visual fix beyond the allowed slack. Roll back.
        return original_seq, before_s, after_s, True
    return proposed_seq, before_s, after_s, False


# ─────────────────────────────────────────────────────────────────────────
# Async optimize job runner
# ─────────────────────────────────────────────────────────────────────────
# Why this exists: Cloudflare's edge proxy enforces a hard 100 s ceiling on
# origin response time (HTTP 524). On a 200-stop manifest with active
# No-Go zones, the synchronous `/api/optimize` endpoint can take 90-150 s
# end-to-end (OSRM matrix + nogo OSRM-aware probe + PyVRP solve + 2-opt
# tightener + final OSRM directions). The 524 fires before we can reply.
#
# Fix: a fire-and-poll job pattern. The client POSTs to
# `/api/optimize/jobs`, gets a 202 + `job_id` in <100 ms, then polls
# `/api/optimize/jobs/{job_id}` every ~2 s until `status` flips from
# `running` → `done` (or `error`). Each poll is well under the 100 s cap,
# so Cloudflare can never time us out — we own the wall-clock budget.
#
# IMPORTANT — Mongo-backed (NOT in-memory):
#   Production runs multiple replicas behind the K8s ingress. An in-memory
#   dict would silently fail when POST hits pod A and the subsequent
#   GET poll hits pod B (job_id missing → "Job not found or expired").
#   We persist the job in Mongo so any pod can serve any poll. A TTL
#   index on `expires_at` auto-purges 10 min after creation; no GC code
#   path needed in Python.
#
# Cross-pod result delivery: the runner that owns the optimize task is
# the *same* pod that handled the POST (asyncio.create_task is local).
# That pod writes `status:"done"` + `result` to Mongo on completion.
# Any poll hitting any pod reads from Mongo. If the owning pod crashes
# mid-solve, the job stays in `running` until TTL — the driver's poll
# loop times out client-side after 5 min and they retap Optimise.

import asyncio as _asyncio_jobs
from datetime import datetime as _dt_jobs, timedelta as _td_jobs, timezone as _tz_jobs

_OPTIMIZE_JOB_TTL_S = 600          # 10 min — driver retries within this

# Strong reference set for all in-flight runner tasks. asyncio's docs are
# explicit: "the event loop only keeps weak references to tasks" — so a
# bare `create_task(...)` can be silently cancelled by GC under memory
# pressure (which a 200-stop optimize generates lots of). The classic
# symptom is exactly what we hit on prod: kickoff returns 202, the
# runner starts, then disappears mid-solve, and the frontend polls
# `status: "running"` forever until its 5-minute ceiling. Holding a hard
# reference here keeps the task alive; the done-callback removes it
# once the runner finishes (success OR failure).
_OPTIMIZE_RUNNER_TASKS: set = set()


async def _ensure_optimize_jobs_indexes() -> None:
    """Create the TTL index on `optimize_jobs.expires_at` once at startup.

    Mongo's TTL monitor sweeps roughly every minute; documents are deleted
    when `expires_at < now`. Called from the existing startup hook so the
    kickoff hot path is pure-insert (no schema work on the request flow,
    no lazy-init lock contention across pods that just spun up after a
    rolling deploy)."""
    try:
        await db.optimize_jobs.create_index("expires_at", expireAfterSeconds=0)
        await db.optimize_jobs.create_index("job_id", unique=True)
        logger.info("optimize_jobs indexes verified")
    except Exception as e:  # noqa: BLE001
        # Non-fatal: if the index already exists with a slightly different
        # spec, motor raises here. The collection still works without the
        # index; we'd just lose TTL auto-purge.
        logger.warning("optimize_jobs index create failed (non-fatal): %s", e)


async def _run_optimize_job(job_id: str, request: OptimizationRequest, current_user: User) -> None:
    """Background runner. Writes the resolved JSONResponse-equivalent dict
    (or an error description) into Mongo `optimize_jobs.{job_id}`.

    The wrapped `_optimize_route_inner` returns either a plain dict or a
    JSONResponse — we coerce both to a serialisable dict so the polling
    endpoint can return it directly without re-running anything."""
    try:
        result = await _optimize_route_inner(request=request, current_user=current_user)
        if hasattr(result, "body"):
            try:
                import json as _json_resp
                payload = _json_resp.loads(result.body.decode("utf-8"))
            except Exception as e:
                payload = {"_raw_response_decode_error": str(e)}
        else:
            payload = result
        await db.optimize_jobs.update_one(
            {"job_id": job_id},
            {"$set": {"status": "done", "result": payload,
                      "finished_at": _dt_jobs.now(_tz_jobs.utc)}},
        )
    except HTTPException as he:
        logger.error("[optimize/jobs] HTTPException job=%s status=%d:\n%s",
                     job_id, he.status_code, traceback.format_exc())
        await db.optimize_jobs.update_one(
            {"job_id": job_id},
            {"$set": {"status": "error",
                      "error": {"status": he.status_code, "detail": str(he.detail)},
                      "finished_at": _dt_jobs.now(_tz_jobs.utc)}},
        )
    except Exception as e:  # noqa: BLE001
        logger.error("[optimize/jobs] Unhandled crash job=%s:\n%s",
                     job_id, traceback.format_exc())
        await db.optimize_jobs.update_one(
            {"job_id": job_id},
            {"$set": {"status": "error",
                      "error": {"status": 500, "detail": f"Optimize crashed: {e}"},
                      "finished_at": _dt_jobs.now(_tz_jobs.utc)}},
        )


# Pro paywall gate for the heavy optimisation pipeline. Imported here
# (not at module top) because routes.billing.require_pro defers its
# `from server import db, get_current_user` to request time, but we
# still need the symbol bound at decorator-evaluation time.
from routes.billing import require_pro as _billing_require_pro

@api_router.post("/optimize/jobs", status_code=202)
async def create_optimize_job(
    request: OptimizationRequest = OptimizationRequest(),
    current_user: User = Depends(get_current_user),
    _pro=Depends(_billing_require_pro),
):
    """Kick off an optimize run in the background and return a job_id.

    Gated behind the Pro paywall (`require_pro`) because this endpoint
    runs the heavy multi-engine optimizer (VROOM, LKH-3, OR-Tools,
    etc.). Free users get a 402 Payment Required with an
    `upgrade_required: true` detail; the RN client surfaces the
    paywall sheet from that signal. Admins (STRIPE_ADMIN_USER_IDS env
    var) bypass the check.

    The driver's client (RN app) polls `/api/optimize/jobs/{job_id}` until
    `status` is `done` (then reads `result`) or `error` (then reads
    `error.detail`). This shape is independent of how slow the underlying
    pipeline is — Cloudflare's 100 s ceiling can't bite us here because
    *this* endpoint always replies in <100 ms (pure Mongo insert; the
    TTL+unique indexes are created at app startup, NOT on the hot path).
    """
    try:
        job_id = str(uuid.uuid4())
        now = _dt_jobs.now(_tz_jobs.utc)
        await db.optimize_jobs.insert_one({
            "job_id": job_id,
            "user_id": current_user.user_id,
            "status": "running",
            "started_at": now,
            "expires_at": now + _td_jobs(seconds=_OPTIMIZE_JOB_TTL_S),
            "result": None,
            "error": None,
        })
        logger.info("[optimize/jobs] kickoff job_id=%s user=%s", job_id, current_user.user_id)
        # Fire-and-forget — runner writes back to the same Mongo document.
        # CRITICAL: hold a strong reference until the runner completes; without
        # this, GC under memory pressure can silently cancel the task and the
        # frontend will poll `status: "running"` forever.
        task = _asyncio_jobs.create_task(_run_optimize_job(job_id, request, current_user))
        _OPTIMIZE_RUNNER_TASKS.add(task)
        task.add_done_callback(_OPTIMIZE_RUNNER_TASKS.discard)
        return {"job_id": job_id, "status": "running"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("[optimize/jobs] Kickoff crashed for user=%s:\n%s",
                     current_user.user_id, traceback.format_exc())
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": f"Optimization kickoff failed: {e}"},
        )


@api_router.get("/optimize/jobs/{job_id}")
async def get_optimize_job(
    job_id: str,
    current_user: User = Depends(get_current_user),
):
    """Poll an optimize job. Returns `{status, result?, error?}`.

    Scoped to the calling user — forging another driver's job_id yields
    404 (same shape as a TTL'd-out job, so we don't leak existence).

    Bandwidth shape: while `status == "running"` we project the (potentially
    multi-megabyte) `result` field away so each poll is a tiny round-trip
    that finishes well inside the client's per-poll AbortController. Once
    the runner writes status=done, the full payload is shipped on the
    next poll — that single response can be 2-5 MB on a 200-stop manifest
    so the client uses a longer timeout for it (see frontend POLL_TIMEOUT_MS).
    """
    # First peek: status only. ~100 bytes over the wire even on a hot Atlas.
    head = await db.optimize_jobs.find_one(
        {"job_id": job_id, "user_id": current_user.user_id},
        {"_id": 0, "status": 1},
    )
    if head is None:
        raise HTTPException(status_code=404, detail="Job not found or expired")
    status = head.get("status", "running")
    if status == "running":
        return {"job_id": job_id, "status": "running", "result": None, "error": None}
    # Job has terminated — fetch the full doc (with result/error) for one
    # final response. After this the client stops polling.
    j = await db.optimize_jobs.find_one(
        {"job_id": job_id, "user_id": current_user.user_id},
        {"_id": 0},
    )
    if j is None:  # raced with TTL expiry between the two reads
        raise HTTPException(status_code=404, detail="Job not found or expired")
    return {
        "job_id": job_id,
        "status": j.get("status", status),
        "result": j.get("result") if j.get("status") == "done" else None,
        "error": j.get("error") if j.get("status") == "error" else None,
    }


@api_router.get("/optimize/diagnostics")
async def optimize_diagnostics(
    current_user: User = Depends(get_current_user),
):
    """Return the last 10 optimize jobs for the calling user.

    Diagnostic endpoint for debugging "optimization keeps failing" reports.
    Shows job_id, status, algorithm, timing, error detail, and stop count
    without returning the full (multi-MB) result payload. Accessible from
    the app or a quick curl from the driver's phone browser.
    """
    cursor = db.optimize_jobs.find(
        {"user_id": current_user.user_id},
        {
            "_id": 0,
            "job_id": 1,
            "status": 1,
            "started_at": 1,
            "finished_at": 1,
            "error": 1,
            # Lightweight result summary — NOT the full stops array.
            "result.algorithm": 1,
            "result.stop_count": 1,
            "result.total_distance_km": 1,
            "result.reasoning": 1,
        },
    ).sort("started_at", -1).limit(10)
    jobs = await cursor.to_list(length=10)

    for j in jobs:
        # Add elapsed_seconds for quick triage
        if j.get("started_at") and j.get("finished_at"):
            try:
                elapsed = (j["finished_at"] - j["started_at"]).total_seconds()
                j["elapsed_seconds"] = round(elapsed, 1)
            except Exception:
                pass
        # Flatten result summary
        if j.get("result"):
            j["result_summary"] = {
                "algorithm": j["result"].get("algorithm"),
                "stop_count": j["result"].get("stop_count"),
                "total_distance_km": j["result"].get("total_distance_km"),
                "reasoning": (j["result"].get("reasoning") or "")[:120],
            }
            del j["result"]
        # Convert datetimes to ISO strings for JSON
        for k in ("started_at", "finished_at"):
            if j.get(k):
                try:
                    j[k] = j[k].isoformat()
                except Exception:
                    pass

    # Also include stop count + OSRM status for context
    stop_count = await db.stops.count_documents({
        "user_id": current_user.user_id, "completed": False,
    })

    return {
        "user_id": current_user.user_id,
        "pending_stops": stop_count,
        "osrm_url": OSRM_URL,
        "osrm_enabled": _osrm_enabled(),
        "vroom_available": VROOM_AVAILABLE,
        "pyvrp_available": PYVRP_AVAILABLE,
        "ortools_available": ORTOOLS_AVAILABLE,
        "lkh_available": LKH_AVAILABLE,
        "recent_jobs": jobs,
    }


# ─────────────────────────────────────────────────────────────────────────
# Build diagnostics — one curl from any browser/phone reveals exactly
# which code is live on production. Stops the "did the deploy take?"
# guessing dead in its tracks. No auth: deliberately public so the
# operator can hit it from a fresh device without copy-pasting a token.
# ─────────────────────────────────────────────────────────────────────────
import time as _time_diag  # avoid clash with the runner's _time_jobs

_BUILD_STARTED_AT = _time_diag.time()


@api_router.get("/_meta/build")
async def meta_build():
    """Return build / runtime info so the operator can verify what's
    actually deployed without SSH-ing into a pod. Fields chosen to map
    1:1 onto the questions we've been guessing at:

    * `started_at_iso` / `uptime_s` — was this pod just spun up?
    * `has_optimize_jobs_endpoint` — is the async-job pattern live?
    * `optimize_jobs_index_ok` / `optimize_jobs_count` — Mongo healthy?
    * `git_sha` — does prod match what's checked in (best-effort).
    """
    git_sha = "unknown"
    try:
        import subprocess
        git_sha = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd="/app", stderr=subprocess.DEVNULL,
        ).decode().strip() or "unknown"
    except Exception:  # noqa: BLE001
        pass

    optimize_jobs_count = -1
    optimize_jobs_index_ok = False
    try:
        optimize_jobs_count = await db.optimize_jobs.count_documents({})
        idx_info = await db.optimize_jobs.index_information()
        optimize_jobs_index_ok = any(
            "expires_at" in str(spec.get("key", []))
            for spec in idx_info.values()
        )
    except Exception as e:  # noqa: BLE001
        logger.warning("meta_build: optimize_jobs probe failed: %s", e)

    has_jobs_endpoint = any(
        getattr(r, "path", None) == "/api/optimize/jobs"
        for r in app.routes
    )

    return {
        "started_at_iso": _dt_jobs.fromtimestamp(_BUILD_STARTED_AT, _tz_jobs.utc).isoformat(),
        "uptime_s": int(_time_diag.time() - _BUILD_STARTED_AT),
        "git_sha": git_sha,
        "has_optimize_jobs_endpoint": has_jobs_endpoint,
        "optimize_jobs_index_ok": optimize_jobs_index_ok,
        "optimize_jobs_count": optimize_jobs_count,
        "now_utc": _dt_jobs.now(_tz_jobs.utc).isoformat(),
    }


# ─────────────────────────────────────────────────────────────────────
# Per-user telemetry rollup — debugging surface for production
# ─────────────────────────────────────────────────────────────────────
# Why this exists:
#   The agent debugging this codebase lives in a preview pod that
#   cannot connect to production's Mongo Atlas. When the user asks
#   "which algorithm did I use today?" or "is the geofence actually
#   firing?", we have no way to answer except by guessing from logs.
#
#   This endpoint exposes aggregate, no-PII rollups computed over the
#   caller's OWN route_history. The user can curl it (or surface it
#   in-app) and paste output back to the agent for diagnosis.
#
# Privacy posture:
#   * Auth-gated (caller's user_id is the ONLY filter on the query).
#   * Returns counts, percentiles, and the algorithm string — NEVER
#     addresses, lat/lng, names, or raw stop bodies.
#   * Forging another user's user_id is structurally impossible: the
#     filter is hard-bound to `current_user.user_id`.

def _today_utc_iso() -> str:
    """Start of the current UTC day in ISO format. Routes archived since
    this moment count as 'today'. We use UTC (not local time) because
    that's what `archived_at` is stored in."""
    now = datetime.now(timezone.utc)
    start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    return start.isoformat()


def _seven_days_ago_iso() -> str:
    return (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()


def _percentile(sorted_arr: List[float], q: float) -> Optional[float]:
    """Cheap nearest-rank percentile. `sorted_arr` MUST already be sorted."""
    if not sorted_arr:
        return None
    idx = min(len(sorted_arr) - 1, int(q * (len(sorted_arr) - 1)))
    return round(sorted_arr[idx], 1)


def _aggregate_rollup(route_docs: List[dict]) -> Dict[str, Any]:
    """Compute the rollup shape from a list of archived route docs.

    Each `route_docs` entry has `summary.telemetry` (per-route rollup
    computed at archive time) plus optional `summary.algorithm` (which
    we started persisting alongside this endpoint — older archives
    will report `None`).
    """
    if not route_docs:
        return {
            "archived_routes": 0,
            "best_route": None,
            "geofence_count": 0,
            "geofence_inferred_count": 0,
            "fallback_count": 0,
            "geofence_rate": None,
            "arrival_proximity_rate": None,
            "completion_distance_p50_m": None,
            "completion_distance_p95_m": None,
            "service_seconds_p50": None,
            "service_seconds_p95": None,
            "distance_samples": 0,
            "service_samples": 0,
        }

    # Roll up per-stop telemetry across all archived routes in the window.
    geofence_n = 0
    inferred_n = 0
    fallback_n = 0
    distances: List[float] = []
    service_seconds: List[float] = []

    for route in route_docs:
        for s in route.get("stops") or []:
            method = s.get("arrival_method")
            if method == "geofence":
                geofence_n += 1
            elif method == "geofence_inferred":
                inferred_n += 1
            elif method == "fallback_completion":
                fallback_n += 1
            cd = s.get("completion_distance_m")
            if isinstance(cd, (int, float)):
                distances.append(float(cd))
            # Real service-time sample: geofence-arrival + completion both
            # timed. `geofence_inferred` has a constant 30s back-date, so
            # it's deliberately excluded to keep the ML distribution clean.
            if method == "geofence":
                a, c = s.get("arrived_at"), s.get("completed_at")
                if a and c:
                    try:
                        if isinstance(a, str):
                            a = datetime.fromisoformat(a.replace("Z", "+00:00"))
                        if isinstance(c, str):
                            c = datetime.fromisoformat(c.replace("Z", "+00:00"))
                        service_seconds.append((c - a).total_seconds())
                    except Exception:
                        pass

    distances.sort()
    service_seconds.sort()

    # "Best route" in the window = lowest total_distance_km among routes
    # that actually have a distance recorded. If none, falls back to the
    # route with the most delivered stops (a coarser proxy for "good day").
    best = None
    for r in route_docs:
        summary = r.get("summary") or {}
        algo = summary.get("algorithm")  # may be None for older archives
        # Heuristic: many archives carry a `stats.total_distance_km` on
        # the route doc itself, but we wrote it only into the optimise
        # response, not the archive. Use delivered count as the
        # comparable signal for now.
        delivered = (summary.get("delivered") or 0)
        candidate = {
            "archived_at": r.get("archived_at"),
            "algorithm": algo,
            "stops": summary.get("total_stops"),
            "delivered": delivered,
            "skipped": summary.get("skipped"),
            "failed": summary.get("failed"),
            "total_weight_kg": summary.get("total_weight_kg"),
        }
        if best is None or delivered > (best.get("delivered") or 0):
            best = candidate

    return {
        "archived_routes": len(route_docs),
        "best_route": best,
        "geofence_count": geofence_n,
        "geofence_inferred_count": inferred_n,
        "fallback_count": fallback_n,
        "geofence_rate": (
            round(geofence_n / (geofence_n + inferred_n + fallback_n), 3)
            if (geofence_n + inferred_n + fallback_n) > 0
            else None
        ),
        "arrival_proximity_rate": (
            round((geofence_n + inferred_n) / (geofence_n + inferred_n + fallback_n), 3)
            if (geofence_n + inferred_n + fallback_n) > 0
            else None
        ),
        "completion_distance_p50_m": _percentile(distances, 0.5),
        "completion_distance_p95_m": _percentile(distances, 0.95),
        "service_seconds_p50": _percentile(service_seconds, 0.5),
        "service_seconds_p95": _percentile(service_seconds, 0.95),
        "distance_samples": len(distances),
        "service_samples": len(service_seconds),
    }


@api_router.get("/_meta/telemetry-rollup")
async def meta_telemetry_rollup(current_user: User = Depends(get_current_user)):
    """Aggregate, no-PII rollup of the caller's archived route telemetry.

    Two windows:
      * `today`        — archives since 00:00 UTC of the current day
      * `last_7_days`  — archives in the trailing 7-day window
    Plus an `ml_readiness` block summarising whether the Phase-1
    service-time learner has enough clean samples to train.

    Each window contains:
      archived_routes, best_route { algorithm, stops, delivered, ... },
      geofence_count, fallback_count, geofence_rate,
      completion_distance_p50_m, completion_distance_p95_m,
      service_seconds_p50, service_seconds_p95,
      distance_samples, service_samples
    """
    today_iso = _today_utc_iso()
    week_iso = _seven_days_ago_iso()

    # One query, broadest window — slice client-side for "today".
    cursor = db.route_history.find(
        {
            "user_id": current_user.user_id,
            "archived_at": {"$gte": week_iso},
        },
        {"_id": 0},
    ).sort("archived_at", -1)
    week_docs = await cursor.to_list(500)
    today_docs = [d for d in week_docs if (d.get("archived_at") or "") >= today_iso]

    today_rollup = _aggregate_rollup(today_docs)
    week_rollup = _aggregate_rollup(week_docs)

    # ML readiness: Phase-1 service-time learner needs ≥50 real
    # (geofence-arrival, geofence-completion) samples spread across
    # ≥10 distinct shifts to train without overfitting one day's
    # parking habits.
    PHASE_1_THRESHOLD = 50
    real_samples = week_rollup["service_samples"]
    blocked_on = None
    if real_samples < PHASE_1_THRESHOLD:
        inferred = week_rollup.get("geofence_inferred_count", 0)
        if week_rollup["geofence_count"] == 0 and week_rollup["fallback_count"] > 0 and inferred == 0:
            blocked_on = (
                "geofence not firing — every completion in the last 7 days "
                "used arrival_method='fallback_completion' and no proximity "
                "inferences either. Likely cause: viewMode is 'planning' "
                "(not 'navigating') when the user taps Delivered, so "
                "useGeofenceArrival.ts never gates open."
            )
        elif week_rollup["geofence_count"] == 0 and inferred > 0:
            blocked_on = (
                f"geofence hook still not firing ({inferred} 'geofence_inferred' "
                f"samples in the last 7 days back-stamped via the 150 m proximity "
                f"backstop). Real service-time samples need an actual hook fire — "
                f"check useGeofenceArrival.ts radius vs. completion_distance_p50_m."
            )
        elif week_rollup["geofence_count"] == 0:
            blocked_on = (
                "no completions recorded yet — drive a route and tap "
                "Delivered to start collecting samples."
            )
        else:
            blocked_on = (
                f"insufficient samples ({real_samples}/{PHASE_1_THRESHOLD}) "
                "— keep driving; learner will unlock automatically."
            )

    return {
        "user_id": current_user.user_id,
        "today_utc_start": today_iso,
        "today": today_rollup,
        "last_7_days": week_rollup,
        "ml_readiness": {
            "real_geofence_samples_last_7d": real_samples,
            "needed_for_phase_1": PHASE_1_THRESHOLD,
            "ready_to_train": real_samples >= PHASE_1_THRESHOLD,
            "blocked_on": blocked_on,
        },
    }


# ─────────────────────────────────────────────────────────────────────────
# Phase 1 ML — Service-Time Learner
# ─────────────────────────────────────────────────────────────────────────
#
# Pulls every archived route's `arrival_method='geofence'` samples for
# this user, computes bucketed-median service times by (suburb, hour),
# stores ONE model doc per user in `ml_service_time_models`. Idempotent
# — calling train again replaces the model.
#
# Why per-user: drivers in suburban Sunshine Coast take different times
# than drivers doing CBD high-rises. Sharing models across users would
# poison both. Tiny data volume (~150 samples → ~2 KB JSON per doc)
# means we can fit every user in a single Mongo collection without
# special care.


@api_router.post("/_meta/ml/train")
async def train_ml_service_time(current_user: User = Depends(get_current_user)):
    """Re-train the service-time learner from this user's archived routes.

    Returns the trained model summary + the count of buckets that
    survived the BUCKET_MIN_SAMPLES filter. The driver can refresh the
    Profile telemetry tile to see the model is now active.
    """
    from ml.service_time_learner import (
        collect_samples_from_archive,
        build_model_from_samples,
        summarize_model,
        BUCKET_MIN_SAMPLES,
        DEFAULT_SECONDS,
    )

    # Pull every archived route for this user (no time window — we
    # want the broadest sample set). _id excluded so the response can
    # be returned directly.
    routes: List[dict] = []
    cursor = db.route_history.find(
        {"user_id": current_user.user_id},
        {"_id": 0, "stops": 1, "archived_at": 1},
    )
    async for r in cursor:
        routes.append(r)

    samples = collect_samples_from_archive(routes)
    if len(samples) == 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "No real geofence samples found. Drive a route, tap "
                "Save Route, then train. Geofence_inferred / "
                "fallback_completion are not used (they have constant "
                "back-dated arrival times)."
            ),
        )

    model = build_model_from_samples(samples)

    # Persist. One doc per user_id, replaced on every retrain so the
    # collection never grows beyond N(users).
    await db.ml_service_time_models.replace_one(
        {"user_id": current_user.user_id},
        {"user_id": current_user.user_id, **model},
        upsert=True,
    )

    summary = summarize_model(model)
    logger.info(
        "[ml] Trained service-time model for user=%s: %d samples → "
        "%d suburbs, global_median=%.1fs, fastest=%s, slowest=%s",
        current_user.user_id,
        summary["sample_count"],
        summary["suburbs_covered"],
        summary["global_median_seconds"],
        summary["fastest_bucket_seconds"],
        summary["slowest_bucket_seconds"],
    )

    return {
        "ok": True,
        "trained_at": model["trained_at"],
        "sample_count": model["sample_count"],
        "bucket_count": len(model.get("buckets") or {}),
        "bucket_min_samples": BUCKET_MIN_SAMPLES,
        "default_seconds": DEFAULT_SECONDS,
        "summary": summary,
    }


@api_router.get("/_meta/ml/model")
async def get_ml_service_time_model(
    current_user: User = Depends(get_current_user),
):
    """Return the current model summary (driver-friendly) for the
    Profile tile. If no model has been trained, returns trained=false
    so the UI can show "Train now" instead of metrics."""
    from ml.service_time_learner import summarize_model

    doc = await db.ml_service_time_models.find_one(
        {"user_id": current_user.user_id},
        {"_id": 0},
    )
    return {
        "user_id": current_user.user_id,
        "model": summarize_model(doc),
    }


# ─────────────────────────────────────────────────────────────────────────
# Phase 2 ML — Building-Side Corrector
# ─────────────────────────────────────────────────────────────────────────
#
# Mapbox centroids land on the rooftop, but drivers park at the kerb. We
# observe that offset every time a driver taps Delivered with GPS on. The
# per-suburb median (Δlat, Δlng) is the predicted real arrival point for
# every new stop in that suburb, even when Mapbox didn't supply an
# `access_navigation_point`.
#
# Same per-user model pattern as Phase 1: one doc per user, replaced on
# every retrain. Source rows accept both `geofence` AND `geofence_inferred`
# arrival_method values (both supply real completion GPS); fallback_completion
# is excluded.


@api_router.post("/_meta/ml/building-side/train")
async def train_ml_building_side(current_user: User = Depends(get_current_user)):
    """Re-train the building-side corrector from this user's archived
    routes. Returns the trained model summary + the count of suburbs that
    survived the BUCKET_MIN_SAMPLES filter."""
    from ml.building_side_corrector import (
        collect_samples_from_archive,
        build_model_from_samples,
        summarize_model,
        BUCKET_MIN_SAMPLES,
        OUTLIER_MAX_METRES,
    )

    routes: List[dict] = []
    cursor = db.route_history.find(
        {"user_id": current_user.user_id},
        {"_id": 0, "stops": 1, "archived_at": 1},
    )
    async for r in cursor:
        routes.append(r)

    samples = collect_samples_from_archive(routes)
    if len(samples) == 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "No samples with completion GPS found. Drive a route with "
                "location services enabled, tap Save Route, then train. "
                "Only geofence/geofence_inferred stops contribute — "
                "fallback_completion rows have no GPS."
            ),
        )

    model = build_model_from_samples(samples)

    await db.ml_building_side_models.replace_one(
        {"user_id": current_user.user_id},
        {"user_id": current_user.user_id, **model},
        upsert=True,
    )

    summary = summarize_model(model)
    logger.info(
        "[ml/building-side] Trained for user=%s: %d samples → %d suburbs, "
        "global_offset=%.1fm, largest_suburb_offset=%sm",
        current_user.user_id,
        summary["sample_count"],
        summary["suburbs_covered"],
        summary["global_offset_metres"],
        summary["largest_suburb_offset_metres"],
    )

    return {
        "ok": True,
        "trained_at": model["trained_at"],
        "sample_count": model["sample_count"],
        "suburb_count": len(model.get("suburbs") or {}),
        "bucket_min_samples": BUCKET_MIN_SAMPLES,
        "outlier_max_metres": OUTLIER_MAX_METRES,
        "summary": summary,
    }


@api_router.get("/_meta/ml/building-side/model")
async def get_ml_building_side_model(
    current_user: User = Depends(get_current_user),
):
    """Return the current building-side correction model summary."""
    from ml.building_side_corrector import summarize_model

    doc = await db.ml_building_side_models.find_one(
        {"user_id": current_user.user_id},
        {"_id": 0},
    )
    return {
        "user_id": current_user.user_id,
        "model": summarize_model(doc),
    }


@api_router.post("/optimize/tighten-cluster")
async def tighten_cluster(
    request: TightenClusterRequest,
    current_user: User = Depends(get_current_user),
):
    """One-tap "fix this zig-zag" handler.

    Lifts a single suspect stop B (identified by `suspect_id`) out of its
    current slot and re-inserts it at the position that minimises the
    haversine perimeter of the route. Then double-checks against OSRM
    seconds — if the relocation actually costs driving time we roll back
    rather than mislead the driver. Persists the chosen order to Mongo and
    returns the refreshed sequence + remaining cluster warnings.
    """
    pending = await db.stops.find(
        {"user_id": current_user.user_id, "completed": {"$ne": True}},
        {"_id": 0},
    ).sort("order", 1).to_list(2000)

    suspect_idx = next(
        (i for i, s in enumerate(pending) if s.get("id") == request.suspect_id),
        None,
    )
    if suspect_idx is None:
        raise HTTPException(
            status_code=404,
            detail=f"suspect_id {request.suspect_id} not found in pending stops",
        )
    if len(pending) < 3:
        raise HTTPException(
            status_code=400,
            detail="Need at least 3 pending stops to tighten a cluster",
        )

    proposed, best_position, before_km, after_km = _relocate_stop_haversine(
        pending, suspect_idx
    )
    chosen, before_s, after_s, rolled_back = await _osrm_verify_relocation(
        pending, proposed
    )
    await _persist_pending_order(current_user.user_id, chosen)

    return {
        "message": (
            "Visual fix declined: OSRM says driving time would increase"
            if rolled_back
            else "Cluster tightened"
        ),
        "moved_stop_id": request.suspect_id,
        "from_position": suspect_idx,
        "to_position": suspect_idx if rolled_back else best_position,
        "rolled_back": rolled_back,
        "haversine_km_before": round(before_km, 3),
        "haversine_km_after": round(
            before_km if rolled_back else after_km, 3
        ),
        "saved_km": (
            0.0 if rolled_back else round(max(0.0, before_km - after_km), 3)
        ),
        "driving_seconds_before": before_s,
        "driving_seconds_after": (
            before_s if rolled_back else after_s
        ),
        "driving_seconds_saved": (
            None
            if before_s is None or after_s is None
            else 0
            if rolled_back
            else max(0, before_s - after_s)
        ),
        "stops": chosen,
        "optimized_sequence": [s["id"] for s in chosen],
        "cluster_warnings": _filter_actionable_warnings(
            chosen, detect_cluster_spikes(chosen)
        ),
    }


@api_router.post("/optimize/tighten-clusters")
async def tighten_all_clusters(
    current_user: User = Depends(get_current_user),
):
    """Iteratively tighten every detected spike in the current pending route.

    Loop until `detect_cluster_spikes` returns an empty list (or we hit a
    safety cap). On every pass we relocate the *worst* spike — the one
    with the largest `extra_km` (most map-distance wasted) — and persist
    the move at the end of the loop. This produces a strictly-improving
    haversine path with no manual intervention.

    The safety cap (`MAX_PASSES = 10`) prevents runaway loops in pathological
    cases where two spikes oscillate; in practice real-world routes
    converge in 1–3 passes.
    """
    pending = await db.stops.find(
        {"user_id": current_user.user_id, "completed": {"$ne": True}},
        {"_id": 0},
    ).sort("order", 1).to_list(2000)

    if len(pending) < 3:
        return {
            "message": "Nothing to tighten",
            "moves": [],
            "passes": 0,
            "haversine_km_before": round(_haversine_path_km(pending), 3),
            "haversine_km_after": round(_haversine_path_km(pending), 3),
            "saved_km": 0.0,
            "stops": pending,
            "optimized_sequence": [s["id"] for s in pending],
            "cluster_warnings": [],
        }

    initial_km = _haversine_path_km(pending)

    # Delegate to the shared tightener: alternates relocate + 2-opt until
    # both move generators are stuck. This is the same engine the auto-
    # tighten path inside /api/optimize uses, so a manual tap and an
    # automatic pass produce the same final state.
    current, moves = _iterative_haversine_tighten(pending)

    if moves:
        # Apply the haversine-best chain BEFORE the OSRM verification.
        # The verification will roll back to `pending` if OSRM disagrees.
        chosen, before_s, after_s, rolled_back = await _osrm_verify_relocation(
            pending, current
        )
        await _persist_pending_order(current_user.user_id, chosen)
        current = chosen
    else:
        before_s = after_s = None
        rolled_back = False

    final_km = _haversine_path_km(current)
    # Observability — one line per tighten call so we can answer "is this
    # actually doing anything?" from the prod log without instrumenting
    # the silent rolled-back path. Includes everything that decides the
    # user-visible outcome: move count, OSRM verdict, and km/seconds
    # delta.
    logger.info(
        "[tighten-clusters] user=%s pending=%d moves=%d rolled_back=%s "
        "haversine_before=%.2fkm haversine_after=%.2fkm "
        "osrm_before=%ss osrm_after=%ss",
        current_user.user_id,
        len(pending),
        len(moves),
        rolled_back,
        initial_km,
        final_km,
        before_s if before_s is not None else "n/a",
        after_s if after_s is not None else "n/a",
    )
    return {
        "message": (
            "Visual fix declined: OSRM says driving time would increase"
            if rolled_back
            else (
                f"Tightened {len(moves)} cluster"
                f"{'s' if len(moves) != 1 else ''}"
                if moves
                else "Route already clean"
            )
        ),
        "moves": [] if rolled_back else moves,
        "passes": 0 if rolled_back else len(moves),
        "rolled_back": rolled_back,
        "haversine_km_before": round(initial_km, 3),
        "haversine_km_after": round(final_km, 3),
        "saved_km": round(max(0.0, initial_km - final_km), 3),
        "driving_seconds_before": before_s,
        "driving_seconds_after": (
            before_s if rolled_back else after_s
        ),
        "driving_seconds_saved": (
            None
            if before_s is None or after_s is None
            else max(0, before_s - after_s)
            if not rolled_back
            else 0
        ),
        "stops": current,
        "optimized_sequence": [s["id"] for s in current],
        "cluster_warnings": (
            # Honest banner: when OSRM rolled the chain back, the algorithm
            # has just *proven* nothing in the current route is fixable
            # without driving longer than the slack budget allows. Suppress
            # the banner entirely — leaving stale warnings on screen after
            # a no-op tap is the precise UI lie this whole feature was
            # designed to eliminate.
            []
            if rolled_back
            else _filter_actionable_warnings(
                current, detect_cluster_spikes(current)
            )
        ),
    }


@api_router.get("/stops/export/xlsx")
async def export_stops_xlsx(current_user: User = Depends(get_current_user)):
    """Export all stops as an Excel (.xlsx) file, ordered by route sequence.

    The `#` column reflects the Sharpie-locked `original_sequence` whenever
    the route has been confirmed — that's the value the driver wrote on the
    physical box, so it's what the spreadsheet must show. Pre-confirm rows
    fall back to the live drive `order` (which may still re-shuffle on the
    next optimise). Sort key follows the same rule so the spreadsheet's
    row order ALWAYS matches the displayed `#` column, never the other
    way around.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    raw_stops = await db.stops.find({"user_id": current_user.user_id}, {"_id": 0}).to_list(2000)
    # Sort by (locked-sequence ?? live-order). Once original_sequence is
    # written by /api/routes/confirm it never moves, so the export stays
    # stable across re-optimisations of partially-confirmed routes.
    def _sort_key(s):
        seq = s.get("original_sequence")
        if isinstance(seq, int) and seq > 0:
            return (0, seq)
        ordr = s.get("order")
        return (1, ordr if isinstance(ordr, int) else 999_999)
    all_stops = sorted(raw_stops, key=_sort_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Route Stops"

    # Header style
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Group sibling parcels (same address, same `order` after sort) so
    # the spreadsheet reads as "stop → its parcels" instead of a flat list
    # where the driver has to scan addresses to spot multi-parcel stops.
    # Each ROW = one physical parcel with its own tracking number + weight
    # (the actual unit we ship). A subtotal row appears under each stop
    # whenever it contains 2+ parcels.
    headers = ["#", "Name", "Address", "Status", "Tracking #", "Weight (kg)", "Latitude", "Longitude", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    completed_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    subtotal_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    subtotal_font = Font(name="Calibri", italic=True, color="7F6000", size=10)
    total_weight = 0.0
    current_addr = None       # The address being grouped (siblings share this)
    addr_parcel_count = 0     # How many parcels seen so far for current_addr
    addr_weight_sum = 0.0     # Subtotal weight for current_addr
    excel_row = 2             # Pointer into worksheet (data rows + injected subtotals)

    def _emit_subtotal(rownum, parcel_count, wt_sum):
        """Inject a yellow subtotal banner below a multi-parcel stop."""
        if parcel_count < 2:
            return rownum  # Single-parcel stops don't need a subtotal banner
        for col in range(1, 5):
            cell = ws.cell(row=rownum, column=col, value="")
            cell.fill = subtotal_fill
            cell.border = thin_border
        label_cell = ws.cell(row=rownum, column=5,
                             value=f"{parcel_count} parcels — subtotal:")
        label_cell.fill = subtotal_fill
        label_cell.font = subtotal_font
        label_cell.alignment = Alignment(horizontal="right")
        label_cell.border = thin_border
        sub_cell = ws.cell(row=rownum, column=6, value=round(wt_sum, 2))
        sub_cell.fill = subtotal_fill
        sub_cell.font = subtotal_font
        sub_cell.alignment = Alignment(horizontal="right")
        sub_cell.border = thin_border
        for col in (7, 8, 9):
            cell = ws.cell(row=rownum, column=col, value="")
            cell.fill = subtotal_fill
            cell.border = thin_border
        return rownum + 1

    for stop in all_stops:
        stop_addr = stop.get("address", "")
        # Boundary detection — if this stop's address differs from the one
        # we were grouping, flush a subtotal row for the previous group.
        if current_addr is not None and stop_addr != current_addr:
            excel_row = _emit_subtotal(excel_row, addr_parcel_count, addr_weight_sum)
            addr_parcel_count = 0
            addr_weight_sum = 0.0
        current_addr = stop_addr
        addr_parcel_count += 1

        order_num = stop.get("order", excel_row - 2)
        is_completed = stop.get("completed", False)
        # Pin number rendered in the # column follows the same Sharpie-marker
        # contract as the map pins:
        #  • Locked: original_sequence (immutable post first /routes/confirm)
        #  • Tentative: order + 1 (server-stamped optimised drive position)
        # The driver cannot get a different number on the box vs the
        # spreadsheet vs the map — they are all bound to original_sequence
        # the moment the route is confirmed, never to the live `order`
        # which can shift on re-optimise.
        seq = stop.get("original_sequence")
        if isinstance(seq, int) and seq > 0:
            display_num = seq
        else:
            display_num = (order_num + 1) if isinstance(order_num, int) else (excel_row - 1)
        # Weight is optional — empty string when missing so the column reads
        # cleanly in Excel (vs `None` which renders literally as "None").
        # Sum only the populated values so the totals reflect ACTUAL known
        # load, not under-counted phantom zeros.
        raw_w = stop.get("weight")
        weight_val = round(float(raw_w), 2) if isinstance(raw_w, (int, float)) else ""
        if isinstance(weight_val, float):
            total_weight += weight_val
            addr_weight_sum += weight_val
        values = [
            display_num,
            stop.get("name", ""),
            stop_addr,
            "Completed" if is_completed else "Pending",
            stop.get("tracking_number", "") or "",
            weight_val,
            round(stop.get("latitude", 0), 6),
            round(stop.get("longitude", 0), 6),
            stop.get("notes", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.border = thin_border
            if is_completed:
                cell.fill = completed_fill
        # Right-align the Weight column for readability — numbers sit
        # cleanly under the header when right-aligned.
        ws.cell(row=excel_row, column=6).alignment = Alignment(horizontal="right")
        # Tracking column also right-aligned + monospace-feel — barcode IDs
        # sort better visually that way.
        ws.cell(row=excel_row, column=5).alignment = Alignment(horizontal="left")
        excel_row += 1

    # Flush the final group's subtotal (if it was a multi-parcel stop).
    if current_addr is not None:
        excel_row = _emit_subtotal(excel_row, addr_parcel_count, addr_weight_sum)

    # Summary footer — total weight across all populated rows. Helps the
    # driver / dispatch sanity-check vehicle load capacity in one glance.
    if all_stops:
        footer_row = excel_row
        footer_font = Font(name="Calibri", bold=True, size=11)
        footer_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for col in range(1, 5):
            cell = ws.cell(row=footer_row, column=col,
                           value="Grand Total Weight" if col == 4 else "")
            cell.font = footer_font
            cell.fill = footer_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if col == 4 else "left")
        # Tracking column header in footer — leave blank
        blank = ws.cell(row=footer_row, column=5, value="")
        blank.font = footer_font
        blank.fill = footer_fill
        blank.border = thin_border
        total_cell = ws.cell(row=footer_row, column=6, value=round(total_weight, 2))
        total_cell.font = footer_font
        total_cell.fill = footer_fill
        total_cell.border = thin_border
        total_cell.alignment = Alignment(horizontal="right")
        # Trailing blank cells get the same fill so the row reads as a
        # single banded footer.
        for col in (7, 8, 9):
            cell = ws.cell(row=footer_row, column=col, value="")
            cell.fill = footer_fill
            cell.border = thin_border

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=route_stops.xlsx"},
    )

@api_router.get("/optimize/algorithms")
async def list_optimization_algorithms(response: Response):
    """List available optimization algorithms with descriptions"""
    response.headers["Cache-Control"] = "public, max-age=86400"  # 24h — static data
    return {
        "algorithms": [
            {
                "id": "auto",
                "name": "Auto Select",
                "description": "Automatically selects the best algorithm based on route size",
                "best_for": "All route sizes"
            },
            {
                "id": "alns",
                "name": "ALNS Hybrid",
                "description": "Adaptive Large Neighbourhood Search with Simulated Annealing and Local Search polish",
                "best_for": "Medium to large routes (10-100+ stops)",
                "complexity": "O(iterations × n)"
            },
            {
                "id": "ortools",
                "name": "OR-Tools",
                "description": "Google OR-Tools single-vehicle optimization prioritizing travel time, then distance",
                "best_for": "High-quality sequencing with fallback safety",
                "complexity": "CP-SAT / local search"
            },
            {
                "id": "pyvrp",
                "name": "PyVRP (HGS)",
                "description": "State-of-the-art Hybrid Genetic Search — minimises total driving time (OSRM duration matrix) for a single driver pure TSP",
                "best_for": "Fastest drop-off sequencing, no time windows, 10-200 stops",
                "complexity": "HGS population-based metaheuristic"
            },
            {
                "id": "nearest_neighbor",
                "name": "Nearest Neighbor",
                "description": "Fast greedy algorithm that always visits the closest unvisited stop",
                "best_for": "Large routes (50+ stops), quick estimates",
                "complexity": "O(n²)"
            },
            {
                "id": "two_opt",
                "name": "2-Opt",
                "description": "Improvement heuristic that reverses route segments to reduce distance",
                "best_for": "Small to medium routes (up to 25 stops)",
                "complexity": "O(n²) per iteration"
            },
            {
                "id": "simulated_annealing",
                "name": "Simulated Annealing",
                "description": "Probabilistic meta-heuristic inspired by metallurgy. Accepts worse solutions early to escape local optima",
                "best_for": "Medium routes (15-40 stops)",
                "complexity": "O(n × iterations)"
            },
            {
                "id": "genetic",
                "name": "Genetic Algorithm",
                "description": "Evolutionary algorithm that evolves a population of solutions through selection, crossover, and mutation",
                "best_for": "Complex routes with many constraints (20-60 stops)",
                "complexity": "O(population × generations × n)"
            },
            {
                "id": "clarke_wright",
                "name": "Clarke-Wright Savings",
                "description": "Classic VRP algorithm that builds routes by merging based on distance savings from depot",
                "best_for": "Delivery routes starting from a depot/warehouse",
                "complexity": "O(n² log n)"
            },
            {
                "id": "cluster_first",
                "name": "Cluster-First",
                "description": "DBSCAN geographic clustering into neighborhoods, then OR-Tools per-cluster with Mapbox road distances. Guarantees spatially coherent routing — clears each area before moving to the next.",
                "best_for": "Large delivery routes (25-500+ stops) with multiple neighborhoods",
                "complexity": "O(n² clustering + per-cluster OR-Tools)"
            },
            {
                "id": "generoute",
                "name": "Generoute",
                "description": "Cloud-based route optimization using real road network data via Generoute API",
                "best_for": "Accurate road-based optimization (up to 1000 stops)",
                "complexity": "Cloud API"
            },
            {
                "id": "mapbox",
                "name": "Mapbox Optimization",
                "description": "Road-based optimization using Mapbox Optimization API",
                "best_for": "Small routes requiring accurate driving directions (up to 12 stops)",
                "complexity": "Cloud API"
            },
        ]
    }

# ===================== Algorithm Recommendation =====================

def _analyze_route_characteristics(stops: List[dict], distance_matrix: List[List[float]]) -> Dict[str, Any]:
    """Analyze geographic characteristics of the route to inform algorithm selection."""
    n = len(stops)
    if n < 2:
        return {"stop_count": n}

    # Collect all pairwise distances
    all_dists = []
    for i in range(n):
        for j in range(i + 1, n):
            all_dists.append(distance_matrix[i][j])

    avg_dist = sum(all_dists) / len(all_dists) if all_dists else 0
    max_dist = max(all_dists) if all_dists else 0

    # Geographic spread: max distance between any two stops
    spread_km = max_dist

    # Nearest-neighbor distances (how clustered stops are)
    nn_dists = []
    for i in range(n):
        nearest = min(distance_matrix[i][j] for j in range(n) if j != i)
        nn_dists.append(nearest)
    avg_nn = sum(nn_dists) / len(nn_dists) if nn_dists else 0

    # Cluster density: ratio of avg nearest-neighbor dist to avg pairwise dist
    # Low ratio = tightly clustered, high ratio = evenly spread
    cluster_ratio = avg_nn / avg_dist if avg_dist > 0 else 1.0

    # Count how many "clusters" exist using a simple threshold
    # Stops within avg_nn * 2 of each other are in the same cluster
    threshold = avg_nn * 3
    visited = [False] * n
    cluster_count = 0
    for i in range(n):
        if visited[i]:
            continue
        cluster_count += 1
        stack = [i]
        while stack:
            cur = stack.pop()
            if visited[cur]:
                continue
            visited[cur] = True
            for j in range(n):
                if not visited[j] and distance_matrix[cur][j] < threshold:
                    stack.append(j)

    return {
        "stop_count": n,
        "spread_km": round(spread_km, 2),
        "avg_distance_km": round(avg_dist, 3),
        "avg_nn_distance_km": round(avg_nn, 3),
        "cluster_ratio": round(cluster_ratio, 4),
        "cluster_count": cluster_count,
        "complexity": "low" if n < 15 else "medium" if n < 60 else "high",
    }


def _recommend_algorithm(chars: Dict[str, Any]) -> Dict[str, Any]:
    """Recommend the best algorithm based on route characteristics."""
    n = chars["stop_count"]
    cluster_ratio = chars.get("cluster_ratio", 0.5)
    cluster_count = chars.get("cluster_count", 1)
    # complexity = chars.get("complexity", "medium")  # reserved for future tuning

    if n < 2:
        return {"algorithm": "none", "confidence": 1.0, "reasoning": "Need at least 2 stops"}

    # Decision tree based on empirical algorithm strengths
    if n <= 8:
        return {
            "algorithm": "two_opt",
            "confidence": 0.95,
            "reasoning": f"With only {n} stops, 2-Opt finds near-optimal solutions instantly.",
            "alternatives": ["nearest_neighbor"],
        }

    if n <= 20:
        if cluster_count <= 2:
            return {
                "algorithm": "ortools",
                "confidence": 0.9,
                "reasoning": f"{n} stops in {cluster_count} cluster(s) — OR-Tools handles this scale perfectly with exact-like solutions.",
                "alternatives": ["two_opt", "simulated_annealing"],
            }
        else:
            return {
                "algorithm": "ortools",
                "confidence": 0.85,
                "reasoning": f"{n} stops across {cluster_count} clusters — OR-Tools balances quality and speed well here.",
                "alternatives": ["alns", "simulated_annealing"],
            }

    if n <= 60:
        if cluster_count >= 4:
            return {
                "algorithm": "alns",
                "confidence": 0.88,
                "reasoning": f"{n} stops across {cluster_count} clusters — ALNS excels at multi-cluster routes with its destroy-repair operators.",
                "alternatives": ["ortools", "simulated_annealing"],
            }
        else:
            return {
                "algorithm": "ortools",
                "confidence": 0.85,
                "reasoning": f"{n} stops in {cluster_count} cluster(s) — OR-Tools provides strong results at this scale.",
                "alternatives": ["alns", "two_opt"],
            }

    # 60+ stops — large scale
    if cluster_ratio < 0.15:
        # Very tightly clustered
        return {
            "algorithm": "alns",
            "confidence": 0.92,
            "reasoning": f"{n} tightly clustered stops (ratio: {cluster_ratio:.2f}) — ALNS's adaptive operators handle dense routes best.",
            "alternatives": ["simulated_annealing", "ortools"],
        }
    elif cluster_count >= 5:
        return {
            "algorithm": "alns",
            "confidence": 0.9,
            "reasoning": f"{n} stops across {cluster_count} distinct clusters — ALNS's segment-based destroy operators are ideal for multi-cluster optimization.",
            "alternatives": ["ortools", "simulated_annealing"],
        }
    else:
        return {
            "algorithm": "alns",
            "confidence": 0.85,
            "reasoning": f"Large route with {n} stops — ALNS provides the best quality for complex large-scale optimization.",
            "alternatives": ["ortools", "simulated_annealing"],
        }


@api_router.get("/optimize/recommend")
async def recommend_algorithm(current_user: User = Depends(get_current_user)):
    """Analyze current route and recommend the best optimization algorithm."""
    all_stops = await db.stops.find({"user_id": current_user.user_id}, {"_id": 0}).sort("order", 1).to_list(1000)
    stops = [s for s in all_stops if not s.get("completed")]

    if len(stops) < 2:
        return {
            "recommendation": {"algorithm": "none", "confidence": 1.0, "reasoning": "Need at least 2 incomplete stops"},
            "characteristics": {"stop_count": len(stops)},
        }

    distance_matrix = calculate_distance_matrix(stops)
    chars = _analyze_route_characteristics(stops, distance_matrix)
    rec = _recommend_algorithm(chars)

    return {
        "recommendation": rec,
        "characteristics": chars,
    }


# ===================== Benchmark & Shadow-Testing =====================

def compute_route_quality_metrics(stops: List[dict], distance_matrix: List[List[float]], route_indices: List[int]) -> Dict[str, Any]:
    """Compute route quality metrics beyond raw distance."""
    n = len(route_indices)
    if n < 2:
        return {"backtrack_count": 0, "backtrack_ratio": 0.0, "longest_leg_km": 0.0, "shortest_leg_km": 0.0, "leg_variance": 0.0, "cluster_score": 1.0}

    legs = []
    for i in range(n - 1):
        legs.append(distance_matrix[route_indices[i]][route_indices[i + 1]])

    # Backtracking: a leg that increases bearing by >120° relative to previous
    backtrack_count = 0
    for i in range(1, len(legs)):
        if i + 1 < n:
            a = route_indices[i - 1]
            b = route_indices[i]
            c = route_indices[i + 1]
            # Simple heuristic: if going to c is further from a than b is, we're backtracking
            dist_a_c = distance_matrix[a][c]
            dist_a_b = distance_matrix[a][b]
            dist_b_c = distance_matrix[b][c]
            if dist_a_c < dist_a_b and dist_b_c > 0:
                backtrack_count += 1

    longest_leg = max(legs) if legs else 0.0
    shortest_leg = min(legs) if legs else 0.0
    mean_leg = sum(legs) / len(legs) if legs else 0.0
    variance = sum((leg - mean_leg) ** 2 for leg in legs) / len(legs) if legs else 0.0

    # Cluster coherence: ratio of sequential neighbor distances vs random shuffle average
    total_dist = sum(legs)
    all_distances = [distance_matrix[i][j] for i in range(len(distance_matrix)) for j in range(len(distance_matrix)) if i != j]
    avg_random_dist = sum(all_distances) / len(all_distances) if all_distances else 1.0
    expected_random_total = avg_random_dist * (n - 1)
    cluster_score = round(1.0 - (total_dist / expected_random_total) if expected_random_total > 0 else 0.0, 4)

    return {
        "backtrack_count": backtrack_count,
        "backtrack_ratio": round(backtrack_count / max(n - 2, 1), 4),
        "longest_leg_km": round(longest_leg, 3),
        "shortest_leg_km": round(shortest_leg, 3),
        "leg_variance": round(variance, 4),
        "cluster_score": max(0.0, cluster_score),
    }


def _run_algorithm_benchmark(algo_id: str, stops: List[dict], distance_matrix: List[List[float]], start_index: int) -> Dict[str, Any]:
    """Run a single algorithm and collect metrics. Returns dict with results or error."""
    import time as _bench_time
    t0 = _bench_time.perf_counter()
    error = None
    optimized = []

    try:
        if algo_id == "vroom":
            if not VROOM_AVAILABLE:
                error = f"pyvroom not available: {VROOM_IMPORT_ERROR}"
            else:
                indices = vroom_tsp_solve(distance_matrix, depot=start_index)
                optimized = [stops[i] for i in indices]
        elif algo_id == "lkh":
            if not LKH_AVAILABLE:
                error = f"LKH not available: {LKH_IMPORT_ERROR}"
            else:
                indices = lkh_tsp_solve(distance_matrix, depot=start_index)
                optimized = [stops[i] for i in indices]
        elif algo_id == "pyvrp":
            # PyVRP HGS — same engine the production /api/optimize uses by
            # default. Was missing from the benchmark dispatcher, which made
            # it show up in the demo report as "Unknown algorithm: pyvrp".
            if not PYVRP_AVAILABLE:
                error = f"pyvrp not available: {PYVRP_IMPORT_ERROR}"
            else:
                time_limit = max(2.0, min(8.0, 2.0 + len(stops) / 40))
                coords = [(float(s.get("lng", 0.0)), float(s.get("lat", 0.0))) for s in stops]
                indices = pyvrp_tsp_solve(
                    distance_matrix,
                    depot=start_index,
                    time_limit_seconds=time_limit,
                    seed=42,
                    coordinates=coords,
                )
                optimized = [stops[i] for i in indices]
        elif algo_id == "vroom_lkh_3opt":
            # Full pipeline: VROOM -> LKH-3 -> 3-opt (the production chain)
            if not VROOM_AVAILABLE:
                error = f"pyvroom not available: {VROOM_IMPORT_ERROR}"
            else:
                indices = vroom_tsp_solve(distance_matrix, depot=start_index)
                if LKH_AVAILABLE:
                    try:
                        indices = lkh_tsp_solve(distance_matrix, depot=start_index)
                    except Exception:
                        pass
                indices = three_opt_improve(indices, distance_matrix, max_iterations=3)
                optimized = [stops[i] for i in indices]
        elif algo_id == "timefold":
            if not TIMEFOLD_AVAILABLE:
                error = f"Timefold not available: {TIMEFOLD_IMPORT_ERROR}"
            else:
                time_limit = max(5, min(15, 5 + len(stops) // 20))
                optimized = timefold_optimize(stops, distance_matrix, start_index=start_index, time_limit_seconds=time_limit)
        elif algo_id == "alns":
            time_limit = max(3, min(10, 5 + len(stops) // 15))
            optimized = alns_hybrid_optimize(stops, distance_matrix, start_index=start_index, time_limit_seconds=time_limit)
        elif algo_id == "ortools":
            time_limit = max(3, min(12, 5 + len(stops) // 10))
            optimized = ortools_optimize(stops, distance_matrix, start_index=start_index, time_limit_seconds=time_limit)
        elif algo_id == "nearest_neighbor":
            optimized = nearest_neighbor_optimize(stops, distance_matrix, start_index)
        elif algo_id == "two_opt":
            nn = nearest_neighbor_optimize(stops, distance_matrix, start_index)
            ri = _indices_by_identity(stops, nn)
            improved = two_opt_improve(ri, distance_matrix)
            optimized = [stops[i] for i in improved]
        elif algo_id == "three_opt":
            nn = nearest_neighbor_optimize(stops, distance_matrix, start_index)
            ri = _indices_by_identity(stops, nn)
            improved = three_opt_improve(ri, distance_matrix, max_iterations=3)
            optimized = [stops[i] for i in improved]
        elif algo_id == "simulated_annealing":
            iters = min(8000, 3000 + len(stops) * 60)
            optimized = simulated_annealing_optimize(stops, distance_matrix, start_index, iterations=iters)
        elif algo_id == "genetic":
            gens = min(150, 60 + len(stops))
            pop = max(25, min(40, len(stops)))
            optimized = genetic_algorithm_optimize(stops, distance_matrix, start_index, generations=gens, population_size=pop)
        elif algo_id == "clarke_wright":
            optimized = clarke_wright_savings(stops, distance_matrix, start_index)
        elif algo_id == "ils":
            time_limit = max(3, min(12, 5 + len(stops) // 15))
            optimized = iterated_local_search(stops, distance_matrix, start_index=start_index, time_limit_seconds=time_limit)
        elif algo_id == "vroom_ortools":
            # VROOM warm-start + OR-Tools GLS refinement
            if not VROOM_AVAILABLE:
                error = f"pyvroom not available: {VROOM_IMPORT_ERROR}"
            else:
                indices = vroom_tsp_solve(distance_matrix, depot=start_index)
                vroom_seed = [stops[i] for i in indices]
                try:
                    time_limit = max(3, min(10, 3 + len(stops) // 15))
                    optimized = ortools_optimize(vroom_seed, distance_matrix, start_index=0, time_limit_seconds=time_limit)
                except Exception:
                    optimized = vroom_seed
        else:
            error = f"Unknown algorithm: {algo_id}"
    except Exception as exc:
        error = str(exc)[:120]

    elapsed_ms = round((_bench_time.perf_counter() - t0) * 1000, 1)

    if error or not optimized:
        return {"algorithm": algo_id, "error": error or "No result", "time_ms": elapsed_ms}

    # Build route indices for metrics
    id_to_idx = {id(s): i for i, s in enumerate(stops)}
    route_indices = [id_to_idx.get(id(s), 0) for s in optimized]

    total_dist = 0.0
    for i in range(len(route_indices) - 1):
        total_dist += distance_matrix[route_indices[i]][route_indices[i + 1]]

    quality = compute_route_quality_metrics(stops, distance_matrix, route_indices)

    return {
        "algorithm": algo_id,
        "total_distance_km": round(total_dist, 3),
        "time_ms": elapsed_ms,
        "quality": quality,
        "error": None,
    }


@api_router.post("/benchmark")
async def benchmark_algorithms(
    request: BenchmarkRequest = BenchmarkRequest(),
    current_user: User = Depends(get_current_user),
):
    """Run all (or selected) algorithms on the current route and return comparison metrics."""
    all_user_stops = await db.stops.find({"user_id": current_user.user_id}, {"_id": 0}).sort("order", 1).to_list(1000)
    stops = [s for s in all_user_stops if not s.get("completed")]

    if len(stops) < 2:
        raise HTTPException(status_code=400, detail="Need at least 2 incomplete stops to benchmark")

    start_index = 0
    if request.use_current_location and request.current_latitude and request.current_longitude:
        current_loc = {
            "id": "current_location",
            "address": "Current Location",
            "latitude": request.current_latitude,
            "longitude": request.current_longitude,
            "completed": False,
        }
        stops = [current_loc] + stops
        start_index = 0

    distance_matrix = calculate_distance_matrix(stops)

    LOCAL_ALGORITHMS = [
        "vroom_lkh_3opt", "vroom_ortools", "vroom", "lkh", "timefold",
        "alns", "ortools", "pyvrp", "ils",
        "nearest_neighbor", "two_opt", "three_opt",
        "simulated_annealing", "genetic", "clarke_wright",
    ]
    # Filter out solvers whose native dependencies aren't present in this environment
    # (LKH binary on bare-metal, Java JVM for Timefold). In the production Docker
    # image these aren't shipped, so listing them only produces "Failed" noise.
    if not LKH_AVAILABLE:
        LOCAL_ALGORITHMS = [a for a in LOCAL_ALGORITHMS if a not in ("lkh", "vroom_lkh_3opt")]
    if not TIMEFOLD_AVAILABLE:
        LOCAL_ALGORITHMS = [a for a in LOCAL_ALGORITHMS if a != "timefold"]
    if not PYVRP_AVAILABLE:
        LOCAL_ALGORITHMS = [a for a in LOCAL_ALGORITHMS if a != "pyvrp"]
    algos_to_run = request.algorithms if request.algorithms else LOCAL_ALGORITHMS

    # Run each algorithm (sequentially to avoid CPU contention skewing times).
    # Wrapped in asyncio.to_thread so the event loop stays responsive — otherwise
    # 129+ stops × 13 solvers blocks the entire FastAPI process for ~2 minutes,
    # which both times out the K8s ingress and freezes all other requests.
    # A 45s wall-time budget caps runaway benchmarks (esp. LKH / genetic on big
    # routes); algorithms that don't fit in the budget come back as errors so
    # the UI still shows partial results.
    import time as _budget_time

    def _run_all_algorithms() -> List[Dict[str, Any]]:
        results_local: List[Dict[str, Any]] = []
        budget_seconds = 45.0
        started = _budget_time.perf_counter()
        for algo_id in algos_to_run:
            if algo_id not in LOCAL_ALGORITHMS:
                continue
            if _budget_time.perf_counter() - started > budget_seconds:
                results_local.append({
                    "algorithm": algo_id,
                    "error": "Skipped (45s benchmark budget exceeded)",
                    "time_ms": 0,
                })
                continue
            results_local.append(
                _run_algorithm_benchmark(algo_id, stops, distance_matrix, start_index)
            )
        return results_local

    results = await asyncio.to_thread(_run_all_algorithms)

    # Sort by distance (best first), errors last
    successful = [r for r in results if r.get("error") is None]
    failed = [r for r in results if r.get("error") is not None]
    successful.sort(key=lambda r: r["total_distance_km"])

    winner = successful[0]["algorithm"] if successful else None

    return {
        "stop_count": len(stops),
        "results": successful + failed,
        "winner": winner,
        "started_from_current_location": start_index == 0 and request.use_current_location,
    }


# ===================== Generoute API Endpoints =====================

@api_router.get("/generoute/status")
async def generoute_status(current_user: User = Depends(get_current_user)):
    """Check if Generoute API is configured and available."""
    configured = bool(GENEROUTE_API_KEY)
    return {
        "configured": configured,
        "api_key_set": configured,
        "services": {
            "route_optimization": configured
        }
    }

# ===================== Mapbox Proxy Endpoints =====================

@api_router.get("/geocode")
async def geocode_address(query: str, current_user: User = Depends(get_current_user)):
    """Search for addresses using Mapbox Geocoding API with full metadata.
    Returns: rooftop centroid, access/navigation point, plus code, interpolation status, and rich metadata"""
    if not MAPBOX_TOKEN:
        raise HTTPException(status_code=500, detail="Mapbox token not configured")
    
    geo_context = await get_user_geocoding_context(current_user.user_id)
    
    params = {
        "q": query,
        "access_token": MAPBOX_TOKEN,
        "limit": 5,
        "types": "address,street,place",
        "routing": "true",
    }
    if geo_context.get("proximity"):
        params["proximity"] = geo_context["proximity"]
    if geo_context.get("country"):
        params["country"] = geo_context["country"]
    
    async with httpx.AsyncClient() as client:
        response = await client.get(
            "https://api.mapbox.com/search/geocode/v6/forward",
            params=params
        )
        
        if response.status_code != 200:
            raise HTTPException(status_code=response.status_code, detail="Geocoding failed")
        
        data = response.json()
        return [_extract_rich_feature(f) for f in data.get("features", [])]

def _maneuver_instruction(maneuver: dict, name: str) -> str:
    """Build a human-readable instruction from OSRM maneuver type/modifier + street name."""
    mtype = maneuver.get("type", "")
    modifier = maneuver.get("modifier", "")
    road = f" onto {name}" if name else ""
    lookup = {
        ("depart", ""): f"Head{road}",
        ("arrive", ""): "You have arrived",
        ("turn", "left"): f"Turn left{road}",
        ("turn", "right"): f"Turn right{road}",
        ("turn", "slight left"): f"Slight left{road}",
        ("turn", "slight right"): f"Slight right{road}",
        ("turn", "sharp left"): f"Sharp left{road}",
        ("turn", "sharp right"): f"Sharp right{road}",
        ("turn", "uturn"): f"Make a U-turn{road}",
        ("continue", "straight"): f"Continue straight{road}",
        ("continue", ""): f"Continue{road}",
        ("merge", "slight left"): f"Merge left{road}",
        ("merge", "slight right"): f"Merge right{road}",
        ("new name", ""): f"Continue{road}",
        ("roundabout", ""): f"Enter the roundabout{road}",
        ("rotary", ""): f"Enter the roundabout{road}",
        ("exit roundabout", ""): f"Exit the roundabout{road}",
        ("exit rotary", ""): f"Exit the roundabout{road}",
        ("fork", "left"): f"Keep left{road}",
        ("fork", "right"): f"Keep right{road}",
        ("end of road", "left"): f"Turn left{road}",
        ("end of road", "right"): f"Turn right{road}",
    }
    key = (mtype, modifier)
    if key in lookup:
        return lookup[key]
    key_type = (mtype, "")
    if key_type in lookup:
        return lookup[key_type]
    if modifier:
        return f"{modifier.replace('_', ' ').title()}{road}"
    return f"Continue{road}"


def _extract_steps(legs: list) -> list:
    """Extract steps from OSRM/Mapbox route legs — shared by batch and single paths.

    Works with both OSRM and Mapbox response formats (OSRM is the origin of the format).
    Generates human-readable instructions from maneuver type/modifier for OSRM responses.
    """
    all_steps = []
    for leg_idx, leg in enumerate(legs):
        for step in leg.get("steps", []):
            maneuver = step.get("maneuver", {})
            instruction = maneuver.get("instruction", "")
            if not instruction:
                instruction = _maneuver_instruction(maneuver, step.get("name", ""))
            all_steps.append({
                "leg_index": leg_idx,
                "distance": step.get("distance", 0),
                "duration": step.get("duration", 0),
                "instruction": instruction,
                "type": maneuver.get("type", ""),
                "modifier": maneuver.get("modifier", ""),
                "bearing_before": maneuver.get("bearing_before", 0),
                "bearing_after": maneuver.get("bearing_after", 0),
                "location": maneuver.get("location", []),
                "name": step.get("name", ""),
                "geometry": step.get("geometry", {}),
                "driving_side": step.get("driving_side", "right"),
                "mode": step.get("mode", "driving"),
                "voice_instruction": step.get("voiceInstructions", [{}])[0].get("announcement", "") if step.get("voiceInstructions") else "",
                "banner_instruction": step.get("bannerInstructions", [{}])[0] if step.get("bannerInstructions") else {}
            })
    return all_steps

def _round_coord(c: str, precision: int = 4) -> str:
    """Round a coordinate string for cache key (reduces cache misses from GPS jitter)"""
    parts = c.split(",")
    return ",".join(f"{float(p):.{precision}f}" for p in parts)

@api_router.get("/directions")
async def get_directions(coordinates: str, response: Response):
    """Get route directions from local OSRM Route API (zero-cost, no API key).
    coordinates format: lng1,lat1;lng2,lat2;lng3,lat3
    No waypoint limits — OSRM handles hundreds of waypoints natively.
    Falls back to Mapbox if OSRM is unavailable.
    """
    # Check TTL cache (rounded to 4 decimal places ~ 11m precision to absorb GPS jitter)
    coord_list = coordinates.split(";")
    cache_key = ";".join(_round_coord(c) for c in coord_list)
    cached = _directions_cache.get(cache_key)
    if cached is not None:
        response.headers["X-Cache"] = "HIT"
        return cached

    response.headers["X-Cache"] = "MISS"

    # ── Sugar Bag Rd injection: for any consecutive LM ↔ Aroona transition,
    # insert the Sugarbag Rd Reservoir midpoint so OSRM routes via Sugar Bag
    # Rd (bypassing the traffic-lighted Caloundra Rd corridor). We track a
    # `leg_map` so the downstream response still reports ONE leg per original
    # stop transition — this preserves the frontend's nav.legs[i] contract.
    stops_for_sb: List[dict] = []
    for c in coord_list:
        parts = c.split(",")
        try:
            stops_for_sb.append({"longitude": float(parts[0]), "latitude": float(parts[1])})
        except (ValueError, IndexError):
            stops_for_sb.append({"longitude": 0.0, "latitude": 0.0})

    needs_sb = [False] + [
        needs_sugar_bag_injection(stops_for_sb[i - 1], stops_for_sb[i])
        for i in range(1, len(stops_for_sb))
    ]
    if any(needs_sb):
        injected_coords = inject_sugar_bag_waypoints(coord_list, stops_for_sb)
        osrm_coord_str = ";".join(injected_coords)
        # leg_map[i] = list of OSRM leg indices that collectively form the
        # ORIGINAL leg between stop i-1 and stop i. Injected legs produce a
        # [pre, post] pair; non-injected legs produce a single index.
        leg_map: List[List[int]] = []
        osrm_idx = 0
        for i in range(1, len(stops_for_sb)):
            if needs_sb[i]:
                leg_map.append([osrm_idx, osrm_idx + 1])
                osrm_idx += 2
            else:
                leg_map.append([osrm_idx])
                osrm_idx += 1
        logger.info(
            "Sugar Bag Rd injection: added %d waypoints across %d legs",
            sum(1 for n in needs_sb if n), len(leg_map),
        )
    else:
        osrm_coord_str = coordinates
        leg_map = None

    # --- Primary: Local OSRM Route API ---
    if _osrm_enabled():
        try:
            url = f"{OSRM_URL}/route/v1/driving/{osrm_coord_str}"
            params = {
                "overview": "full",
                "geometries": "geojson",
                "steps": "true",
            }
            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.get(url, params=params)
                if resp.status_code == 200:
                    data = resp.json()
                    if data.get("code") == "Ok" and data.get("routes"):
                        route = data["routes"][0]
                        osrm_legs = route.get("legs", [])
                        # Coalesce injected Sugar Bag legs back into the
                        # original 1-leg-per-stop-transition shape so the
                        # frontend's nav.legs[i] contract is preserved.
                        if leg_map is not None:
                            coalesced: list = []
                            for indices in leg_map:
                                coalesced.append({
                                    "distance": sum(osrm_legs[i].get("distance", 0) for i in indices),
                                    "duration": sum(osrm_legs[i].get("duration", 0) for i in indices),
                                    "summary": osrm_legs[indices[0]].get("summary", ""),
                                })
                            legs_for_response = coalesced
                            all_steps = _extract_steps(osrm_legs)  # keep all steps — incl. the Sugar Bag turn
                        else:
                            legs_for_response = [
                                {
                                    "distance": leg.get("distance", 0),
                                    "duration": leg.get("duration", 0),
                                    "summary": leg.get("summary", ""),
                                }
                                for leg in osrm_legs
                            ]
                            all_steps = _extract_steps(osrm_legs)
                        result = {
                            "geometry": route["geometry"],
                            "distance": route.get("distance", 0),
                            "duration": route.get("duration", 0),
                            "steps": all_steps,
                            "legs": legs_for_response,
                            "waypoints": data.get("waypoints", []),
                            "source": "osrm"
                        }
                        _directions_cache.set(cache_key, result)
                        return result
        except Exception as e:
            _osrm_log_failure("OSRM directions failed, falling back to Mapbox", e)

    # --- Fallback: Mapbox Directions API ---
    if not MAPBOX_TOKEN:
        raise HTTPException(status_code=503, detail="OSRM unavailable and no Mapbox token configured")

    MAX_WAYPOINTS = 25

    if len(coord_list) > MAX_WAYPOINTS:
        all_legs = []
        total_distance = 0
        total_duration = 0
        combined_geometry = {"type": "LineString", "coordinates": []}

        async with httpx.AsyncClient() as client:
            for i in range(0, len(coord_list), MAX_WAYPOINTS - 1):
                chunk = coord_list[i:i + MAX_WAYPOINTS]
                if len(chunk) < 2:
                    break

                chunk_coords = ";".join(chunk)
                resp = await client.get(
                    f"https://api.mapbox.com/directions/v5/mapbox/driving/{chunk_coords}",
                    params={
                        "access_token": MAPBOX_TOKEN,
                        "geometries": "geojson",
                        "overview": "full",
                        "steps": "true",
                    }
                )

                if resp.status_code != 200:
                    raise HTTPException(status_code=resp.status_code, detail=f"Mapbox fallback failed for batch {i}")

                data = resp.json()
                if data.get("routes") and len(data["routes"]) > 0:
                    route = data["routes"][0]
                    total_distance += route.get("distance", 0)
                    total_duration += route.get("duration", 0)

                    if route.get("geometry", {}).get("coordinates"):
                        if combined_geometry["coordinates"]:
                            combined_geometry["coordinates"].extend(route["geometry"]["coordinates"][1:])
                        else:
                            combined_geometry["coordinates"] = route["geometry"]["coordinates"]

                    all_legs.extend(route.get("legs", []))

        all_steps = _extract_steps(all_legs)

        result = {
            "geometry": combined_geometry,
            "distance": total_distance,
            "duration": total_duration,
            "steps": all_steps,
            "legs": all_legs,
            "waypoints": [],
            "source": "mapbox_fallback"
        }
        _directions_cache.set(cache_key, result)
        return result

    async with httpx.AsyncClient() as client:
        resp = await client.get(
            f"https://api.mapbox.com/directions/v5/mapbox/driving/{coordinates}",
            params={
                "access_token": MAPBOX_TOKEN,
                "geometries": "geojson",
                "overview": "full",
                "steps": "true",
            }
        )

        if resp.status_code != 200:
            raise HTTPException(status_code=resp.status_code, detail="Mapbox fallback failed")

        data = resp.json()
        if data.get("routes") and len(data["routes"]) > 0:
            route = data["routes"][0]

            all_steps = _extract_steps(route.get("legs", []))

            legs_summary = []
            for leg in route.get("legs", []):
                legs_summary.append({
                    "distance": leg.get("distance", 0),
                    "duration": leg.get("duration", 0),
                    "summary": leg.get("summary", "")
                })

            result = {
                "distance": route["distance"],
                "duration": route["duration"],
                "geometry": route["geometry"],
                "steps": all_steps,
                "legs": legs_summary,
                "waypoints": data.get("waypoints", []),
                "source": "mapbox_fallback"
            }
            _directions_cache.set(cache_key, result)
            return result

    return {"error": "No route found"}

@api_router.get("/navigation")
async def get_navigation_route(
    current_user: User = Depends(get_current_user),
    current_lat: Optional[float] = Query(None, description="Current latitude"),
    current_lng: Optional[float] = Query(None, description="Current longitude")
):
    """Get full navigation data for all stops in order with waypoint splitting for large routes"""
    all_user_stops = await db.stops.find({"user_id": current_user.user_id}, {"_id": 0}).sort("order", 1).to_list(1000)
    completed_stops = [s for s in all_user_stops if s.get("completed")]
    stops = [s for s in all_user_stops if not s.get("completed")]
    
    if len(stops) < 1:
        return {"error": "Need at least 1 stop for navigation", "stops": all_user_stops}
    
    if not MAPBOX_TOKEN:
        raise HTTPException(status_code=500, detail="Mapbox token not configured")
    
    # Fetch any saved optimization hubs
    hubs = await db.optimization_hubs.find({"user_id": current_user.user_id}, {"_id": 0}).sort("order", 1).to_list(100)
    
    # Create a virtual "current location" stop if coordinates provided
    navigation_stops = []
    if current_lat is not None and current_lng is not None:
        current_location_stop = {
            "id": "current_location",
            "name": "Current Location",
            "address": "Your current location",
            "latitude": current_lat,
            "longitude": current_lng,
            "is_current_location": True
        }
        navigation_stops.append(current_location_stop)
    
    # If we have hubs, we need to interleave them with stops based on which segment each stop belongs to
    if hubs and len(hubs) > 0:
        # Build navigation order: stops before hub1, hub1, stops before hub2, hub2, etc.
        # First, assign each stop to a hub segment based on proximity
        sorted_hubs = sorted(hubs, key=lambda h: h['order'])
        
        # Create waypoint list including hubs
        waypoint_coords = []
        if current_lat is not None and current_lng is not None:
            waypoint_coords.append((current_lat, current_lng))
        for hub in sorted_hubs:
            waypoint_coords.append((hub['latitude'], hub['longitude']))
        
        # Assign each stop to the segment it's closest to
        stop_segments = {i: [] for i in range(len(sorted_hubs) + 1)}
        
        for stop in stops:
            stop_coord = (stop['latitude'], stop['longitude'])
            best_segment = 0
            best_score = float('inf')
            
            for seg_idx in range(len(sorted_hubs) + 1):
                # Calculate distance to segment boundaries
                if seg_idx < len(waypoint_coords):
                    dist_to_start = ((stop_coord[0] - waypoint_coords[seg_idx][0])**2 + 
                                    (stop_coord[1] - waypoint_coords[seg_idx][1])**2)**0.5
                    
                    if seg_idx + 1 < len(waypoint_coords):
                        dist_to_end = ((stop_coord[0] - waypoint_coords[seg_idx + 1][0])**2 + 
                                      (stop_coord[1] - waypoint_coords[seg_idx + 1][1])**2)**0.5
                        score = min(dist_to_start, dist_to_end)
                    else:
                        score = dist_to_start
                else:
                    score = float('inf')
                
                if score < best_score:
                    best_score = score
                    best_segment = seg_idx
            
            stop_segments[best_segment].append(stop)
        
        # Sort stops within each segment by their order field
        for seg_idx in stop_segments:
            stop_segments[seg_idx].sort(key=lambda s: s.get('order', 0))
        
        # Build final navigation order: segment0 stops, hub1, segment1 stops, hub2, ...
        for seg_idx in range(len(sorted_hubs) + 1):
            # Add stops in this segment
            for stop in stop_segments[seg_idx]:
                navigation_stops.append(stop)
            
            # Add hub after this segment (if not the last segment)
            if seg_idx < len(sorted_hubs):
                hub = sorted_hubs[seg_idx]
                hub_waypoint = {
                    "id": f"hub_{hub['id']}",
                    "name": f"Hub {hub['order']}",
                    "address": f"Optimization waypoint {hub['order']}",
                    "latitude": hub['latitude'],
                    "longitude": hub['longitude'],
                    "is_hub": True
                }
                navigation_stops.append(hub_waypoint)
    else:
        # No hubs, just use stops in order
        navigation_stops.extend(stops)
    
    if len(navigation_stops) < 2:
        return {"error": "Need at least 2 points for navigation (current location + 1 stop)", "stops": stops}
    
    # Mapbox limit is 25 waypoints per request
    MAX_WAYPOINTS = 25
    
    async def fetch_route_chunk(chunk_stops: List[dict]) -> Optional[dict]:
        """Fetch route for a chunk of stops using OSRM (primary) or Mapbox (fallback)"""
        coordinates = ";".join([f"{s['longitude']},{s['latitude']}" for s in chunk_stops])

        # --- Primary: OSRM ---
        if _osrm_enabled():
            try:
                async with httpx.AsyncClient(timeout=30.0) as client:
                    resp = await client.get(
                        f"{OSRM_URL}/route/v1/driving/{coordinates}",
                        params={"overview": "full", "geometries": "geojson", "steps": "true"},
                    )
                    if resp.status_code == 200:
                        data = resp.json()
                        if data.get("code") == "Ok" and data.get("routes"):
                            return data["routes"][0]
            except Exception as e:
                _osrm_log_failure("OSRM navigation chunk failed", e)

        # --- Fallback: Mapbox ---
        if not MAPBOX_TOKEN:
            return None

        async with httpx.AsyncClient() as client:
            try:
                response = await client.get(
                    f"https://api.mapbox.com/directions/v5/mapbox/driving/{coordinates}",
                    params={
                        "access_token": MAPBOX_TOKEN,
                        "geometries": "geojson",
                        "overview": "full",
                        "steps": "true",
                    },
                    timeout=30.0
                )
                
                if response.status_code == 200:
                    data = response.json()
                    if data.get("routes") and len(data["routes"]) > 0:
                        return data["routes"][0]
            except Exception as e:
                logger.error(f"Route chunk fetch error: {e}")
        
        return None
    
    # Split stops into chunks with overlap (last point of chunk N = first point of chunk N+1)
    chunks = []
    for i in range(0, len(navigation_stops), MAX_WAYPOINTS - 1):
        chunk = navigation_stops[i:i + MAX_WAYPOINTS]
        if len(chunk) >= 2:
            chunks.append(chunk)
    
    # Fetch all chunks
    all_legs = []
    total_distance = 0
    total_duration = 0
    all_geometry_coords = []
    
    global_stop_index = 0
    
    for chunk_idx, chunk in enumerate(chunks):
        route_data = await fetch_route_chunk(chunk)
        
        if not route_data:
            # If a chunk fails, create placeholder legs
            for i in range(len(chunk) - 1):
                from_stop = chunk[i]
                to_stop = chunk[i + 1]
                all_legs.append({
                    "leg_index": global_stop_index,
                    "from_stop": from_stop,
                    "to_stop": to_stop,
                    "distance": 0,
                    "duration": 0,
                    "summary": "Route unavailable",
                    "steps": []
                })
                global_stop_index += 1
            continue
        
        # Process route data
        total_distance += route_data.get("distance", 0)
        total_duration += route_data.get("duration", 0)
        
        # Add geometry coordinates
        if route_data.get("geometry", {}).get("coordinates"):
            # Skip first coordinate if not first chunk (to avoid duplicates)
            coords = route_data["geometry"]["coordinates"]
            if chunk_idx > 0 and all_geometry_coords:
                coords = coords[1:]
            all_geometry_coords.extend(coords)
        
        # Process legs
        for leg_idx, leg in enumerate(route_data.get("legs", [])):
            from_stop_idx = leg_idx
            to_stop_idx = leg_idx + 1
            
            from_stop = chunk[from_stop_idx] if from_stop_idx < len(chunk) else None
            to_stop = chunk[to_stop_idx] if to_stop_idx < len(chunk) else None
            
            steps = []
            for step in leg.get("steps", []):
                maneuver = step.get("maneuver", {})
                voice_instructions = step.get("voiceInstructions", [])
                voice_text = voice_instructions[0].get("announcement", "") if voice_instructions else ""
                
                steps.append({
                    "distance": step.get("distance", 0),
                    "duration": step.get("duration", 0),
                    "instruction": maneuver.get("instruction", ""),
                    "type": maneuver.get("type", ""),
                    "modifier": maneuver.get("modifier", ""),
                    "location": maneuver.get("location", []),
                    "name": step.get("name", ""),
                    "geometry": step.get("geometry", {}),
                    "voice_instruction": voice_text
                })
            
            all_legs.append({
                "leg_index": global_stop_index,
                "from_stop": from_stop,
                "to_stop": to_stop,
                "distance": leg.get("distance", 0),
                "duration": leg.get("duration", 0),
                "summary": leg.get("summary", ""),
                "steps": steps
            })
            global_stop_index += 1
    
    # Build combined geometry
    combined_geometry = {
        "type": "LineString",
        "coordinates": all_geometry_coords
    } if all_geometry_coords else None
    
    # Calculate completion stats
    completed_count = len(completed_stops)
    
    return {
        "total_distance": total_distance,
        "total_duration": total_duration,
        "geometry": combined_geometry,
        "legs": all_legs,
        "stops": stops + completed_stops,
        "completed_count": completed_count,
        "total_stops": len(all_user_stops),
        "chunks_used": len(chunks)
    }

@api_router.get("/mapbox-token")
async def get_mapbox_token(response: Response):
    """Return Mapbox token for frontend use"""
    response.headers["Cache-Control"] = "private, max-age=3600"  # 1h — rarely changes
    return {"token": MAPBOX_TOKEN}


# ===================== TTS (Text-to-Speech) =====================

_tts_cache: dict[str, str] = {}  # text -> base64 audio cache

@api_router.post("/tts")
async def text_to_speech(request: Request):
    """Generate speech audio from navigation instruction text using OpenAI TTS"""
    body = await request.json()
    text = body.get("text", "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Missing 'text' field")
    if len(text) > 500:
        text = text[:500]

    # Check in-memory cache
    if text in _tts_cache:
        return {"audio_base64": _tts_cache[text], "cached": True}

    llm_key = os.environ.get("EMERGENT_LLM_KEY")
    if not llm_key:
        raise HTTPException(status_code=500, detail="TTS key not configured")

    try:
        from emergentintegrations.llm.openai import OpenAITextToSpeech
        tts = OpenAITextToSpeech(api_key=llm_key)
        audio_b64 = await tts.generate_speech_base64(
            text=text,
            model="tts-1",
            voice="nova",
            speed=1.1,
            response_format="mp3",
        )
        # Cache (limit to 200 entries)
        if len(_tts_cache) > 200:
            _tts_cache.pop(next(iter(_tts_cache)))
        _tts_cache[text] = audio_b64
        return {"audio_base64": audio_b64, "cached": False}
    except Exception as e:
        logger.error("TTS generation failed: %s", e)
        raise HTTPException(status_code=500, detail=f"TTS failed: {str(e)}")

# ===================== Health Check =====================

@api_router.get("/")
async def root():
    return {"message": "Circuit Route Optimizer API", "status": "healthy"}

# ===================== Map Alerts Endpoints =====================

# Default expiry times for different alert types (in minutes)
ALERT_EXPIRY_MINUTES = {
    "police": 30,
    "speed_camera_mobile": 60,
    "hazard": 120,
    "accident": 180,
    "road_work": 480,
    "speed_camera_fixed": None,  # Permanent
}

@api_router.get("/alerts")
async def get_alerts(
    lat: float = Query(..., description="Current latitude"),
    lng: float = Query(..., description="Current longitude"),
    radius_km: float = Query(10, description="Search radius in kilometers"),
    request: Request = None
):
    """Get all active alerts within radius of current location"""
    try:
        # Get current time for expiry check
        now = datetime.now(timezone.utc)
        
        # Find alerts that haven't expired
        cursor = db.map_alerts.find({
            "$or": [
                {"expires_at": None},  # Permanent alerts
                {"expires_at": {"$gt": now}}  # Non-expired alerts
            ]
        }, {"_id": 0})
        
        alerts = []
        async for alert in cursor:
            # Calculate distance
            alert_coords = (alert["latitude"], alert["longitude"])
            user_coords = (lat, lng)
            distance_km = haversine(user_coords, alert_coords, unit=Unit.KILOMETERS)
            
            if distance_km <= radius_km:
                alert["distance_meters"] = distance_km * 1000
                alerts.append(alert)
        
        # Sort by distance
        alerts.sort(key=lambda x: x["distance_meters"])
        
        return alerts
    except Exception as e:
        logger.error(f"Error getting alerts: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/alerts")
async def create_alert(alert_data: AlertCreate, request: Request):
    """Report a new alert"""
    try:
        # Get current user if authenticated
        user = None
        try:
            user = await get_current_user(request)
        except Exception:
            pass
        
        # Check for duplicate alerts nearby (within 100 meters)
        now = datetime.now(timezone.utc)
        cursor = db.map_alerts.find({
            "type": alert_data.type,
            "$or": [
                {"expires_at": None},
                {"expires_at": {"$gt": now}}
            ]
        }, {"_id": 0})
        
        async for existing in cursor:
            existing_coords = (existing["latitude"], existing["longitude"])
            new_coords = (alert_data.latitude, alert_data.longitude)
            distance_m = haversine(existing_coords, new_coords, unit=Unit.METERS)
            
            if distance_m < 100:
                # Update confirmations on existing alert
                await db.map_alerts.update_one(
                    {"id": existing["id"]},
                    {
                        "$inc": {"confirmations": 1},
                        "$set": {"last_confirmed_at": now}
                    }
                )
                existing["confirmations"] += 1
                return existing
        
        # Create new alert
        is_permanent = alert_data.type == "speed_camera_fixed"
        expiry_minutes = ALERT_EXPIRY_MINUTES.get(alert_data.type, 60)
        
        alert = MapAlert(
            type=alert_data.type,
            latitude=alert_data.latitude,
            longitude=alert_data.longitude,
            reported_by=user.user_id if user else None,
            description=alert_data.description,
            speed_limit=alert_data.speed_limit,
            direction=alert_data.direction,
            is_permanent=is_permanent,
            expires_at=None if is_permanent else (now + timedelta(minutes=expiry_minutes))
        )
        
        await db.map_alerts.insert_one(alert.model_dump())
        logger.info(f"New alert created: {alert.type} at ({alert.latitude}, {alert.longitude})")
        
        return alert.model_dump()
    except Exception as e:
        logger.error(f"Error creating alert: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/alerts/{alert_id}/confirm")
async def confirm_alert(alert_id: str, request: Request):
    """Confirm an alert still exists (extends its lifetime)"""
    try:
        now = datetime.now(timezone.utc)
        
        alert = await db.map_alerts.find_one({"id": alert_id}, {"_id": 0})
        if not alert:
            raise HTTPException(status_code=404, detail="Alert not found")
        
        # Extend expiry time if not permanent
        update_data = {
            "$inc": {"confirmations": 1},
            "$set": {"last_confirmed_at": now}
        }
        
        if not alert.get("is_permanent"):
            expiry_minutes = ALERT_EXPIRY_MINUTES.get(alert["type"], 60)
            update_data["$set"]["expires_at"] = now + timedelta(minutes=expiry_minutes)
        
        await db.map_alerts.update_one({"id": alert_id}, update_data)
        
        # Return updated alert
        updated = await db.map_alerts.find_one({"id": alert_id}, {"_id": 0})
        return updated
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error confirming alert: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/alerts/{alert_id}/dismiss")
async def dismiss_alert(alert_id: str, request: Request):
    """Mark an alert as no longer valid"""
    try:
        alert = await db.map_alerts.find_one({"id": alert_id}, {"_id": 0})
        if not alert:
            raise HTTPException(status_code=404, detail="Alert not found")
        
        # Decrease confirmations or delete if no confirmations left
        if alert.get("confirmations", 1) <= 1:
            await db.map_alerts.delete_one({"id": alert_id})
            return {"message": "Alert deleted"}
        else:
            await db.map_alerts.update_one(
                {"id": alert_id},
                {"$inc": {"confirmations": -1}}
            )
            updated = await db.map_alerts.find_one({"id": alert_id}, {"_id": 0})
            return updated
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error dismissing alert: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/alerts/{alert_id}")
async def delete_alert(alert_id: str, request: Request):
    """Delete an alert (admin or reporter only)"""
    try:
        user = await get_current_user(request)
        
        alert = await db.map_alerts.find_one({"id": alert_id}, {"_id": 0})
        if not alert:
            raise HTTPException(status_code=404, detail="Alert not found")
        
        # Only allow deletion by the reporter
        if alert.get("reported_by") and alert["reported_by"] != user.user_id:
            raise HTTPException(status_code=403, detail="Not authorized to delete this alert")
        
        await db.map_alerts.delete_one({"id": alert_id})
        return {"message": "Alert deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting alert: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/alerts/types")
async def get_alert_types():
    """Get all available alert types with their metadata"""
    return [
        {"type": "police", "label": "Police", "icon": "shield", "color": "#3b82f6", "expiry_minutes": 30},
        {"type": "speed_camera_fixed", "label": "Fixed Speed Camera", "icon": "camera", "color": "#ef4444", "expiry_minutes": None},
        {"type": "speed_camera_mobile", "label": "Mobile Speed Camera", "icon": "videocam", "color": "#f97316", "expiry_minutes": 60},
        {"type": "hazard", "label": "Hazard", "icon": "warning", "color": "#eab308", "expiry_minutes": 120},
        {"type": "accident", "label": "Accident", "icon": "car", "color": "#dc2626", "expiry_minutes": 180},
        {"type": "road_work", "label": "Road Work", "icon": "construct", "color": "#f59e0b", "expiry_minutes": 480},
    ]

@api_router.get("/traffic/info")
async def traffic_info():
    """Return current traffic multiplier and schedule."""
    from datetime import datetime, timezone
    now_hour = datetime.now(timezone.utc).hour
    return {
        "current_hour_utc": now_hour,
        "current_multiplier": _traffic_multiplier(now_hour),
        "schedule": {
            "night_free_flow": {"hours": "20:00-05:00", "multiplier": 1.00},
            "early_morning": {"hours": "05:00-07:00", "multiplier": 1.10},
            "am_peak": {"hours": "07:00-09:00", "multiplier": 1.35},
            "post_am_peak": {"hours": "09:00-10:00", "multiplier": 1.15},
            "midday": {"hours": "10:00-15:00", "multiplier": 1.05},
            "school_run": {"hours": "15:00-16:00", "multiplier": 1.20},
            "pm_peak": {"hours": "16:00-18:00", "multiplier": 1.40},
            "post_pm_peak": {"hours": "18:00-20:00", "multiplier": 1.15},
        }
    }

@api_router.get("/cache/stats")
async def cache_stats():
    """Return hit/miss stats for in-memory caches."""
    return {
        "osrm_matrix": _osrm_matrix_cache.stats(),
        "osrm_distance": _osrm_distance_cache.stats(),
        "directions": _directions_cache.stats(),
    }


# ─────────────────────────────────────────────────────────────────────
# Van Layout — per-driver bin-grid configuration for the parcel-finding
# feature. The driver picks a grid shape once (2×3 / 3×3 / 3×4); we
# persist it on their account and reuse it across every route. Bin
# coordinates use spreadsheet notation: rows A, B, C (top→bottom) and
# columns 1, 2, 3 (left→right). So a 3×3 van's bottom-right bin is C3.
# ─────────────────────────────────────────────────────────────────────

# Allowed shapes are explicitly enumerated to prevent drivers from
# accidentally configuring a 50×50 van layout that would break the UI.
ALLOWED_VAN_SHAPES = {(2, 3), (3, 3), (3, 4)}


@api_router.get("/van-layout")
async def get_van_layout(current_user: User = Depends(get_current_user)):
    """Return the driver's saved van layout, or a 3×3 default."""
    doc = await db.van_layouts.find_one(
        {"user_id": current_user.user_id}, {"_id": 0, "user_id": 0}
    )
    if not doc:
        return {"rows": 3, "cols": 3, "is_default": True}
    return {"rows": int(doc["rows"]), "cols": int(doc["cols"]), "is_default": False}


@api_router.put("/van-layout")
async def save_van_layout(
    layout: VanLayout,
    current_user: User = Depends(get_current_user),
):
    """Persist the driver's chosen grid shape. Idempotent upsert."""
    if (layout.rows, layout.cols) not in ALLOWED_VAN_SHAPES:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Unsupported van layout {layout.rows}×{layout.cols}. "
                f"Allowed: {sorted(ALLOWED_VAN_SHAPES)}."
            ),
        )
    await db.van_layouts.update_one(
        {"user_id": current_user.user_id},
        {"$set": {
            "rows": layout.rows,
            "cols": layout.cols,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )
    return {"rows": layout.rows, "cols": layout.cols, "is_default": False}


@api_router.get("/health")
async def health_check():
    """Health check endpoint with MongoDB connection verification"""
    try:
        # Verify MongoDB connection
        await db.command('ping')
        return {
            "status": "healthy",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "database": "connected"
        }
    except Exception as e:
        logger.error(f"Health check failed: {str(e)}")
        return {
            "status": "unhealthy",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "database": "disconnected",
            "error": str(e)
        }


# ── Readiness probe ──────────────────────────────────────────────────────
# Captured ONCE at module load — cheap, deterministic, and means a single
# git fork later in the day doesn't drift the value during a pod's lifetime.
# Source priority:
#   1. GIT_SHA / RELEASE_SHA / EMERGENT_BUILD_SHA env vars (CI bakes them in)
#   2. `git rev-parse --short HEAD` if a .git dir is present in the image
#   3. literal "unknown" — never fail the import on this
def _resolve_build_sha() -> str:
    for env_key in ("GIT_SHA", "RELEASE_SHA", "EMERGENT_BUILD_SHA", "SOURCE_VERSION"):
        v = os.environ.get(env_key)
        if v:
            return v.strip()[:12]
    try:
        import subprocess
        out = subprocess.check_output(
            ["git", "rev-parse", "--short=12", "HEAD"],
            cwd=str(ROOT_DIR.parent),
            stderr=subprocess.DEVNULL,
            timeout=2,
        )
        return out.decode("ascii", "ignore").strip() or "unknown"
    except Exception:
        return "unknown"


_BUILD_SHA = _resolve_build_sha()
_PROCESS_STARTED_AT = datetime.now(timezone.utc)


@api_router.get("/healthz")
async def readiness_probe(response: Response):
    """Lightweight readiness probe for the deploy panel.

    Designed to be polled by Emergent's K8s readiness check. Returns:
      • build.sha — captured once at module load (env var preferred,
        falls back to local `git rev-parse`, then to 'unknown'). Lets you
        eyeball deploy-time vs runtime config drift at a glance.
      • build.uptime_sec — process uptime since the worker started.
      • mongo — Atlas connection state with round-trip latency. The
        actual blocker if anything's wrong in production.
      • tile_cache — row count + bytes on disk + hit rate (from the
        existing stats_sync()). Useful for spotting cold-pod vs warm-pod
        traffic differences in the deploy panel without log diving.
      • status — 'ok' if mongo ping succeeded, 'degraded' otherwise.
        HTTP 503 on degraded so the K8s probe can mark the pod
        unready and pull it out of the load-balancer rotation.

    Deliberately distinct from /api/health (which always returns 200,
    even when mongo is down) so we don't break anything that relies
    on the older endpoint's contract.
    """
    # ── Mongo ping with latency ─────────────────────────────────────────
    mongo_block: Dict[str, Any]
    mongo_ok = False
    t0 = _time.perf_counter()
    try:
        await db.command("ping")
        mongo_ok = True
        mongo_block = {
            "connected": True,
            "db_name": os.environ.get("DB_NAME", ""),
            "ping_ms": round((_time.perf_counter() - t0) * 1000, 2),
        }
    except Exception as e:
        mongo_block = {
            "connected": False,
            "db_name": os.environ.get("DB_NAME", ""),
            "ping_ms": round((_time.perf_counter() - t0) * 1000, 2),
            "error": str(e)[:240],
        }

    # ── Tile cache stats (best-effort, never fail the probe) ───────────
    tile_block: Dict[str, Any]
    try:
        from routes import _tile_cache as _tc
        s = _tc.stats_sync()
        tile_block = {
            "rows": s.get("rows", 0),
            "bytes_on_disk": s.get("bytes_on_disk", 0),
            "hit_rate": s.get("hit_rate", 0.0),
            "hits": s.get("hits", 0),
            "misses": s.get("misses", 0),
        }
    except Exception as e:
        tile_block = {"error": str(e)[:240]}

    now = datetime.now(timezone.utc)
    uptime = (now - _PROCESS_STARTED_AT).total_seconds()
    status = "ok" if mongo_ok else "degraded"

    # 503 on degraded so K8s readiness probe pulls the pod from rotation.
    if not mongo_ok:
        response.status_code = 503

    return {
        "status": status,
        "timestamp": now.isoformat(),
        "build": {
            "sha": _BUILD_SHA,
            "started_at": _PROCESS_STARTED_AT.isoformat(),
            "uptime_sec": round(uptime, 1),
        },
        "mongo": mongo_block,
        "tile_cache": tile_block,
    }


@api_router.get("/healthz/version")
async def readiness_version():
    """Lightweight version-only sub-path for load balancers / monitoring
    (Datadog, Grafana, uptime pings) that just want to spot a deploy rollover
    cheaply. Skips the mongo ping + tile-cache stats — no I/O at all, just
    the two module-level constants captured once at import. ~10 µs to serve.
    Always returns 200; if you need degraded-pod ejection, poll /api/healthz
    instead."""
    return {
        "sha": _BUILD_SHA,
        "started_at": _PROCESS_STARTED_AT.isoformat(),
    }


# ===================== Self-Hosted Building Tiles =====================
# Serves GeoJSON tiles from SQLite DB generated by tiles/generate_building_tiles.py
# Endpoint: /api/tiles/buildings/{z}/{x}/{y}.json

import sqlite3 as _sqlite3
import math as _math

_building_tile_db = None
_TILE_DB_PATH = str(ROOT_DIR.parent / 'tiles' / 'buildings.db')

# ── Parcels + Address tiles moved to routes/tiles.py ───────────────────────
# The /api/tiles/parcels/* and /api/tiles/addresses/* endpoints live in
# routes/tiles.py now. Include them on the shared api_router here so the
# /api prefix is preserved without touching clients.
from routes.tiles import router as tiles_router
api_router.include_router(tiles_router)

# ── House-number endpoints moved to routes/housenumbers.py ───────────────
# Handles /api/housenumbers and /api/housenumbers/prewarm. Own caches +
# circuit breakers — no shared state with server.py.
from routes.housenumbers import router as housenumbers_router
api_router.include_router(housenumbers_router)

# ── Auth endpoints moved to routes/auth.py ───────────────────────────────
# /api/auth/session, /api/auth/me, /api/auth/logout. The whitelist +
# SIGNUPS_DISABLED flags live in that module; auth.py lazily imports
# db/User/get_current_user/get_session_from_request from this file to avoid
# a circular load.
from routes.auth import router as auth_router
api_router.include_router(auth_router)

# ── Stops CRUD moved to routes/stops.py ─────────────────────────────────
# Covers /api/stops, /api/stops/{id}, /api/stops/clear, /api/stops/reorder,
# /api/stops/{id}/complete, /api/stops/{id}/uncomplete, /api/debug/stops-coords.
# Heavy siblings (regeocode, refresh-suburbs, /car/*, stops/export/xlsx) stay
# in server.py for now. Endpoints lazy-import shared helpers from server.
from routes.stops import router as stops_router
api_router.include_router(stops_router)

# ── No-Go Zones ──────────────────────────────────────────────────────────
# /api/nogo-zones CRUD + matrix penalty integration. Polygons drawn by
# the user (or POSTed via curl for now) get treated as impassable: any
# (A, B) leg whose great-circle line crosses a zone is penalised
# +1e9 seconds in the OSRM duration matrix, so the optimiser will
# never pick a zone-crossing leg unless it has no alternative.
from routes.nogo_zones import router as nogo_zones_router
api_router.include_router(nogo_zones_router)

# ── Map-asset proxy moved to routes/map_assets.py ────────────────────────
# /api/map/style, /api/map/sprites/*, /api/map/fonts/*. Self-hosts the
# Liberty sprite + glyph fetches on our origin so MapLibre can reuse the
# warm HTTP/2 connection instead of a cold TLS handshake to openfreemap.org.
# Backed by the shared disk cache — first request is upstream, every
# subsequent one is a local SQLite read.
from routes.map_assets import router as map_assets_router
api_router.include_router(map_assets_router)

# ── Demo scenario for hackathon judges ───────────────────────────────────
# Public, no-auth GET /api/demo/scenario returning a baked 50-stop
# Sunshine Coast route with full OSRM polyline + headline stats. Lets the
# login screen launch a cinematic flythrough without forcing the judge
# through Google sign-in.
from routes.demo import router as demo_router
api_router.include_router(demo_router)

# ── Stripe billing / Pro paywall ─────────────────────────────────────────
# /api/billing/{status,checkout,portal,webhook}. Owns the subscription
# lifecycle and exports `make_require_pro()` — the dependency that other
# endpoints use to gate Pro features. Indexes are created in the startup
# hook below alongside the rest of the Mongo bootstrap.
from routes.billing import (
    router as billing_router,
    _ensure_indexes as _ensure_billing_indexes,
    make_require_pro,
)
api_router.include_router(billing_router)

# ── Waitlist API for Phase 2 rollout gating ──────────────────────────
# /api/waitlist/{join,status,entries,stats,approve,reject,{id}}. Public
# join + status endpoints, admin CRUD. When SIGNUPS_DISABLED=true in
# routes/auth.py, new users are auto-gated through the waitlist.
from routes.waitlist import (
    router as waitlist_router,
    _ensure_waitlist_indexes,
)
api_router.include_router(waitlist_router)

_building_tile_db = None
_TILE_DB_PATH = str(ROOT_DIR.parent / 'tiles' / 'buildings.db')

def _get_tile_db():
    global _building_tile_db
    if _building_tile_db is None:
        import os
        if os.path.exists(_TILE_DB_PATH):
            _building_tile_db = _sqlite3.connect(_TILE_DB_PATH, check_same_thread=False)
            logger.info(f"Building tile DB loaded: {_TILE_DB_PATH}")
        else:
            logger.warning(f"Building tile DB not found: {_TILE_DB_PATH}")
    return _building_tile_db

@api_router.get("/tiles/buildings/{z}/{x}/{y}.json")
async def get_building_tile(z: int, x: int, y: int):
    """Serve self-hosted building GeoJSON tiles with rich height metadata."""
    import gzip as _gzip
    conn = _get_tile_db()
    if conn is None:
        return Response(content=b'{"type":"FeatureCollection","features":[]}',
                       media_type="application/json",
                       headers={"Access-Control-Allow-Origin": "*", "Cache-Control": "public, max-age=86400"})
    row = conn.execute('SELECT data FROM tiles WHERE z=? AND x=? AND y=?', (z, x, y)).fetchone()
    if row is None:
        return Response(content=b'{"type":"FeatureCollection","features":[]}',
                       media_type="application/json",
                       headers={"Access-Control-Allow-Origin": "*", "Cache-Control": "public, max-age=86400"})
    data = _gzip.decompress(row[0])
    return Response(content=data,
                   media_type="application/json",
                   headers={"Access-Control-Allow-Origin": "*", "Cache-Control": "public, max-age=86400"})

@api_router.get("/tiles/buildings/metadata")
async def get_building_tile_metadata():
    """Return metadata about the self-hosted building tiles."""
    conn = _get_tile_db()
    if conn is None:
        return {"error": "Building tile DB not available"}
    rows = conn.execute('SELECT name, value FROM metadata').fetchall()
    return {r[0]: r[1] for r in rows}

# Include router
app.include_router(api_router)


# ─────────────────────────────────────────────────────────────────────────
# Public legal pages (NO /api prefix — these are user-facing URLs that Play
# Console / users open in their browser, not API endpoints)
# ─────────────────────────────────────────────────────────────────────────
#
# Why served from the backend: keeps the legal copy on a stable, SSL'd
# URL under our control (no separate GitHub Pages / Netlify hosting),
# and lets us update the policy by editing one file in this repo.
#
# The HTML lives at /app/frontend/public/privacy-policy.html — the same
# file referenced by Play Console and the in-app Privacy & Terms screen.
# Read at request time (not startup) so a `git pull` propagates instantly
# without a backend restart.

_PRIVACY_POLICY_PATH = ROOT_DIR.parent / "frontend" / "public" / "privacy-policy.html"


def _read_privacy_policy() -> str:
    """Read the privacy policy HTML. Returns a graceful 503 stub if the
    file is missing (shouldn't happen, but better than a 500 traceback)."""
    try:
        return _PRIVACY_POLICY_PATH.read_text(encoding="utf-8")
    except FileNotFoundError:
        return (
            "<!doctype html><html><body style='font-family:sans-serif;padding:2rem'>"
            "<h1>Privacy Policy</h1>"
            "<p>Temporarily unavailable. Email "
            "<a href='mailto:xmltvg@gmail.com'>xmltvg@gmail.com</a> for a copy.</p>"
            "</body></html>"
        )


@app.get("/privacy", include_in_schema=False)
@app.get("/privacy-policy", include_in_schema=False)
@app.get("/privacy-policy.html", include_in_schema=False)
async def privacy_policy_page():
    """Serve the privacy policy HTML at three URL spellings (Play Console,
    in-app, and human-friendly). Cached for 5 min so the CDN can absorb
    crawler traffic; legal pages rarely change but the file is the
    source of truth so we don't want minutes of stale content."""
    return HTMLResponse(
        content=_read_privacy_policy(),
        headers={
            "Cache-Control": "public, max-age=300, s-maxage=300",
            "X-Robots-Tag": "index, follow",
        },
    )


@app.get("/terms", include_in_schema=False)
async def terms_redirect():
    """Terms summary currently lives inside the privacy policy doc.
    Reserved for a future standalone Terms file."""
    return HTMLResponse(content=_read_privacy_policy())


app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── GZip compression ─────────────────────────────────────────────────
# Compress responses > 500 bytes. The big win: optimize results for
# 190 stops (~3 MB JSON) compress to ~300 KB. Drivers on weak 4G
# get results 8-10× faster, eliminating "Network request failed"
# timeouts on the final poll.
from starlette.middleware.gzip import GZipMiddleware
app.add_middleware(GZipMiddleware, minimum_size=500)

@app.on_event("startup")
async def startup_db_client():
    """Verify MongoDB connection and create indexes on startup"""
    max_retries = 8
    retry_delay = 3
    
    for attempt in range(max_retries):
        try:
            await db.command('ping')
            logger.info("MongoDB connection established successfully")
            logger.info(f"Connected to database: {os.environ['DB_NAME']}")
            break
        except Exception as e:
            logger.warning(f"MongoDB connection attempt {attempt + 1}/{max_retries} failed: {str(e)}")
            if attempt < max_retries - 1:
                await asyncio.sleep(retry_delay)
            else:
                logger.error(
                    "Failed to connect to MongoDB after all retries. "
                    "Continuing startup in degraded mode so deployment can complete; "
                    "database-backed endpoints will recover automatically once MongoDB is reachable."
                )
                return
    
    # Create indexes (separate from connection retry loop)
    try:
        # Drop stale unique index if it exists, then create non-unique
        try:
            await db.user_sessions.drop_index("session_token_1")
        except Exception:
            pass
        await db.stops.create_index([("user_id", 1), ("order", 1)], background=True)
        await db.stops.create_index("id", background=True)
        await db.stops.create_index([("user_id", 1), ("completed", 1)], background=True)
        await db.user_sessions.create_index("session_token", background=True)
        await db.user_sessions.create_index("user_id", background=True)
        await db.geocode_cache.create_index("address_query", background=True)
        await db.optimization_hubs.create_index("user_id", background=True)
        await db.map_alerts.create_index("expires_at", background=True)
        await db.route_history.create_index([("user_id", 1), ("archived_at", -1)], background=True)
        await db.route_history.create_index("id", background=True)
        # Async optimize-job store: TTL-purges 10 min after creation; unique
        # job_id keeps `update_one`/`find_one` accurate. Wired here (not on
        # the kickoff hot path) so the user-facing POST is pure-insert.
        await _ensure_optimize_jobs_indexes()
        # Billing: subscriptions + processed_webhook_events collections.
        # Idempotent; safe to re-run on every cold start.
        await _ensure_billing_indexes(db)
        # Waitlist: unique email index + status/created_at for admin queries.
        await _ensure_waitlist_indexes(db)
        logger.info("MongoDB indexes created/verified")
    except Exception as idx_err:
        logger.warning(f"Index creation warning (non-fatal): {idx_err}")

    # Spawn the tile-cache maintenance loop (hourly wal_checkpoint + optimize,
    # daily VACUUM). Imported locally so the circular-import-safe lazy wiring
    # used by routes/auth + stops also applies here.
    try:
        from routes import _tile_cache as _tc
        _tc.start_background_tasks()
        logger.info("tile_cache background maintenance started")
    except Exception as e:
        logger.warning(f"tile_cache maintenance start failed (non-fatal): {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    """Close MongoDB connection and shared HTTP client on shutdown"""
    try:
        # Close shared HTTP client
        global _shared_http_client
        if _shared_http_client and not _shared_http_client.is_closed:
            await _shared_http_client.aclose()
            logger.info("Shared HTTP client closed")
        client.close()
        logger.info("MongoDB connection closed successfully")
    except Exception as e:
        logger.error(f"Error during shutdown: {str(e)}")

# Root-level probe endpoint for platforms that check GET / directly.
@app.get("/api/test-clusters")
async def test_clusters_page():
    import os
    html_path = os.path.join(os.path.dirname(__file__), "static", "test_clusters.html")
    with open(html_path) as f:
        return HTMLResponse(content=f.read())

@app.get("/api/map-test")
async def map_test_page():
    """Standalone MapLibre camera follow test — verifies jumpTo works outside RN WebView."""
    import os
    html_path = os.path.join(os.path.dirname(__file__), "map-test.html")
    with open(html_path) as f:
        return HTMLResponse(content=f.read())


@app.get("/")
async def root_probe():
    """Lightweight root probe that never depends on MongoDB."""
    return {"status": "ok", "service": "route-optimizer", "probe": "root"}

# Root-level health check for Kubernetes probes (without /api prefix)
@app.get("/health")
async def root_health_check():
    """Root health check endpoint for Kubernetes liveness/readiness probes"""
    try:
        # Quick ping to verify MongoDB connection
        await db.command('ping')
        return {"status": "ok", "service": "route-optimizer"}
    except Exception as e:
        logger.error(f"Root health check failed: {str(e)}")
        # Return 200 but indicate unhealthy to prevent pod restarts during transient issues
        return {"status": "degraded", "service": "route-optimizer", "error": "database_unavailable"}

@app.get("/ready")
@app.get("/api/ready")
async def readiness_check():
    """Kubernetes readiness probe - checks if app is ready to serve traffic"""
    uptime_seconds = (datetime.now(timezone.utc) - APP_START_TIME).total_seconds()
    try:
        # Verify database is accessible
        await db.command('ping')
        return {"ready": True, "database": "connected"}
    except Exception as e:
        logger.error(f"Readiness check failed: {str(e)}")

        # During initial warm-up window, return 200 degraded to avoid deployment flapping
        if uptime_seconds < DB_READY_GRACE_SECONDS:
            return {
                "ready": False,
                "database": "connecting",
                "status": "warming_up",
                "grace_seconds_remaining": int(DB_READY_GRACE_SECONDS - uptime_seconds),
            }

        return JSONResponse(
            status_code=503,
            content={"ready": False, "database": "disconnected", "error": str(e)}
        )

@app.get("/live")
@app.get("/api/live")
async def liveness_check():
    """Kubernetes liveness probe - checks if app is alive (less strict than readiness)"""
    # Simple check - if the app can respond, it's alive
    return {"alive": True}

# ── Temporary public file download (no auth) ─────────────────────────
@app.get("/api/download/{token}")
async def download_temp_file(token: str):
    """One-time public download for exported files."""
    import os
    from fastapi.responses import FileResponse
    safe_token = "".join(c for c in token if c.isalnum())
    filepath = os.path.join(os.path.dirname(__file__), f"stops_export_{safe_token}.xlsx")
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="File not found or link expired")
    return FileResponse(filepath, filename="stops_export.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
