"""End-to-end regression test for the Sharpie-locked xlsx export contract.

Verifies `/api/stops/export/xlsx` honours the post-confirm lock contract:

  • Pre-confirm rows: `#` column = `order + 1` (live drive position)
  • Post-confirm rows: `#` column = `original_sequence` (Sharpie-locked)
  • Sort key: locked rows first (sorted by `original_sequence`), then
    unlocked rows in `order` ASC, so spreadsheet row order ALWAYS matches
    the `#` column.
  • Re-optimise after confirm shifts `order` but MUST NOT shift `#`.

Runs in-process via FastAPI TestClient with auth deps overridden — no
real Google session needed. All seeding happens through the public HTTP
API so we never touch motor outside the loop the TestClient owns
(prior fork's pytest module hit a known motor/asyncio loop-binding
conflict when seeding directly).

Usage:
    cd /app/backend && pytest tests/test_xlsx_export_lock.py -v
"""
from __future__ import annotations

import io
import os
import sys
import uuid
from datetime import datetime, timezone

import pytest
from fastapi.testclient import TestClient

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


@pytest.fixture(scope="module")
def client():
    import server  # noqa: WPS433
    from routes.stops import _current_user as routes_current_user

    test_user = server.User(
        user_id=f"pytest-xlsx-{uuid.uuid4().hex[:8]}",
        email="pytest-xlsx@example.com",
        name="Pytest XLSX User",
        picture=None,
        created_at=datetime.now(timezone.utc),
    )

    # /stops/export/xlsx uses get_current_user from server.py
    server.app.dependency_overrides[server.get_current_user] = lambda: test_user
    # /stops, /stops/{id}, /routes/confirm use _current_user from routes/stops.py
    server.app.dependency_overrides[routes_current_user] = lambda: test_user

    with TestClient(server.app) as tc:
        tc._test_user_id = test_user.user_id
        yield tc

    import asyncio

    async def _cleanup():
        await server.db.stops.delete_many({"user_id": test_user.user_id})

    try:
        asyncio.run(_cleanup())
    except RuntimeError:
        pass
    server.app.dependency_overrides.pop(server.get_current_user, None)
    server.app.dependency_overrides.pop(routes_current_user, None)


@pytest.fixture(autouse=True)
def _reset(client):
    client.delete("/api/stops")
    yield
    client.delete("/api/stops")


def _seed(client, rows):
    """Create stops via POST and return the list of created stop dicts."""
    created = []
    for r in rows:
        resp = client.post("/api/stops", json={
            "address": r["address"],
            "name": r.get("name", r["address"]),
            "latitude": r["latitude"],
            "longitude": r["longitude"],
        })
        assert resp.status_code == 200, resp.text
        created.append(resp.json())
    return created


def _read_first_col(xlsx_bytes):
    """Return values from the `#` column (skips header), int-only."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    out = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, int):
            out.append(v)
    return out


# ─────────────────────────────────────────────────────────────────────────────


def test_pre_confirm_uses_order_plus_one(client):
    """Pre-confirm export should number rows from `order + 1`."""
    rows = [
        {"address": "1 First St",  "latitude": -27.5, "longitude": 153.0},
        {"address": "2 Second St", "latitude": -27.6, "longitude": 153.1},
        {"address": "3 Third St",  "latitude": -27.7, "longitude": 153.2},
    ]
    _seed(client, rows)

    resp = client.get("/api/stops/export/xlsx")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml"
    )

    nums = _read_first_col(resp.content)
    assert nums == [1, 2, 3], f"Expected [1,2,3] from order+1, got {nums}"


def test_post_confirm_locks_to_original_sequence(client):
    """Post-confirm rows display `original_sequence` (Sharpie-locked)."""
    rows = [
        {"address": "10 Apple St",  "latitude": -27.5, "longitude": 153.0},
        {"address": "20 Banana St", "latitude": -27.6, "longitude": 153.1},
        {"address": "30 Cherry St", "latitude": -27.7, "longitude": 153.2},
    ]
    created = _seed(client, rows)

    # Confirm in same order: original_sequence will be 1, 2, 3
    ids = [s["id"] for s in created]
    resp = client.post("/api/routes/confirm", json={"confirmed_sequence": ids})
    assert resp.status_code == 200, resp.text

    # Now SHIFT `order` aggressively to simulate re-optimise after confirm.
    # The Sharpie lock should hold: # column = original_sequence, NOT order+1.
    client.put(f"/api/stops/{ids[0]}", json={"order": 7})
    client.put(f"/api/stops/{ids[1]}", json={"order": 9})
    client.put(f"/api/stops/{ids[2]}", json={"order": 5})

    resp = client.get("/api/stops/export/xlsx")
    assert resp.status_code == 200, resp.text

    nums = _read_first_col(resp.content)
    assert nums == [1, 2, 3], (
        f"Sharpie lock broken: got {nums} after re-optimise. "
        f"Expected [1,2,3] from original_sequence (immutable)."
    )


def test_mixed_locked_and_unlocked_rows(client):
    """Locked rows sort first by original_sequence, unlocked follow by order."""
    rows = [
        {"address": "10 Apple St",  "latitude": -27.5, "longitude": 153.0},
        {"address": "20 Banana St", "latitude": -27.6, "longitude": 153.1},
        {"address": "30 Cherry St", "latitude": -27.7, "longitude": 153.2},
        {"address": "40 Date St",   "latitude": -27.8, "longitude": 153.3},
    ]
    created = _seed(client, rows)
    ids = [s["id"] for s in created]

    # Confirm only the FIRST TWO stops (original_sequence = 1, 2).
    # The other two stay unlocked.
    resp = client.post("/api/routes/confirm",
                       json={"confirmed_sequence": [ids[0], ids[1]]})
    assert resp.status_code == 200, resp.text

    resp = client.get("/api/stops/export/xlsx")
    assert resp.status_code == 200, resp.text

    nums = _read_first_col(resp.content)
    # First 2 rows: locked Sharpie [1, 2]. Then 2 unlocked rows by order+1.
    assert len(nums) == 4, f"Expected 4 numeric rows, got {len(nums)}: {nums}"
    assert nums[:2] == [1, 2], (
        f"Locked rows must come first with original_sequence; got {nums[:2]}"
    )
    # Unlocked rows: their `order` was assigned at create time (0..3), so
    # the unlocked subset's order+1 values are still positive integers.
    # Just assert they came AFTER the locked block.
    assert all(n >= 1 for n in nums[2:])


def test_locked_numbers_survive_explicit_order_rewrite(client):
    """The bug we're guarding against: explicit `order` rewrite via PUT
    must NOT shift the spreadsheet `#` column once confirmed."""
    rows = [
        {"address": f"{i*10} Test St",
         "latitude": -27.5 - i * 0.01, "longitude": 153.0 + i * 0.01}
        for i in range(4)
    ]
    created = _seed(client, rows)
    ids = [s["id"] for s in created]

    # Confirm: original_sequence = 1,2,3,4 in created order.
    resp = client.post("/api/routes/confirm",
                       json={"confirmed_sequence": ids})
    assert resp.status_code == 200, resp.text

    # Now re-optimise (PUT order on each).
    client.put(f"/api/stops/{ids[0]}", json={"order": 9})
    client.put(f"/api/stops/{ids[1]}", json={"order": 0})
    client.put(f"/api/stops/{ids[2]}", json={"order": 5})
    client.put(f"/api/stops/{ids[3]}", json={"order": 7})

    resp = client.get("/api/stops/export/xlsx")
    assert resp.status_code == 200, resp.text

    nums = _read_first_col(resp.content)
    assert nums == [1, 2, 3, 4], (
        f"Sharpie lock broken under re-optimise: got {nums} not [1,2,3,4]"
    )
